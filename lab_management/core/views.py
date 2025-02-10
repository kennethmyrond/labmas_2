import openpyxl
import pandas as pd
from openpyxl.styles import Alignment
from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.core.serializers.json import DjangoJSONEncoder
from django.core.exceptions import PermissionDenied, ValidationError
from django.contrib.auth.password_validation import validate_password
from django.contrib.auth import authenticate, login as auth_login, logout, update_session_auth_hash, get_user_model
from django.contrib.auth.hashers import check_password
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models.functions import TruncDate, Coalesce, Greatest, Concat, TruncDay, TruncMonth, TruncYear, Abs
from django.db.models import Q, Sum , Prefetch, F, Count, Avg , CharField, Value,  Case, When, ExpressionWrapper, IntegerField, Max, Min
from django.db import connection, models, DatabaseError, IntegrityError, transaction
from django.utils import timezone
from django.utils.dateformat import DateFormat
from django.utils.timezone import now
from django.http import HttpResponse, JsonResponse, Http404, HttpResponseRedirect
from django.urls import reverse
from django import forms
from django.views.decorators.csrf import csrf_exempt
from collections import defaultdict
from functools import wraps

from allauth.socialaccount.models import SocialAccount

# from allauth.socialaccount.helpers import provider_login_url
from .forms import LoginForm, InventoryItemForm
from .models import RoomTable, laboratory, Module, item_description, item_types, item_inventory, suppliers, user, suppliers, item_expirations, item_handling
from .models import borrow_info, borrowed_items, borrowing_config, reported_items
from .models import rooms, laboratory_reservations, reservation_config
from .models import laboratory_users, laboratory_roles, laboratory_permissions, permissions, Notification
from .models import WorkInProgress
from .decorators import lab_permission_required, superuser_or_lab_permission_required
from datetime import timedelta, date, datetime
from calendar import monthrange
from pyzbar.pyzbar import decode
from PIL import Image, ImageDraw, ImageFont 
from io import BytesIO
# from ratelimit.decorators import ratelimit 


# python image library pillow

import json, qrcode, base64, threading, time, re

import logging

sec_logger = logging.getLogger('django.security')

logger = logging.getLogger('custom_logger')

import pint

def get_unit_choices():
    ureg = pint.UnitRegistry()

    unit_categories = {
        "Capacity/Volume": ["milliliter", "liter"],
        "Mass": ["gram", "kilogram", "milligram", "pound", "ounce"],
        "Length": ["millimeter", "centimeter", "meter", "inch", "foot", "yard"],
        # "Pieces": ["set", "unit"]
    }

    unit_choices = {}
    for category, units in unit_categories.items():
        choices = []
        for unit in units:
            try:
                unit_obj = ureg(unit)
                symbol = f"{unit_obj.units:~P}"  # Correctly get unit symbol
                choices.append((unit_obj.units, f"{unit_obj.units} ({symbol})"))
            except pint.errors.UndefinedUnitError:
                choices.append((unit, unit))  # Fallback if no unit symbol
        
        unit_choices[category] = choices

    return unit_choices


prev_day = ''
# thread every midnight query items to check due date if today (for on holding past due date )
@transaction.atomic
def late_borrow(request):
    global prev_day
    prev_day = timezone.localtime().day  # Start with the current day
    output = 'None'
    # Get current date and time
    current_datetime = timezone.localtime()
    current_day = current_datetime.day

    # Check if the day has changed (execute once every new day)
    if current_day != prev_day:
    # if 1==1 :
        # Update `prev_day` to track the day change
        prev_day = current_day

        # 1. Cancel borrows if their `borrow_date` is past today and status is still 'Accepted' ('A')
        expired_borrows = borrow_info.objects.filter(
            status='A',  # Accepted borrows
            borrow_date__lt=current_datetime.date()  # Past borrow dates
        )
        for borrow in expired_borrows:
            borrow.status = 'L'  # Cancel the borrow
            borrow.remarks = "Automatically cancelled due to expired borrow date."
            borrow.save()

        # 2. Handle borrowed items with past due dates
        overdue_borrows = borrow_info.objects.filter(
            status='B',  # Borrowed status
            due_date__lt=current_datetime.date()  # Past due dates
        )
        for borrow in overdue_borrows:
            # Get the borrowed items
            borrowed_items_list = borrowed_items.objects.filter(borrow=borrow)

            for item in borrowed_items_list:
                # Check if the item hasn't been fully returned
                if item.qty > item.returned_qty:
                    # Create a reported item for overdue borrowed items
                    reported_item = reported_items.objects.create(
                        borrow=borrow,
                        item=item.item,
                        qty_reported=item.qty - item.returned_qty,
                        report_reason="Overdue item",
                        amount_to_pay=99999,  # You can set an amount if needed
                        status=1  # Pending
                    )
                    reported_item.save()

            # Put borrower's clearance on hold
            borrow.status = 'B'  # Mark borrow as 'Cancelled/On Hold'
            borrow.remarks = "Automatically placed on hold due to overdue items."
            borrow.save()
        output='late changed'
    data = {
        'value': output
    }

    return JsonResponse(data)

# Start the thread for daily checking
# x = threading.Thread(target=thread_function)
# x.start()

# functions
def get_inventory_history(item_description_instance):
    inventory_items = item_inventory.objects.filter(item=item_description_instance)
    history = item_handling.objects.filter(inventory_item__in=inventory_items).order_by('-timestamp')
    return history

def check_item_expiration(request, item_id):
    # Get the item_description instance by item_id
    print(item_id)
    item = get_object_or_404(item_description, item_id=item_id)
    # Return JSON response with rec_expiration value
    return JsonResponse({
        'rec_expiration': item.expiry_type,
        'rec_per_inv': item.rec_per_inv
    })

def suggest_items(request):
    query = request.GET.get('query', '')
    selected_laboratory_id = request.session.get('selected_lab')
    
    # Fetch suggestions from the database
    # suggestions = item_description.objects.filter(
    #     item_name__icontains=query, 
    #     laboratory_id=selected_laboratory_id, 
    #     is_disabled=0
    # )[:5]

    suggestions = item_description.objects.filter(
        laboratory_id=selected_laboratory_id, 
        is_disabled=0
    ).filter(
        Q(item_name__icontains=query) | Q(item_id=query)
    )[:5]

    data = []
    for item in suggestions:
        # Parse the add_cols JSON string into a dictionary
        try:
            add_cols = json.loads(item.add_cols) if item.add_cols else {}
        except json.JSONDecodeError:
            add_cols = {}

        # Convert add_cols dictionary to a formatted string for display
        add_cols_str = ', '.join([f"{key}: {value}" for key, value in add_cols.items()])
        
        # Prepare the response data
        item_data = {
            'item_id': item.item_id,
            'item_name': item.item_name,
            'rec_expiration': item.expiry_type,
            'add_cols': add_cols_str  # Send formatted string
        }
        
        data.append(item_data)
    return JsonResponse(data, safe=False)

def suggest_inventory_items(request, item_id):
    # Fetch inventory items associated with the given item_id that have quantity > 0
    item = get_object_or_404(item_description, item_id=item_id)
    inventory_items = item_inventory.objects.filter(item__item_id=item_id).exclude(qty=0)

    data = []
    for inventory_item in inventory_items:
        # Get expiration date if it exists, else set to None
        expiration = item_expirations.objects.filter(inventory_item=inventory_item).first()
        if item.expiry_type == 'Date':
            expiration_data = DateFormat(expiration.expired_date).format('Y-m-d') if expiration else "None"
        elif item.expiry_type == 'Usage':
            expiration_data = expiration.remaining_uses if expiration else "None"
        elif item.expiry_type == 'Maintenance':
            expiration_data = expiration.next_maintenance_date if expiration else "None"
        else:
            expiration_data = "N/A"
        
        # Format item data for response
        data.append({
            'inventory_item_id': inventory_item.inventory_item_id,
            'expiry_type': item.expiry_type,
            'expiration_date': expiration_data,
            'qty': inventory_item.qty
        })

    return JsonResponse(data, safe=False)

def suggest_suppliers(request):
    if not request.user.is_authenticated:
        return redirect('userlogin')
    
    selected_laboratory_id = request.session.get('selected_lab')
    query = request.GET.get('query', '')
    supplier_suggestions = suppliers.objects.filter(supplier_name__icontains=query, laboratory=selected_laboratory_id, is_disabled=0)[:5]  # Limit the results to 5

    data = []
    for supplier in supplier_suggestions:
        data.append({
            'suppliers_id': supplier.suppliers_id,
            'supplier_name': supplier.supplier_name
        })

    return JsonResponse(data, safe=False)

def suggest_users(request):
    print("users:")
    query = request.GET.get('query', '')
    lab_id = request.GET.get('lab_id', None)

    # Filter to get users not already assigned to the lab
    assigned_user_ids = laboratory_users.objects.filter(laboratory_id=lab_id, status='A', is_active=1).values_list('user_id', flat=True)
    users = user.objects.exclude(Q(user_id__in=assigned_user_ids) | Q(is_superuser=1)).filter(
        Q(username__icontains=query) | Q(firstname__icontains=query) | Q(lastname__icontains=query)
    ).annotate(fullname=Concat(F('email'), Value(' | '), F('firstname'), Value(' '), F('lastname'), output_field=CharField()))

    results = [{'user_id': u.user_id, 'fullname': u.fullname} for u in users]
    
    return JsonResponse(results, safe=False)

def suggest_report_users(request):
    print("users:")
    query = request.GET.get('query', '')
    lab_id = request.session.get('selected_lab')
    
    users = laboratory_users.objects.filter(
        Q(user__username__icontains=query) | Q(user__firstname__icontains=query) | Q(user__lastname__icontains=query),
        laboratory_id=lab_id, 
        status='A', 
        is_active=True  # Ensure this matches the field type
    ).annotate(
        fullname=Concat( F('user__personal_id'), Value(' | '), F('user__firstname'), Value(' '), F('user__lastname'), Value(' | '),F('user__email') ,
        output_field=CharField())
    )

    results = [{'user_id': u.user.user_id, 'fullname': u.fullname} for u in users]
    
    return JsonResponse(results, safe=False)

@transaction.atomic
@login_required
def remove_item_from_inventory(inventory_item, amount, user, change_type, remarks=''):
    if inventory_item.qty < amount:
        raise ValueError("Not enough items in inventory to remove.")

    inventory_item.qty -= amount
    inventory_item.save()

    # Record the item handling
    item_handling.objects.create(
        inventory_item=inventory_item,
        updated_by=user,
        changes=change_type,
        qty=0-amount,
        remarks=remarks,
    )

@transaction.atomic
def generate_qr_code(item_QRize, details):

    # Create a QR code instance
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(item_QRize)
    qr.make(fit=True)

    # Generate the QR code image
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")

    # Add text below the QR code
    text = details
    font_size = 20
    try:
        # Load a default font; replace "arial.ttf" with the path to your font file if needed
        font = ImageFont.truetype("arial.ttf", font_size)
    except IOError:
        # Fallback to default PIL font if the specified font is not found
        font = ImageFont.load_default()

    # Split the text into multiple lines
    lines = text.split("\n")
    line_heights = [font.getbbox(line)[3] - font.getbbox(line)[1] for line in lines]
    total_text_height = sum(line_heights) + (len(lines) - 1) * 5  # 5px spacing between lines

    # Create a new image with space for the QR code and text
    img_width, img_height = qr_img.size
    new_height = img_height + total_text_height + 10  # Add space for text and margins
    new_img = Image.new("RGBA", (img_width, new_height), "white")
    new_img.paste(qr_img, (0, 0))

    # Draw text on the new image
    draw = ImageDraw.Draw(new_img)
    current_y = img_height + 5  # Start drawing below the QR code
    for line in lines:
        text_width = font.getbbox(line)[2] - font.getbbox(line)[0]
        text_x = (img_width - text_width) // 2  # Center text horizontally
        draw.text((text_x, current_y), line, fill="black", font=font)
        current_y += font.getbbox(line)[3] - font.getbbox(line)[1] + 5  # Move to the next line

    # Save the QR code with text as a base64 image
    buffered = BytesIO()
    new_img.save(buffered, format="PNG")
    qr_base64 = base64.b64encode(buffered.getvalue()).decode()
    qr_code = f"data:image/png;base64,{qr_base64}"

    return qr_code  # Return the base64 encoded string

def create_notification(user, message):
    try:
        Notification.objects.create(user=user, message=message)
        logger.info(f"Notification created for {user}: {message}")
    except Exception as e:
        logger.error(f"Failed to create notification for {user}: {e}", exc_info=True)
    
def get_notifications(request):
    try:
        return Notification.objects.filter(user=request.user, is_read=False)
    except Exception as e:
        logger.error(f"Error fetching notifications for {request.user}: {e}", exc_info=True)
        return Notification.objects.none()


def mark_all_notifications_read(request):
    # Mark all notifications for the current user as read
    notifications = Notification.objects.filter(user=request.user, is_read=False)
    notifications.update(is_read=True)  # Update all notifications to read

    # Optionally, add a success message
    messages.success(request, "All notifications have been marked as read.")

    # Redirect back to the referer (previous page the user was on)
    referer = request.META.get('HTTP_REFERER', 'default_page')  # Default to 'default_page' if no referer is found
    return HttpResponseRedirect(referer)

# forms
class ItemEditForm(forms.ModelForm):
    itemType = forms.ModelChoiceField(
        queryset=item_types.objects.none(),  # Initialize with an empty queryset
        empty_label="Select Item Type",
        required=False,
        label="Item Type",
        widget=forms.Select(attrs={'class': 'form-control'})
    )

    class Meta:
        model = item_description
        fields = ['item_name', 'itemType']

    def __init__(self, *args, **kwargs):
        # Capture selected laboratory ID passed via kwargs
        selected_laboratory_id = kwargs.pop('selected_laboratory_id', None)
        super(ItemEditForm, self).__init__(*args, **kwargs)
        
        # Filter itemType based on the selected laboratory
        if selected_laboratory_id:
            self.fields['itemType'].queryset = item_types.objects.filter(laboratory_id=selected_laboratory_id)


# views

# misc views
User = get_user_model()

def register(request):
    if request.method == "POST":
        firstname = request.POST.get('firstname')
        lastname = request.POST.get('lastname')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        
        # Check if email exists
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists.")
            return render(request, "register.html")
        
        # Validate password policy
        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
        else:
            try:
                validate_password(password)
            except ValidationError as e:
                messages.error(request, e)
            else:
                # Create user
                user = User.objects.create_user(
                    email=email, 
                    firstname=firstname, 
                    lastname=lastname, 
                    password=password
                )
                messages.success(request, "Account created successfully. Please log in.")
                return redirect("userlogin")

    return render(request, "register.html")

def custom_login(request):
    if 'error' in request.GET and request.GET['error'] == 'access_denied':
        messages.error(request, "You cancelled the login process. Please try again.")
        return redirect(reverse('userlogin'))
    return redirect(reverse('userlogin'))

# @ratelimit(key='post:email', rate='5/m', block=True)
def userlogin(request):
    # try:
    if request.method == "POST":
        email = request.POST.get('email')
        password = request.POST.get('password')

        # Authenticate using email and password
        user = authenticate(request, email=email, password=password)
        if user is not None:
            # Check if the user is deactivated
            if user.is_deactivated == 1:  # Assuming is_deactivated is an integer field
                messages.error(request, "Invalid email or password")
            else:
                auth_login(request, user)
                return redirect('home')  # Redirect to a specific page after login
        else:
            messages.error(request, "Invalid email or password")
    
    # Check if the user came back from Google OAuth
    if request.user.is_authenticated and SocialAccount.objects.filter(user=request.user).exists():
        return redirect('home')  # Redirect to home if authenticated via Google
    
    # Render the login template
    return render(request, "user_login.html")
    # except Ratelimited:
    #     messages.error(request, "Too many login attempts. Please try again later.")
    #     return redirect('login')  # Redirect to the login page

@login_required(login_url='/login')
def home(request):
    pending_labs = laboratory.objects.filter(
            is_available=True, 
            laboratory_users__user=request.user, 
            laboratory_users__is_active=True,
            laboratory_users__status='P'
        )
    
    context = {
        'user': request.user, 
        'pending_labs':pending_labs
    }

    if request.user.is_superuser:
        return redirect('setup_createlab')
    else:
        return render(request, "home.html", context)

@login_required
def request_laboratory(request):
    if request.method == "POST":
        lab_id = request.POST.get("laboratory_id")
        role_id = request.POST.get("role_id")
        
        # Check if laboratory ID is valid
        try:
            lab = laboratory.objects.get(laboratory_id=lab_id)
        except laboratory.DoesNotExist:
            messages.error(request, "Invalid laboratory ID.")
            return redirect("home")

        # Check if role ID is valid
        try:
            role = laboratory_roles.objects.get(roles_id=role_id)
        except laboratory_roles.DoesNotExist:
            messages.error(request, "Invalid role ID.")
            return redirect("home")

        # Check if user is already added to this laboratory
        if laboratory_users.objects.filter(user=request.user, laboratory=lab, status__in=['A', 'P', 'I'], is_active=True).exists():
            messages.error(request, "You are already registered in this laboratory.")
            return redirect("home")

        # Create a new laboratory user with pending status
        laboratory_users.objects.create(
            user=request.user,
            laboratory=lab,
            role=role,
            status='P'  # Set to Pending
        )
        messages.success(request, "Laboratory access request submitted successfully.")
        return redirect("home")
    return redirect("home")

@login_required
def set_lab(request, laboratory_id):
    # Get the chosen laboratory
    try:
        lab = get_object_or_404(laboratory, laboratory_id=laboratory_id)
        
        # Check if the laboratory is available
        if lab.is_available == 1 and laboratory_users.objects.filter(user=request.user, laboratory=lab, is_active=True, status='A').exists():
                # Set the chosen laboratory in the session
                request.session['selected_lab'] = lab.laboratory_id
                request.session['selected_lab_name'] = lab.name
        return redirect('home')  # Redirect back to the previous page
    except:
        return redirect('home')


@login_required
def logout_view(request):
    logout(request)
    return redirect("/login")

def error_page(request, message=None):
    """
    Renders a generic error page.
    :param request: HTTP request
    :param message: The error message to display (optional)
    """
    message = request.GET.get('message', 'An error occurred.')
    return render(request, 'error_page.html', {'message': message})

def custom_404_view(request, exception):
    return render(request, 'error_page.html', status=404)

@login_required
def my_profile(request):
    user = request.user
    return render(request, 'my_profile.html', {'user': user})

@login_required
@transaction.atomic
def deactivate_account(request):
    user = request.user
    user.is_deactivated = True
    user.save()
    messages.success(request, "Account deactivated successfully.")
    return redirect('logout')  # Redirect to logout or another appropriate page

@login_required
@transaction.atomic
def edit_profile(request):
    user = request.user
    if request.method == 'POST':
        firstname = request.POST.get('firstname')
        lastname = request.POST.get('lastname')
        username = request.POST.get('username')
        email = request.POST.get('email')
        personal_id = request.POST.get('personal_id')
        
        # Check if the email or username is already taken by another user
        if User.objects.filter(email=email).exclude(pk=user.pk).exists():
            messages.error(request, "Email is already registered by another user.")
        elif User.objects.filter(username=username).exclude(pk=user.pk).exists():
            messages.error(request, "Username is already taken by another user.")
        else:
            user.firstname = firstname
            user.lastname = lastname
            user.username = username
            user.email = email
            user.personal_id = personal_id
            user.save()
            messages.success(request, "Profile updated successfully.")
            return redirect('my_profile')
    
    return render(request, 'edit_profile.html', {'user': user})

def signup_redirect(request):
    messages.error(request, "Something wrong here, it may be that you already have account!")
    return redirect("homepage")


# inventory
@login_required
@lab_permission_required('view_inventory')
def inventory_view(request):
    if not request.user.is_authenticated:
        logger.warning(f"Unauthorized attempt to access inventory_view by {request.user}")
        return redirect('userlogin')
    
    selected_laboratory_id = request.session.get('selected_lab')  # Get the selected laboratory from the session
    item_types_list = item_types.objects.filter(laboratory_id=selected_laboratory_id)  # Get all item types for the selected laboratory
    selected_item_type = request.GET.get('item_type')  # Get the selected item_type from the GET parameters
    current_date = timezone.localtime().date()  # Fetch current date for expiration comparison

    # Fetch inventory items based on item_type and lab
    try:
        if selected_item_type:
            if selected_item_type == '0':
                inventory_items = item_description.objects.filter(
                    laboratory_id=selected_laboratory_id,
                    is_disabled=0,
                    itemType_id=None  # Only get items that are enabled
                ).annotate(total_qty=Coalesce(Sum('item_inventory__qty'), 0))  # Calculate total quantity
                add_cols = []
            else:
                inventory_items = item_description.objects.filter(
                    itemType_id=selected_item_type,
                    laboratory_id=selected_laboratory_id,
                    is_disabled=0  # Only get items that are enabled
                ).annotate(total_qty=Coalesce(Sum('item_inventory__qty'), 0))  # Calculate total quantity
                selected_item_type_instance = item_types.objects.get(pk=selected_item_type)
                add_cols = json.loads(selected_item_type_instance.add_cols)
        else:
            inventory_items = item_description.objects.filter(
                laboratory_id=selected_laboratory_id,
                is_disabled=0  # Only get items that are enabled
            ).annotate(total_qty=Coalesce(Sum('item_inventory__qty'), 0))  # Calculate total quantity
            add_cols = []
        
        # Log inventory count
        logger.debug(f"Inventory fetched: {inventory_items.count()} items found.")
        notifications_sent = False  # Flag to prevent duplicate notifications
        notified_items = set()  # To track which items have been notified

        # Loop through each inventory item and check conditions
        for item in inventory_items:
            expirations = item_expirations.objects.filter(inventory_item__item=item)
            expiration_warnings = ''

            # Check expiration conditions
            for exp in expirations:
                # Date Expiry
                if item.expiry_type == 'Date' and exp.expired_date <= current_date + timedelta(days=7):
                    if exp.inventory_item.qty > 0 and item.item_id not in notified_items:
                        expiration_warnings = 'D'
                        message = f"Item '{item.item_name}' will expire soon!"
                        existing_notification = Notification.objects.filter(
                            user=request.user,
                            message=message
                        ).first()
                        if not existing_notification:
                            create_notification(request.user, message)
                            notified_items.add(item.item_id)  # Mark as notified

                # Usage Limit Reached
                elif item.expiry_type == 'Usage' and exp.remaining_uses <= 0:
                    if item.item_id not in notified_items:
                        expiration_warnings = 'U'
                        message = f"Item '{item.item_name}' usage limit reached!"
                        existing_notification = Notification.objects.filter(
                            user=request.user,
                            message=message
                        ).first()
                        if not existing_notification:
                            create_notification(request.user, message)
                            notified_items.add(item.item_id)  # Mark as notified

                # Maintenance Due
                elif item.expiry_type == 'Maintenance' and exp.next_maintenance_date <= current_date + timedelta(days=7):
                    if item.item_id not in notified_items:
                        expiration_warnings = 'M'
                        message = f"Item '{item.item_name}' needs maintenance soon!"
                        existing_notification = Notification.objects.filter(
                            user=request.user,
                            message=message
                        ).first()
                        if not existing_notification:
                            create_notification(request.user, message)
                            notified_items.add(item.item_id)  # Mark as notified

            # Check for low stock (e.g., less than 5 items remaining)
            if item.total_qty < 5 and item.item_id not in notified_items:
                message = f"Item '{item.item_name}' is running low on stock!"
                existing_notification = Notification.objects.filter(
                    user=request.user,
                    message=message
                ).first()
                if not existing_notification:
                    create_notification(request.user, message)
                    notified_items.add(item.item_id)  # Mark as notified

            item.expiration_warning = expiration_warnings  # Set expiration warning

        # Send notifications for any issue detected
        if notifications_sent:
            logger.info("Notifications have been sent for inventory items.")

        # Get unread notifications for the user
        notifications = get_notifications(request)

        return render(request, 'mod_inventory/view_inventory.html', {
            'inventory_items': inventory_items,
            'item_types': item_types_list,
            'selected_item_type': int(selected_item_type) if selected_item_type else None,
            'add_cols': add_cols,
            'notifications': notifications,  # Pass notifications to the template
        })
    except Exception as e:
        logger.error(f"Error fetching inventory: {e}", exc_info=True)
        messages.error(request, "An error occurred while fetching inventory.")
        return redirect('home')


@transaction.atomic
@login_required
@lab_permission_required('view_inventory')
def update_maintenance(request, inventory_item_id):
    try:
        inventory_item = get_object_or_404(item_inventory, inventory_item_id=inventory_item_id)
        expiration_entry = get_object_or_404(item_expirations, inventory_item=inventory_item)
        
        maintenance_interval = inventory_item.item.maintenance_interval or 1  # Default 1 year if not set
        next_maintenance_date = now().date().replace(year=now().year + maintenance_interval)
        
        expiration_entry.next_maintenance_date = next_maintenance_date
        expiration_entry.save()
        
        messages.success(request, f"Maintenance updated for {inventory_item_id}! Next maintenance on {next_maintenance_date.strftime('%Y-%m-%d')}")
    except Exception as e:
        logger.error(f"Error fetching inventory: {e}", exc_info=True)
        messages.error(request, "An error occurred while updating maintenance date.")

    return redirect('inventory_itemDetails_view', item_id=inventory_item.item.item_id)


@login_required
@lab_permission_required('view_inventory')
def inventory_itemDetails_view(request, item_id):
    if not request.user.is_authenticated:
        logger.warning(f"Unauthorized access attempt to inventory_itemDetails_view by {request.user}")
        return redirect('userlogin')
    
    selected_laboratory_id = request.session.get('selected_lab')

    logger.info(f"User {request.user.email} is viewing details for item ID: {item_id} in lab {selected_laboratory_id}")

    try:
        item = get_object_or_404(item_description, item_id=item_id) # Get the item_description instance
        add_cols_data = json.loads(item.add_cols) if item.add_cols else {} # Parse add_cols JSON
        item_inventories = item_inventory.objects.filter(item=item).select_related('supplier')
        
        item_type = item.itemType# Get the related itemType instance
        date_today = date.today()
        lab = get_object_or_404(laboratory, laboratory_id=selected_laboratory_id)


        # Gather expiration data for each inventory item if item records expiration
        expiration_data = {}
        if item.expiry_type!=None:
            expirations = item_expirations.objects.filter(inventory_item__in=item_inventories)
            if item.expiry_type=='Date':
                expiration_data = {exp.inventory_item.inventory_item_id: exp.expired_date for exp in expirations}
            elif item.expiry_type=='Usage':
                expiration_data = {exp.inventory_item.inventory_item_id: exp.remaining_uses for exp in expirations}
            elif item.expiry_type=='Date':
                expiration_data = {exp.inventory_item.inventory_item_id: exp.next_maintenance_date for exp in expirations}


        # inventory items
        # Initialize the inventory data list
        inventory_data = []

        # Gather data for each inventory item
        for inventory in item_inventories:
            # Get the latest handling record for each inventory item
            last_handling = inventory.item_handling_set.order_by('-timestamp').first()
            first_handling = inventory.item_handling_set.order_by('timestamp').first()
                
            expiration_date = expiration_data.get(inventory.inventory_item_id, 'None') if item.expiry_type!=None else 'N/A'
            
            last_updated_date = last_handling.timestamp if last_handling else 'N/A'
            last_updated_by = last_handling.updated_by if last_handling and last_handling.updated_by else 'N/A'
            date_created = first_handling.timestamp if first_handling else 'N/A'

            # Append data to the inventory_data list
            inventory_data.append({
                'inventory_item_id': inventory.inventory_item_id,
                'current_qty': inventory.qty,
                'expiration_date': expiration_date,
                'last_updated_date': last_updated_date,
                'last_updated_by': last_updated_by,
                'date_created': date_created
            })

        # Prefetch item_handling related to item_inventory, ordered by timestamp in descending order
        item_handling_prefetch = Prefetch('item_handling_set', queryset=item_handling.objects.all().order_by('-timestamp'))

        item_inventories = item_inventory.objects.filter(item=item)\
            .select_related('supplier')\
            .prefetch_related(item_handling_prefetch)\
            .annotate(latest_handling_timestamp=Max('item_handling__timestamp'))\
            .order_by('-latest_handling_timestamp')  # Use the annotated field for ordering

        # print(item_inventories)

        # Calculate the total quantity
        total_qty = item_inventories.aggregate(Sum('qty'))['qty__sum'] or 0

        # Generate QR codes
        qr_data = f"{item.item_id}"
        qr_details = f"{item.item_name}"
        qr_code_data_item_only = generate_qr_code(qr_data, qr_details)

        if item.expiry_type != None or item.rec_per_inv == True:
            for item_inv in item_inventories:
                qr_data = f"{item_inv.item.item_id}, {item_inv.inventory_item_id}"
                qr_details = f"{item.item_name}\nInvID: {item_inv.inventory_item_id}"
                qr_code_data = generate_qr_code(qr_data, qr_details)
                item_inv.qr_code = qr_code_data

        # Fetch expiration data and attach it to each inventory item if applicable
        if item.expiry_type != None :
            expirations = item_expirations.objects.filter(inventory_item__in=item_inventories)
            # expiration_data = {exp.inventory_item.inventory_item_id: exp.expired_date for exp in expirations}
            if item.expiry_type=='Date':
                expiration_data = {exp.inventory_item.inventory_item_id: exp.expired_date for exp in expirations}
            elif item.expiry_type=='Usage':
                expiration_data = {exp.inventory_item.inventory_item_id: exp.remaining_uses for exp in expirations}
            elif item.expiry_type=='Maintenance':
                expiration_data = {exp.inventory_item.inventory_item_id: exp.next_maintenance_date for exp in expirations}

            # Attach expiration_date directly to each inventory instance
            for inventory in item_inventories:
                latest_handling = inventory.item_handling_set.first()  # Access related item_handling
                inventory.expiration_date = expiration_data.get(inventory.inventory_item_id)
        else:
            # If no expirations are recorded, set expiration_date to None for consistency
            for inventory in item_inventories:
                latest_handling = inventory.item_handling_set.first()  # Access related item_handling
                inventory.expiration_date = None

        # suppliers table
        supplier_data = []
        for inventory in item_inventories:
            if inventory.supplier_id != None:
                first_handling = (
                    item_handling.objects.filter(inventory_item=inventory)
                    .order_by('timestamp')
                    .first()
                )
                
                date_purchased = inventory.date_purchased
                date_received = inventory.date_received
                duration = (date_received - date_purchased).days if date_purchased and date_received else None
                expiration_date = expiration_data.get(inventory.inventory_item_id, 'None')

                supplier_data.append({
                    'inventory_id': inventory.inventory_item_id,
                    'supplier_name': inventory.supplier.supplier_name if inventory.supplier else 'N/A',
                    'date_purchased': date_purchased,
                    'date_received': date_received,
                    'duration': f"{duration} days" if duration else '0',
                    'qty': first_handling.qty if first_handling else 'Invalid',
                    'purchase_price': inventory.purchase_price,
                    'expiration': expiration_date
                })
        
        logger.debug(f"Fetched {len(item_inventories)} inventory items for item {item_id}.")

        context = {
            'item': item,
            'itemType_name': item_type.itemType_name if item_type else None,
            'laboratory_name': lab.name if lab else None,
            'item_inventories': item_inventories,
            'total_qty': total_qty,
            'add_cols_data': add_cols_data,
            'is_edit_mode': False,  # Not in edit mode
            'qr_code_data': qr_code_data_item_only,
            'date_today': date_today,

            'lab_name': lab.name,
            'supplier_data': supplier_data,

            'inventory_data': inventory_data,
        }

        return render(request, 'mod_inventory/inventory_itemDetails.html', context)
    except Exception as e:
        logger.error(f"Error fetching inventory item details: {e}", exc_info=True)
        messages.error(request, "An error occurred while fetching item details.")
        return redirect('home')

@transaction.atomic
@login_required
@lab_permission_required('add_new_item')
def inventory_addNewItem_view(request):
    if not request.user.is_authenticated:
        return redirect('userlogin')

    selected_lab = request.session.get('selected_lab')

    
    if selected_lab:
        item_types_list = item_types.objects.filter(laboratory_id=selected_lab)
    else:
        item_types_list = item_types.objects.none()  # No lab selected, show nothing or handle it accordingly

    qr_code_image = None  # Placeholder for QR code image
    
    if request.method == "POST":
        # Extract form data
        item_name = request.POST.get('item_name')
        item_type_id = request.POST.get('item_type')
        alert_qty = request.POST.get('alert_qty')
        rec_per_inv = request.POST.get('rec_per_inv') == 'on'

        # Handle expiration type
        expiry_type = request.POST.get('expiration_type')  # 'date', 'usage', or 'maintenance'
        if expiry_type == "null":
            expiry_type = None
        max_uses = request.POST.get('max_uses', None) if expiry_type == 'Usage' else None
        maintenance_interval = request.POST.get('maintenance_interval', None) if expiry_type == 'Maintenance' else None

        logger.info(f"User {request.user} attempting to add a new item: {item_name}, type: {item_type_id}")

        try:
            # Dynamic fields from additional columns (based on the selected item_type)
            item_type = item_types.objects.get(itemType_id=item_type_id)
            add_cols_dict = {}
            if item_type.add_cols:
                add_cols = json.loads(item_type.add_cols)
                for col in add_cols:
                    # Replace spaces with underscores to match how they are generated in HTML
                    field_name = f'add_col_{col.replace(" ", "_").lower()}'
                    field_value = request.POST.get(field_name)
                    add_cols_dict[col] = field_value if field_value else None  # Handle None values for empty inputs

            # Convert the additional columns to a JSON string
            add_cols_json = json.dumps(add_cols_dict)


            # Save the data to the database
            new_item = item_description(
                laboratory_id=selected_lab,
                item_name=item_name,
                itemType_id=item_type_id,
                add_cols=add_cols_json,
                alert_qty=alert_qty,
                rec_per_inv=rec_per_inv,
                expiry_type=expiry_type,
                max_uses = max_uses,
                maintenance_interval = maintenance_interval,
            )
            new_item.save()

            qr_details=f"{item_name}"
            qr_code_data = generate_qr_code(new_item.item_id, qr_details)
            # Render the page with QR code and modal trigger
            return render(request, 'mod_inventory/inventory_addNewItem.html', {
                'item_types': item_types_list,
                'qr_code_data': qr_code_data,
                'new_item': new_item,  # Pass new item details for modal
                'show_modal': True,    # Flag to show the modal
            })

        except item_types.DoesNotExist:
            logger.error(f"Item type {item_type_id} does not exist. User: {request.user}")
            messages.error(request, "Invalid item type selected.")
        
        except Exception as e:
            logger.error(f"Failed to add item: {str(e)}", exc_info=True)
            messages.error(request, "An error occurred while adding the item. Please try again.")

    return render(request, 'mod_inventory/inventory_addNewItem.html', {
        'item_types': item_types_list,
        'selected_lab_name': request.session.get('selected_lab_name'),
    })

@transaction.atomic
@login_required
@lab_permission_required('update_item_inventory')
def inventory_updateItem_view(request):
    if not request.user.is_authenticated:
        logger.warning(f"Unauthorized access attempt to inventory_updateItem_view by {request.user}")
        return redirect('userlogin')
    
    if request.method == 'POST':
        selected_laboratory_id = request.session.get('selected_lab')
        current_user = request.user
        action_type = request.POST.get('action_type')
        item_id = request.POST.get('item_name') 
        if not item_id:
            messages.error(request, 'Invalid Item Input')
            return redirect('inventory_updateItem')
        
        inventory_item_id = request.POST.get('inventory_item_id') 
        remarks = request.POST.get('remarks', '')
        remaining_qty=0
        
        try:
            # Fetch item and supplier
            item_instance = get_object_or_404(item_description, item_id=item_id)
            
            # Add or remove inventory logic
            if action_type == 'add':
                qty_add = request.POST.get('amount')
                if not qty_add:
                    messages.error(request, 'Invalid Quantity Input')
                    return redirect('inventory_updateItem')
                qty_add = int(request.POST.get('amount', 0))

                supplier_id = request.POST.get('item_supplier') or None
                date_purchased = request.POST.get('item_date_purchased', None) or None
                date_received = request.POST.get('item_date_received', None) or None
                purchase_price = request.POST.get('item_price', 0.0) or 0.0
                expiration_date = request.POST.get('expiration_date') if item_instance.expiry_type == 'Date' else None
                maintenance_date = request.POST.get('maintenance_date') if item_instance.expiry_type == 'Maintenance' else None
                new_inventory_item = item_inventory.objects.create(
                    item=item_instance,
                    supplier_id=supplier_id,
                    date_purchased=date_purchased,
                    date_received=date_received,
                    purchase_price=purchase_price,
                    qty=qty_add
                )

                item_handling.objects.create(
                    inventory_item=new_inventory_item,
                    updated_by=current_user,
                    changes='A',
                    qty=qty_add,
                    remarks='Add to Inventory'
                )
                # If expiration date is provided, save it
                
                if item_instance.expiry_type != None:
                    max_usage = item_instance.max_uses
                    item_expirations.objects.create(
                        inventory_item=new_inventory_item,
                        expired_date=expiration_date if item_instance.expiry_type == 'Date' else None,
                        next_maintenance_date = maintenance_date if item_instance.expiry_type == 'Maintenance' else None,
                        remaining_uses = max_usage if item_instance.expiry_type == 'Usage' else None,
                    )
                

                # Inside the 'add' action block in your view
                if item_instance.expiry_type=='Date' or item_instance.rec_per_inv == True:
                    qr_data = f"{item_id}, {new_inventory_item.inventory_item_id}"
                else:
                    qr_data = f"{item_id}, 0"

                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=10,
                    border=4,
                )
                qr.add_data(qr_data)
                qr.make(fit=True)

                # Save the QR code as a base64 image
                img = qr.make_image(fill_color="black", back_color="white")
                buffered = BytesIO()
                img.save(buffered, format="PNG")
                qr_base64 = base64.b64encode(buffered.getvalue()).decode()
                qr_code = f"data:image/png;base64,{qr_base64}"

                logger.info(f"User {current_user} added {qty_add} of {item_instance.item_name} (Item ID: {item_id}) to inventory.")

                return JsonResponse({
                    'success': True,
                    'new_inventory_item_id': new_inventory_item.inventory_item_id,
                    'item_qrcode': qr_code,  # Include the QR code in the response
                    'item_name': item_instance.item_name,
                    'action_type': action_type
                })
            
            elif action_type in ['remove', 'report']:# Handle remove & report from inventory
                if action_type=='remove':
                    qty_remove = int(request.POST.get('quantity_removed', 0))
                else:
                    qty_damaged = int(request.POST.get('quantity_damaged', 0))

                quantity = qty_remove if action_type == 'remove' else qty_damaged
                if inventory_item_id:
                    # Fetch specific inventory item for expiration-controlled items
                    inventory_item = get_object_or_404(item_inventory, inventory_item_id=inventory_item_id)
                    if quantity > inventory_item.qty:
                        messages.error(request, "The quantity to remove/report exceeds the available stock.")
                        # return redirect('inventory_updateItem')
                        return JsonResponse({'success': True})
                    else:
                        # Deduct quantity and save the item handling action
                        inventory_item.qty -= quantity
                        inventory_item.save()
                        item_handling.objects.create(
                            inventory_item=inventory_item,
                            timestamp=timezone.localtime(),
                            updated_by=current_user,
                            changes='R' if action_type == 'remove' else 'D',
                            qty=0-quantity,
                            remarks= 'Remove from inventory' if action_type == 'remove' else remarks
                        )
                else:
                    item_totalqty = item_inventory.objects.filter(item=item_instance).aggregate(total_qty=Sum('qty'))
                    total_qty = item_totalqty['total_qty'] or 0  # Use 0 if total_qty is None
                    print(quantity, '>', total_qty)
                    if quantity > total_qty:
                        messages.error(request, "The quantity to remove/report exceeds the available stock.")
                        return JsonResponse({'success': True})

                    remaining_qty = quantity
                    inventory_qs = item_inventory.objects.filter(item=item_instance, qty__gt=0).order_by('date_received')
                    for inventory_item in inventory_qs:
                        if remaining_qty <= 0:
                            break
                        if inventory_item.qty >= remaining_qty:
                            inventory_item.qty -= remaining_qty
                            inventory_item.save()
                            item_handling.objects.create(
                                inventory_item=inventory_item,
                                timestamp=timezone.localtime(),
                                updated_by=current_user,
                                changes='R' if action_type == 'remove' else 'D',
                                qty=0-remaining_qty,
                                remarks= 'Remove from inventory' if action_type == 'remove' else remarks
                            )
                            remaining_qty = 0
                        else:
                            handled_qty = inventory_item.qty
                            remaining_qty -= handled_qty
                            inventory_item.qty = 0
                            inventory_item.save()
                            item_handling.objects.create(
                                inventory_item=inventory_item,
                                timestamp=timezone.localtime(),
                                updated_by=current_user,
                                changes='R' if action_type == 'remove' else 'D',
                                qty=0-handled_qty,
                                remarks= 'Remove from inventory' if action_type == 'remove' else remarks
                            )

                if remaining_qty > 0:
                    messages.error(request, f"Unable to fully remove requested quantity. Remaining: {remaining_qty}")
                    logger.warning(f"User {current_user} attempted to remove more than available stock for {item_instance.item_name} (ID: {item_id}).")
                    return JsonResponse({'success': True})
                else:
                    messages.success(request, f"{action_type.capitalize()} action completed successfully.")
                    logger.info(f"User {current_user} {action_type}d {quantity} of {item_instance.item_name} (ID: {item_id}) from inventory.")
                    return JsonResponse({'success': True})

            return render(request, 'mod_inventory/inventory_updateItem.html')
        
        except Exception as e:
            logger.error(f"Error in inventory_updateItem_view: {str(e)}", exc_info=True)
            messages.error(request, "An error occurred while updating inventory.")
            return JsonResponse({'success': False})


    return render(request, 'mod_inventory/inventory_updateItem.html')

@transaction.atomic
@login_required
@lab_permission_required('view_inventory')
def inventory_itemEdit_view(request, item_id):
    if not request.user.is_authenticated:
        logger.warning(f"Unauthorized access attempt to edit inventory item {item_id} by {request.user}")
        return redirect('userlogin')

    # Get the selected laboratory from the session
    selected_laboratory_id = request.session.get('selected_lab')

    # Get the item_description instance for the selected lab
    item = get_object_or_404(
        item_description,
        item_id=item_id,
        laboratory_id=selected_laboratory_id
    )

    # Parse existing add_cols JSON
    add_cols_data = json.loads(item.add_cols) if item.add_cols else {}

    if request.method == 'POST':
        form = ItemEditForm(request.POST, instance=item, selected_laboratory_id=selected_laboratory_id)

        # Handle additional form fields (rec_expiration, alert_qty)
        expiry_type = request.POST.get('expiration_type', None)
        if expiry_type == "null":
            expiry_type = None

        alert_qty_disabled = request.POST.get('disable_alert_qty', 'off') == 'on'
        rec_per_inv = request.POST.get('rec_per_inv', 'off') == 'on'

        max_uses = int(request.POST.get('max_uses', None)) if expiry_type == 'Usage' else None
        maintenance_interval = int(request.POST.get('maintenance_interval', None)) if expiry_type == 'Maintenance' else None
        
        if form.is_valid():
            print("Form is valid. Saving data...")
            # Save form fields
            form.save()
            # Fetch additional columns based on the selected (or unchanged) itemType
            selected_item_type = form.cleaned_data.get('itemType') or item.itemType
            new_add_cols = json.loads(selected_item_type.add_cols) if selected_item_type and selected_item_type.add_cols else {}

            # Retain common values between old and new additional columns
            updated_add_cols = {}
            for label in new_add_cols:
                # Use the exact label name as it appears in the form
                field_name = label.split('(')[0].strip() if '(' in label else label
                updated_add_cols[label] = request.POST.get(field_name, add_cols_data.get(label, ''))
                print(f"Saving field '{field_name}': {updated_add_cols[label]}")

            # Update the item attributes
            item.add_cols = json.dumps(updated_add_cols)
            item.rec_per_inv = rec_per_inv
            item.expiry_type = expiry_type
            item.max_uses = max_uses
            item.maintenance_interval = maintenance_interval
            item.alert_qty = 0 if alert_qty_disabled else request.POST.get('alert_qty', item.alert_qty)

            # Force save with update_fields to ensure data is saved
            item.save(update_fields=['add_cols', 'rec_per_inv', 'expiry_type', 'alert_qty', 'max_uses', 'maintenance_interval'])
            logger.info(f"User {request.user} updated item {item.item_name} (ID: {item.item_id}) with new attributes.")

            # Verify save by reloading the item
            item.refresh_from_db()
            # print("Reloaded item from DB:", item.add_cols)

            # Redirect after saving
            return redirect('inventory_itemDetails_view', item_id=item_id)
        else:
            # print("Form is not valid. Errors:", form.errors)
            logger.warning(f"User {request.user} failed to update item {item.item_name}. Errors: {form.errors}")

    else:
        form = ItemEditForm(instance=item, selected_laboratory_id=selected_laboratory_id)

    # Pass the additional columns with options to the template
    dropdown_fields = {}
    for label, value in add_cols_data.items():
        # Check if the field has specific options, e.g., "Grade (A, B, C)"
        if '(' in label and ')' in label:
            options = label[label.find('(') + 1:label.find(')')].split(',')
            dropdown_fields[label] = [option.strip() for option in options]

    return render(request, 'mod_inventory/inventory_itemEdit.html', {
        'form': form,
        'item': item,
        'add_cols_data': add_cols_data,
        'dropdown_fields': dropdown_fields,  # New context variable
        'is_alert_disabled': item.alert_qty == 0,
    })

@transaction.atomic
@login_required
@lab_permission_required('view_inventory')
def inventory_itemDelete_view(request, item_id):
    if not request.user.is_authenticated:
        logger.warning(f"Unauthorized access attempt to delete item {item_id} by {request.user}")
        return redirect('userlogin')
    
    try:
    
        item = get_object_or_404(item_description, item_id=item_id)

        # Redirect to view inventory after disabling
        if request.method == 'POST':
            item.is_disabled = 1  # Mark item as disabled
            item.save()
            logger.info(f"User {request.user} disabled item '{item.item_name}' (Item ID: {item_id}) in lab {item.laboratory_id}.")
            return redirect('inventory_view')

        # Prepare context for rendering the confirmation template
        item_type = item.itemType
        lab = get_object_or_404(laboratory, laboratory_id=item.laboratory_id)
        context = {
            'item': item,
            'itemType_name': item_type.itemType_name if item_type else None,
            'laboratory_name': lab.name if lab else None,
        }

        return render(request, 'mod_inventory/inventory_itemDelete.html', context)

    except Exception as e:
        logger.error(f"Error while deleting item {item_id} by {request.user}: {str(e)}", exc_info=True)
        messages.error(request, "An error occurred while deleting the item.")
        return redirect('home')


@login_required
@lab_permission_required('view_inventory')
def get_item_type_add_cols(request, itemType_id):
    if not request.user.is_authenticated:
        logger.warning(f"Unauthorized attempt to access get_item_type_add_cols by {request.user}")
        return redirect('userlogin')
    
    try:
        item_type = item_types.objects.get(itemType_id=itemType_id)
        add_cols = json.loads(item_type.add_cols) if item_type.add_cols else []
        logger.info(f"User {request.user} fetched additional columns for itemType {itemType_id}: {add_cols}")
        return JsonResponse({'add_cols': add_cols})
    
    except item_types.DoesNotExist:
        logger.error(f"ItemType {itemType_id} not found. Requested by {request.user}", exc_info=True)
        return JsonResponse({'error': 'Item type not found'}, status=404)

    except Exception as e:
        logger.error(f"Unexpected error while fetching itemType {itemType_id} for {request.user}: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Internal Server Error'}, status=500)
    
@login_required
@lab_permission_required('physical_count')
def inventory_physicalCount_view(request):
    if not request.user.is_authenticated:
        logger.warning(f"Unauthorized attempt to access inventory_physicalCount_view by {request.user}")
        return redirect('userlogin')
    
    try:
        # Get the selected laboratory from the session
        selected_laboratory_id = request.session.get('selected_lab')
        item_types_list = item_types.objects.filter(laboratory_id=selected_laboratory_id)
        selected_item_type = request.GET.get('item_type')
        current_user = request.user

        # Filter items by laboratory and selected item type
        if selected_item_type:
            inventory_items = item_description.objects.filter(
                laboratory_id=selected_laboratory_id, 
                itemType_id=selected_item_type, 
                is_disabled=0
            ).annotate(total_qty=Sum('item_inventory__qty'))
        else:
            inventory_items = item_description.objects.filter(
                laboratory_id=selected_laboratory_id, 
                is_disabled=0
            ).annotate(total_qty=Sum('item_inventory__qty'))
            
        # Parse add_cols JSON and fetch individual inventory records if necessary
        for item in inventory_items:
            # Parse additional columns if available
            item.parsed_add_cols = json.loads(item.add_cols) if item.add_cols else {}
            
            # If item tracks expiration or per-inventory quantity, fetch individual inventories
            if item.expiry_type != None or item.rec_per_inv:
                item.individual_inventories = item_inventory.objects.filter(item=item, qty__gt=0).order_by('inventory_item_id')
            else:
                # Total quantity for items without inventory-level tracking
                item.total_qty = item_inventory.objects.filter(item=item).aggregate(total_qty=Sum('qty'))['total_qty'] or 0

        # Check if the form is submitted
        if request.method == "POST":
            for item in inventory_items:
                if item.expiry_type != None or item.rec_per_inv:
                    # Handle item_inventory updates individually
                    for inventory in item.individual_inventories:
                        try:
                            count_qty = request.POST.get(f'count_qty_{inventory.inventory_item_id}')
                            print(f'Item ID: {item.item_name}, Count Qty: {count_qty}')  # For debugging
                            if count_qty is not None and count_qty != '':
                                count_qty = int(count_qty)
                            else:
                                count_qty = inventory.qty

                            discrepancy_qty = count_qty - inventory.qty

                            # Add or remove discrepancy from inventory
                            if discrepancy_qty != 0:
                                adjust_inventory_item('inv',inventory, discrepancy_qty, current_user, "P", "Physical count adjustment")
                        except Exception as e:
                            logger.error(f"Error processing inventory {inventory.inventory_item_id} during physical count: {e}", exc_info=True)
                            messages.error(request, f"Error updating item {inventory.inventory_item_id}")

                else:
                    try:
                        count_qty = request.POST.get(f'count_qty_{item.item_id}')
                        print(f'Item ID: {item.item_name}, Count Qty: {count_qty}')  # For debugging
                        if count_qty is not None and count_qty != '':
                            count_qty = int(count_qty)
                        else:
                            count_qty = item.total_qty
                        
                        discrepancy_qty = count_qty - item.total_qty
                        if discrepancy_qty != 0:
                            adjust_inventory_item('item',item, discrepancy_qty, current_user, "P", "Physical count adjustment")
                    except Exception as e:
                        logger.error(f"Error processing item {item.item_id} during physical count: {e}", exc_info=True)
                        messages.error(request, f"Error updating item {item.item_id}")

            messages.success(request, 'Physical count saved successfully!')
            return redirect('inventory_physicalCount')

        return render(request, 'mod_inventory/inventory_physicalCount.html', {
            'inventory_items': inventory_items,
            'item_types': item_types_list,
            'selected_item_type': int(selected_item_type) if selected_item_type else None,
        })
    except Exception as e:
        logger.error(f"Critical error in inventory_physicalCount_view: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('home')

@transaction.atomic
@login_required
def adjust_inventory_item(status, inventory, discrepancy_qty, user, change_type, remark):
    try:
        if discrepancy_qty > 0:
            if status=='item':
                item_inventory_instance = item_inventory.objects.create(
                    item=inventory,
                    qty=discrepancy_qty,
                    date_purchased=timezone.localtime(),
                    date_received=timezone.localtime(),
                    supplier=None  # Supplier can be left as None for physical adjustments
                )
                item_handling.objects.create(
                    inventory_item=item_inventory_instance,
                    updated_by=user,
                    changes=change_type,
                    qty=discrepancy_qty,
                    remarks=remark
                )
                logger.info(f"Added {discrepancy_qty} to item {inventory.item_name} by {user}")

            else:
                item_handling.objects.create(
                    inventory_item=inventory,
                    updated_by=user,
                    changes=change_type,
                    qty=discrepancy_qty,
                    remarks=remark
                )
                inventory.qty += discrepancy_qty
                inventory.save()
                logger.info(f"Added {discrepancy_qty} to inventory {inventory.inventory_item_id} by {user}")

        elif discrepancy_qty < 0:
            remaining_amount = abs(discrepancy_qty)
            if status=='item':
                print('pass', inventory)
                item_description_instance = get_object_or_404(item_description, item_id=inventory.item_id)
                if item_description_instance.expiry_type != None:
                    item_inventory_queryset = item_inventory.objects.filter(
                            item=item_description_instance,
                            qty__gt=0
                        ).annotate(
                            expired_date=F('item_expirations__expired_date')
                        ).order_by('expired_date')
                else:
                    item_inventory_queryset = item_inventory.objects.filter(
                        item=item_description_instance,
                        qty__gt=0
                    ).order_by('date_received')
                
                for item_inventory_instance in item_inventory_queryset:
                    if remaining_amount <= 0:
                        break

                    if item_inventory_instance.qty >= remaining_amount:
                        try:
                            remove_item_from_inventory(item_inventory_instance, remaining_amount, user, 'P', 'Physical Count Adjustment')
                            remaining_amount = 0
                        except ValueError as e:
                            print(e)
                            break
                    else:
                        try:
                            remove_item_from_inventory(item_inventory_instance, item_inventory_instance.qty, user, 'P', 'Physical Count Adjustment')
                            remaining_amount -= item_inventory_instance.qty
                        except ValueError as e:
                            logger.error(f"Error removing inventory item {item_inventory_instance.inventory_item_id}: {e}", exc_info=True)
                            break
            else:
                remove_item_from_inventory(inventory, remaining_amount, user, change_type, remark)
    except Exception as e:
        logger.error(f"Unexpected error adjusting inventory {inventory}: {e}", exc_info=True)

@transaction.atomic
@login_required
@lab_permission_required('manage_suppliers')
def inventory_manageSuppliers_view(request):
    if not request.user.is_authenticated:
        logger.warning(f"Unauthorized attempt to access inventory_manageSuppliers_view by {request.user}")
        return redirect('userlogin')

    try:
        selected_laboratory_id = request.session.get('selected_lab')
        
        # Annotate suppliers with the count of items supplied in item_inventory
        lab_suppliers = suppliers.objects.filter(laboratory=selected_laboratory_id, is_disabled=0).annotate(
            supplied_items_count=models.Count('item_inventory')
        )

        if request.method == "POST":
            supplier_name = request.POST.get("supplier_name")
            contact_person = request.POST.get("contact_person")
            contact_number = request.POST.get("contact_number") or None
            supplier_desc = request.POST.get("description")
            email = request.POST.get("email")

            new_supplier = suppliers(
                laboratory_id=selected_laboratory_id,
                supplier_name=supplier_name,
                contact_person=contact_person,
                contact_number=contact_number,
                description=supplier_desc,
                email=email
            )
            new_supplier.save()
            messages.success(request, 'Supplier added successfully.')
            return redirect('inventory_manageSuppliers')

        return render(request, 'mod_inventory/inventory_manageSuppliers.html', {
            'suppliers': lab_suppliers,
        })
    except Exception as e:
        logger.error(f"Unexpected error in inventory_manageSuppliers_view: {e}", exc_info=True)
        messages.error(request, "An error occurred. Please try again.")
        return redirect('home')

@transaction.atomic
@lab_permission_required('manage_suppliers')
def inventory_supplierDetails_view(request, supplier_id):
    try:
        if not request.user.is_authenticated:
            logger.warning(f"Unauthorized attempt to access inventory_supplierDetails_view by {request.user}")
            return redirect('userlogin')

        # Get the supplier details using the supplier_id
        supplier = get_object_or_404(suppliers, suppliers_id=supplier_id)

        # Get item handling entries associated with the supplier
        item_handling_entries = item_handling.objects.filter(
            inventory_item__supplier=supplier,
            changes='A'
        ).select_related('inventory_item')

        if request.method == "POST":
            if 'edit_supplier' in request.POST:
                try:
                    # Handle supplier edit
                    supplier.supplier_name = request.POST.get("supplier_name")
                    supplier.contact_person = request.POST.get("contact_person") or None
                    supplier.contact_number = request.POST.get("contact_number") or None
                    supplier.description = request.POST.get("description") or None
                    supplier.email = request.POST.get("email") or None
                    supplier.save()
                    messages.success(request, 'Supplier details edited successfully.')
                    return redirect('inventory_supplierDetails', supplier_id=supplier.suppliers_id)
                except Exception as e:
                        logger.error(f"Error updating supplier {supplier.supplier_name}: {e}", exc_info=True)
                        messages.error(request, "Error updating supplier. Please try again.")

            elif 'disable_supplier' in request.POST:
                # Handle supplier disable
                try:
                    supplier.is_disabled = True
                    supplier.save()
                    messages.success(request, 'Supplier deleted successfully.')
                    return redirect('inventory_manageSuppliers')
                except Exception as e:
                        logger.error(f"Error disabling supplier {supplier.supplier_name}: {e}", exc_info=True)
                        messages.error(request, "Error disabling supplier. Please try again.")

        return render(request, 'mod_inventory/inventory_supplierDetails.html', {
            'supplier': supplier,
            'item_handling_entries': item_handling_entries,
        })
    except Exception as e:
        logger.error(f"Unexpected error in inventory_supplierDetails_view: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('inventory_manageSuppliers')
    
@login_required
@lab_permission_required('configure_inventory')
def inventory_config_view(request):
    try:
        if not request.user.is_authenticated:
            logger.warning(f"Unauthorized attempt to access inventory_config_view by {request.user}")
            return redirect('userlogin')


        # Retrieve the selected laboratory ID from the session
        selected_lab = request.session.get('selected_lab')
        
        # Query categories based on the selected laboratory ID
        if selected_lab:
            categories = item_types.objects.filter(laboratory_id=selected_lab)
        else:
            categories = item_types.objects.none()  # No categories if no lab is selected

        # Get attributes for the first category if it exists
        attributes = json.loads(categories[0].add_cols) if categories.exists() else []

        # Prepare the context for rendering the template
        context = {
            'categories': categories,
            'attributes': attributes
        }

        return render(request, 'mod_inventory/inventory_config.html', context)
    except Exception as e:
        logger.error(f"Error loading inventory config: {e}", exc_info=True)
        messages.error(request, "Error loading inventory configuration. Please try again.")
        return redirect('home')
    
@transaction.atomic
@login_required
@lab_permission_required('configure_inventory')
def add_category(request):
    try:
        if not request.user.is_authenticated:
            logger.warning(f"Unauthorized attempt to access add_category by {request.user}")
            return redirect('userlogin')

        # Get laboratory_id from the session
        laboratory_id = request.session.get('selected_lab')

        if not laboratory_id:
            messages.error(request, "No laboratory selected.")
            return redirect('inventory_config')

        if request.method == 'POST':
            category_name = request.POST.get('category')  # Change 'category_name' to 'category'
            add_cols = request.POST.get('add_cols')  # Ensure you have this input field in your form

            # Debugging output to check what is received
            print(f"Received category_name: {category_name}, add_cols: {add_cols}")  

            if not category_name:  # Check if category_name is empty
                messages.error(request, "Category name cannot be empty.")
                return redirect('inventory_config')

            new_category = item_types(
                laboratory_id=laboratory_id,
                itemType_name=category_name,
                # add_cols=add_cols
            )
            new_category.save()

            messages.success(request, 'Category added successfully!')
            return redirect('inventory_config')

        return render(request, 'mod_inventory/add_category.html')
    except Exception as e:
        logger.critical(f"Unexpected error in add_category: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('inventory_config')

@login_required
@lab_permission_required('configure_inventory')
def delete_category(request, category_id):
    try:
        if not request.user.is_authenticated:
            logger.warning(f"Unauthorized attempt to access delete_category by {request.user}")
            return redirect('userlogin')
    
        if request.method == 'POST':
            category = get_object_or_404(item_types, pk=category_id)

            # Optional: Ensure that the category belongs to the selected laboratory
            selected_lab = request.session.get('selected_lab')
            if category.laboratory_id != selected_lab:
                logger.warning(f"Unauthorized attempt to delete category {category_id} by {request.user}")
                return JsonResponse({'success': False, 'message': "Category does not belong to the selected laboratory."}, status=400)

            category.delete()  # Delete the category from item_types
            logger.info(f"Category {category_id} deleted by {request.user}")
            messages.success(request, 'Category deleted successfully!')
            return JsonResponse({'success': True})

        return JsonResponse({'success': False}, status=400)
    except Exception as e:
        logger.error(f"Error deleting category {category_id}: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': "An error occurred while deleting the category."}, status=500)

@transaction.atomic
@login_required
@lab_permission_required('configure_inventory')
def add_attributes(request):
    try:
        if not request.user.is_authenticated:
            logger.warning(f"Unauthorized attempt to access add_attributes by {request.user}")
            return redirect('userlogin')
        
        if request.method == 'POST':
            category_id = request.POST['category']
            attribute_name = request.POST['attributeName']
            fixed_values = request.POST.getlist('fixedValues')

            category = get_object_or_404(item_types, pk=category_id)

            # Check if the category belongs to the selected laboratory
            selected_lab = request.session.get('selected_lab')
            if category.laboratory_id != selected_lab:
                logger.warning(f"Unauthorized attempt to add attribute to category {category_id} by {request.user}")
                return JsonResponse({'success': False, 'message': "Category does not belong to the selected laboratory."}, status=400)

            # Initialize add_cols if necessary
            if category.add_cols is None:
                add_cols = []  # Initialize to empty list
            else:
                add_cols = json.loads(category.add_cols)

            # Create a combined attribute string with fixed values if any
            combined_attribute = f"{attribute_name} ({', '.join(fixed_values)})" if fixed_values else attribute_name

            if combined_attribute not in add_cols:  # Prevent duplicate attributes
                add_cols.append(combined_attribute)
                category.add_cols = json.dumps(add_cols)
                category.save()
                logger.info(f"Attribute '{combined_attribute}' added to category {category_id} by {request.user}")
                return JsonResponse({'success': True, 'attribute': combined_attribute})
            else:
                return JsonResponse({'success': False, 'message': "Attribute already exists."})

        return JsonResponse({'success': False, 'message': "Invalid request."})
    except Exception as e:
        logger.error(f"Error adding attribute to category {category_id}: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': "An error occurred while adding the attribute."}, status=500)

@login_required
@lab_permission_required('configure_inventory')
def get_fixed_choices(request, category_id):
    try:
        if not request.user.is_authenticated:
            logger.warning(f"Unauthorized attempt to access get_fixed_choices by {request.user}")
            return redirect('userlogin')
        
        category = get_object_or_404(item_types, pk=category_id)
        return JsonResponse({'fixed_choices': category.fixed_choices})
    
    except Exception as e:
        logger.error(f"Error retrieving fixed choices for category {category_id}: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': "An error occurred while fetching fixed choices."}, status=500)

@transaction.atomic
@login_required
@lab_permission_required('configure_inventory')
def delete_attribute(request, category_id, attribute_name):
    try:
        if request.method == 'POST':
            category = get_object_or_404(item_types, pk=category_id)

            selected_lab = request.session.get('selected_lab')
            if category.laboratory_id != selected_lab:
                logger.warning(f"Unauthorized attempt by {request.user} to delete attribute '{attribute_name}' from category {category_id}")
                return JsonResponse({'success': False, 'message': "Category does not belong to the selected laboratory."}, status=400)

            try:
                add_cols = json.loads(category.add_cols) if category.add_cols else []
            except json.JSONDecodeError:
                add_cols = []

            if attribute_name in add_cols:
                add_cols.remove(attribute_name)
                category.add_cols = json.dumps(add_cols)
                category.save()
                logger.info(f"Attribute '{attribute_name}' removed from category {category_id} by {request.user}")
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'message': 'Attribute not found.'}, status=404)

        return JsonResponse({'success': False}, status=400)

    except Exception as e:
        logger.error(f"Error deleting attribute '{attribute_name}' from category {category_id}: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': "An error occurred while deleting the attribute."}, status=500)

@login_required   
@lab_permission_required('configure_inventory')
def get_add_cols(request, category_id):
    try:
        category = get_object_or_404(item_types, pk=category_id)

        selected_lab = request.session.get('selected_lab')
        if category.laboratory_id != selected_lab:
            logger.warning(f"Unauthorized attempt by {request.user} to access add_cols for category {category_id}")
            return JsonResponse({'success': False, 'message': "Category does not belong to the selected laboratory."}, status=400)

        add_cols = json.loads(category.add_cols) if category.add_cols else []
        return JsonResponse({'add_cols': add_cols})
    
    except Exception as e:
        logger.error(f"Error retrieving add_cols for category {category_id}: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': "An error occurred while retrieving add_cols."}, status=500)

@login_required
def inventory_tobuyList(request):
    try:
        return render(request, 'mod_inventory/inventory_tobuyList.html')
    
    except Exception as e:
        logger.error(f"Error rendering to-buy list: {e}", exc_info=True)
        messages.error(request, "An error occurred while loading the to-buy list.")
        return redirect('home')  # Redirect user to a safe page if error occurs


# =====================================================
#BORROWING
@login_required
@lab_permission_required('borrow_items')
def borrowing_view(request):
    return render(request, 'mod_borrowing/borrowing.html')

@login_required
@lab_permission_required('borrow_items')
def borrowing_student_prebookview(request):
    try:
        unit_choices = get_unit_choices()
        laboratory_id = request.session.get('selected_lab')
        user = request.user
        lab = get_object_or_404(borrowing_config, laboratory_id=laboratory_id)

        # Check if pre-booking is allowed
        if not lab.allow_prebook:
            logger.warning(f"Unauthorized prebook attempt by {request.user} in lab {laboratory_id}")
            return render(request, 'error_page.html', {'message': 'Pre-booking is not allowed for this laboratory.'})

        # Fetch the prebook-specific questions
        prebook_questions = lab.get_questions(mode='prebook')

        # Fetch all inventory items
        inventory_items = item_description.objects.filter(
            laboratory_id=laboratory_id,
            is_disabled=0,
            allow_borrow=1  # Only get items that can be borrowed
        ).select_related('itemType').annotate(total_qty=Sum('item_inventory__qty'))

       # Group items by their itemType
        items_by_type = {}
        for item in inventory_items:
            item_type = item.itemType.itemType_name
            # Parse add_cols if it exists and is a JSON string
            add_cols = {}
            if item.add_cols:
                try:
                    add_cols = json.loads(item.add_cols)  # Parse JSON string into a dictionary
                except ValueError:
                    add_cols = {}
            item.add_cols = add_cols  # Attach parsed add_cols to the item object

            if item_type not in items_by_type:
                items_by_type[item_type] = []
            items_by_type[item_type].append(item)

        if request.method == 'POST':
            # Fetch form data
            borrowing_type = request.POST.get('borrowing-type')
            one_day_date = request.POST.get('one_day_booking_date')
            from_date = request.POST.get('from_date')
            to_date = request.POST.get('to_date')
            equipment_rows = request.POST.getlist('equipment_ids[]')  # List of equipment items
            quantities = request.POST.getlist('quantities[]')         # Corresponding quantities
            units = request.POST.getlist('units[]')  

            # Collect responses to the custom questions
            custom_question_responses = {}
            for question in prebook_questions:
                custom_question_responses[question['question_text']] = request.POST.get(question['question_text'])


            # Get the current time in the configured timezone
            current_time = timezone.localtime()

            # Format the date as "Month DD, YYYY"
            request_date = current_time.date()

            # request_date = timezone.localtime().date()

            # Determine borrow and due dates based on borrowing type
            error_message = None         
    
           
            if borrowing_type == 'oneday':
                borrow_date = one_day_date
                due_date = one_day_date
                type_status = 'A'
                # Validate that the borrowing date is not in the past
                if one_day_date < request_date.strftime('%Y-%m-%d'):
                    error_message = 'The borrowing date cannot be earlier than today for one-day borrowing.'
                # Validate the one-day borrowing: must be at least 3 days from the request date
                min_borrow_date = request_date + timedelta(days=int(lab.prebook_lead_time))
                if one_day_date <= min_borrow_date.strftime('%Y-%m-%d'):
                    error_message = f'For one-day borrowing, the requested date must be at least {lab.prebook_lead_time} days from today.'
            else:
                borrow_date = from_date
                due_date = to_date
                type_status = 'P'
                # Validate the long-term borrowing
                min_from_date = request_date + timedelta(days=int(lab.prebook_lead_time))
                if from_date < min_from_date.strftime('%Y-%m-%d'):
                    error_message = f'The "From" date for long-term borrowing must be at least {lab.prebook_lead_time} days from the request date.'

                if to_date < from_date:
                    error_message = '"To" date cannot be earlier than the "From" date.'

            # Validate quantity limits for equipment
            equipment_rows = request.POST.getlist('equipment_ids[]')  # List of equipment items
            quantities = request.POST.getlist('quantities[]')   
            units = request.POST.getlist('units[]')     # Corresponding quantities
            error_message_qty = None
            


            if not equipment_rows:
                error_message = "You must select at least one equipment to proceed."
                return render(request, 'mod_borrowing/borrowing_studentPrebook.html', {
                    'error_message': error_message,
                    'current_date': request_date,
                    'items_by_type': items_by_type, 
                    'prebook_questions': prebook_questions, 
                    'unit_choices':unit_choices
            })

    
            for i, item_id in enumerate(equipment_rows):
                quantity = int(quantities[i])
                item = item_description.objects.get(item_id=item_id)                

                # If qty_limit is not null and quantity exceeds the limit, show error
                if item.qty_limit is not None and quantity > item.qty_limit:
                    error_message_qty = f"Quantity for '{item.item_name}' exceeds the quantity limit of {item.qty_limit}."
                    break  # Stop checking after the first error

            # If there was any error related to quantity, re-render the form with the error
            if error_message or error_message_qty:
                return render(request, 'mod_borrowing/borrowing_studentPrebook.html', {
                    'error_message': error_message or error_message_qty,
                    'current_date': request_date,
                    'items_by_type': items_by_type,  # Include grouped items here
                    'unit_choices':unit_choices
                })

            # If validation passes, proceed with insertion
            borrow_entry = borrow_info.objects.create(
                laboratory_id=laboratory_id,
                user=user,
                request_date=timezone.localtime(),  # Use current timestamp
                borrow_date=borrow_date,
                due_date=due_date,
                status=type_status,  # Set initial status to 'Pending'
                questions_responses=custom_question_responses
            )

            # Process equipment details
            equipment_rows = request.POST.getlist('equipment_ids[]')  # List of equipment items
            quantities = request.POST.getlist('quantities[]')       # Corresponding quantities
            units = request.POST.getlist('units[]')

            for i, item_id in enumerate(equipment_rows):
                quantity = int(quantities[i])
                unit = units[i]
                # Fetch the item from core_item_description
                item = item_description.objects.get(item_id=item_id)
                
                # Insert the item into borrowed_items table
                borrowed_items.objects.create(
                    borrow=borrow_entry,
                    item=item,
                    qty=quantity,
                    unit=unit
                )
            return redirect('borrowing_studentviewPreBookRequests')

        # Get the current time in the configured timezone
        current_time = timezone.localtime()

        # Format the date as "Month DD, YYYY"
        current_date = current_time.strftime('%Y-%m-%d')


    except Http404:
        logger.error(f"Lab not found for user {request.user}")
        return render(request, 'error_page.html', {'message': 'The laboratory was not found.'})
    except Exception as e:
        logger.error(f"Unexpected error in borrowing_student_prebookview: {e}", exc_info=True)
        return render(request, 'error_page.html', {'message': 'An unexpected error occurred. Please try again.'})

    return render(request, 'mod_borrowing/borrowing_studentPrebook.html', {
        'current_date': current_date,
        'items_by_type': items_by_type,  # Pass grouped items to the template
        'prebook_questions': prebook_questions,  # Pass the prebook questions to the template
        'unit_choices':unit_choices
    })

@login_required
@lab_permission_required('borrow_items')
def get_items_by_type(request, item_type_id):
    try:
        items = item_description.objects.filter(itemType_id=item_type_id, is_disabled=0, allow_borrow=1)
        item_list = []

        for item in items:
            total_qty = item_inventory.objects.filter(item_id=item.item_id).aggregate(total_qty=Sum('qty'))['total_qty'] or 0
            item_list.append({
                'item_id': item.item_id,
                'item_name': item.item_name,
                'total_qty': total_qty
            })

        return JsonResponse(item_list, safe=False)
    except Exception as e:
        logger.error(f"Error fetching items for type {item_type_id}: {e}", exc_info=True)
        return JsonResponse({'error': "Failed to fetch items"}, status=500)

@login_required
@lab_permission_required('borrow_items')
def get_quantity_for_item(request, item_id):
    try:
        total_quantity = item_inventory.objects.filter(item_id=item_id).aggregate(total_qty=Sum('qty'))['total_qty'] or 0
        return JsonResponse({'total_qty': total_quantity})
    except Exception as e:
        logger.error(f"Error fetching quantity for item {item_id}: {e}", exc_info=True)
        return JsonResponse({'error': "Failed to fetch quantity"}, status=500)


@login_required
@lab_permission_required('borrow_items')
def borrowing_student_walkinview(request):
    try:
        current_time = timezone.localtime()  
        unit_choices = get_unit_choices()

        # Get session details
        laboratory_id = request.session.get('selected_lab')
        user_id = request.user

        # Fetch the lab's borrowing configuration
        lab = get_object_or_404(borrowing_config, laboratory_id=laboratory_id)

        # Check if walk-ins are allowed
        if not lab.allow_walkin:
            return render(request, 'error_page.html', {'message': 'Walk-ins are not allowed for this laboratory.'})

        # Fetch the walk-in-specific questions
        walkin_questions = lab.get_questions(mode='walkin')

        # Fetch all inventory items
        inventory_items = item_description.objects.filter(
            laboratory_id=laboratory_id,
            is_disabled=0,
            allow_borrow=1  # Only get items that can be borrowed
        ).select_related('itemType').annotate(total_qty=Sum('item_inventory__qty'))
        
        # Group items by their itemType
        items_by_type = {}
        for item in inventory_items:
                item_type = item.itemType.itemType_name
                # Parse add_cols if it exists and is a JSON string
                add_cols = {}
                if item.add_cols:
                    try:
                        add_cols = json.loads(item.add_cols)  # Parse JSON string into a dictionary
                    except ValueError:
                        add_cols = {}
                item.add_cols = add_cols  # Attach parsed add_cols to the item object

                if item_type not in items_by_type:
                    items_by_type[item_type] = []
                items_by_type[item_type].append(item)

        if request.method == 'POST':
            # request_date = timezone.localtime()
            # Format the date as "Month DD, YYYY"
            request_datetime = current_time.strftime('%Y-%m-%d %H:%M:%S')
            request_date = current_time.strftime('%Y-%m-%d')
            print(request_date)

            borrow_date = request_date
            due_date = request_date

            # Collect responses to the custom walk-in questions
            custom_question_responses = {}
            for question in walkin_questions:
                response = request.POST.get(question['question_text'])
                custom_question_responses[question['question_text']] = response

            # Insert into core_borrow_info
            borrow_entry = borrow_info.objects.create(
                laboratory_id=laboratory_id,
                user=user_id,
                request_date=request_datetime,
                borrow_date=borrow_date,
                due_date=due_date,
                status='A',  # Set initial status to 'Pending'
                questions_responses=custom_question_responses  # Save the user's responses to the questions
            )

            # Process equipment details
            equipment_rows = request.POST.getlist('equipment_ids[]')  # List of equipment items
            quantities = request.POST.getlist('quantities[]')       # Corresponding quantities
            units = request.POST.getlist('units[]')

            # Validate equipment quantities and items
            error_message = None
            if not equipment_rows:  # If no equipment is selected
                error_message = 'Please select at least one equipment item to borrow.'
                return render(request, 'mod_borrowing/borrowing_studentWalkIn.html', {
                    'current_date': request_date,
                    'equipment_list': item_description.objects.filter(laboratory_id=laboratory_id, is_disabled=0, allow_borrow=1),
                    'error_message': error_message,
                    'inventory_items': inventory_items,
                    'walkin_questions': walkin_questions,  # Pass walk-in questions back to the template
                    'items_by_type': items_by_type,  # Pass grouped items to the template
                    'lab_config': lab,
                    'unit_choices': unit_choices
                })

        
            
            for i, item_id in enumerate(equipment_rows):
                try:
                    quantity = int(quantities[i])
                    if quantity <= 0:
                        error_message = 'Quantity must be greater than 0 for each item.'
                        break

                    # Check if item exists and is borrowable
                    item = item_description.objects.filter(item_id=item_id, is_disabled=0, allow_borrow=1).first()
                    if not item:
                        error_message = f'Item with ID {item_id} is not available for borrowing.'
                        break
                # Check if requested quantity exceeds the qty_limit of the item
                    if item.qty_limit is not None and quantity > item.qty_limit:
                        error_message = f'Quantity requested for {item.item_name} exceeds the available limit ({item.qty_limit}).'
                        break
                except (ValueError, IndexError) as e:
                    error_message = 'Invalid quantity or item ID.'
                    break

            if error_message:
                return render(request, 'mod_borrowing/borrowing_studentWalkIn.html', {
                    'current_date': request_date,
                    'equipment_list': item_description.objects.filter(laboratory_id=laboratory_id, is_disabled=0, allow_borrow=1),
                    'error_message': error_message,
                    'inventory_items': inventory_items,
                    'walkin_questions': walkin_questions,  # Pass walk-in questions back to the template
                    'items_by_type': items_by_type,  # Pass grouped items to the template
                    'lab_config': lab,
                    'unit_choices': unit_choices 
                })

            # If validation passes, insert items into borrowed_items
            for i, item_id in enumerate(equipment_rows):
                quantity = int(quantities[i])
                unit = units[i]
                if quantity <= 0:
                    continue  # Skip if quantity is invalid

                item = item_description.objects.get(item_id=item_id)

                # Check if item already exists in borrowed_items
                existing_borrowed_item = borrowed_items.objects.filter(borrow=borrow_entry, item=item).first()
                if existing_borrowed_item:
                    continue  # Skip insertion if it already exists

                # Insert the item into borrowed_items table
                borrowed_item = borrowed_items.objects.create(
                    borrow=borrow_entry,
                    item=item,
                    qty=quantity,
                    unit=unit
                )
            logger.info(f"Walk-in request created successfully by {request.user} in lab {laboratory_id}")
            return redirect('borrowing_studentviewPreBookRequests')

        # Fetch the current date and all equipment items including chemicals
        # current_date = current_time.strftime('%B %d, %Y')
        current_date = current_time.strftime('%Y-%m-%d')

        # Get unique item types for dropdown filtering
        item_types_list = item_types.objects.filter(laboratory_id=laboratory_id)

        # Fetch all equipment items including their total quantities
        equipment_list = item_description.objects.filter(
            laboratory_id=laboratory_id,
            is_disabled=0,
            allow_borrow=1  # Only include items that can be borrowed
        ).annotate(total_qty=Sum('item_inventory__qty'))  # Annotate with total quantity

        # Get all item types for the selected laboratory
        item_types_list = item_types.objects.filter(laboratory_id=laboratory_id)

        # Fetch inventory items and order them by item type name
        inventory_items = item_description.objects.filter(
            laboratory_id=laboratory_id,
            is_disabled=0,  # Only get items that are enabled
            allow_borrow=1  # Only get items that can be borrowed
        ).annotate(total_qty=Sum('item_inventory__qty'))  # Calculate total quantity
        inventory_items = inventory_items.select_related('itemType').order_by('itemType__itemType_name')  # Order by itemType name

        return render(request, 'mod_borrowing/borrowing_studentWalkIn.html', {
            'current_date': current_date,
            'equipment_list': equipment_list,
            'item_types': item_types_list,  # Pass item types to the template
            'inventory_items': inventory_items,
            'walkin_questions': walkin_questions,  # Pass the walk-in questions to the template
            'items_by_type': items_by_type,  # Pass grouped items to the template
            'unit_choices': unit_choices 
        })
    except Exception as e:
        logger.error(f"Unexpected error in borrowing_student_walkinview: {e}", exc_info=True)
        return render(request, 'error_page.html', {'message': 'An unexpected error occurred. Please try again.'})

@login_required
@lab_permission_required('borrow_items')
def borrowing_student_viewPreBookRequestsview(request):
    try:
        if not request.user.is_authenticated:
            logger.warning(f"Unauthorized access attempt to Pre-Book Requests by {request.user}")
            return redirect('userlogin')

        current_user = request.user

        # Get the borrowing configuration for the laboratory
        laboratory_id = request.session.get('selected_lab')
        lab_config = get_object_or_404(borrowing_config, laboratory_id=laboratory_id)

        # Filter borrow_info based on statuses
        prebook_requests = borrow_info.objects.annotate(
            truncated_request_date=TruncDate('request_date')
        ).filter(
            user=current_user,
            request_date__isnull=False,
            borrow_date__isnull=False
        ).exclude(
            truncated_request_date=F('borrow_date')  # Exclude same-day requests (walk-ins)
        ).order_by('-request_date')

        # Walk-in requests: where request_date == borrow_date
        walkin_requests = borrow_info.objects.annotate(
            request_date_only=TruncDate('request_date'),
            borrow_date_only=TruncDate('borrow_date')
        ).filter(
            user=current_user,
            request_date_only=F('borrow_date_only'),
            request_date__isnull=False,
            borrow_date__isnull=False
        ).order_by('-request_date')

        walkin_requests = borrow_info.objects.extra(
            select={
                'request_date_only': "DATE(request_date)",
                'borrow_date_only': "DATE(borrow_date)"
            },
            where=["DATE(request_date) = DATE(borrow_date)"]
        ).filter(
            user=current_user,
            request_date__isnull=False,
            borrow_date__isnull=False
        ).order_by('-request_date')

        # Filter pre-book requests by status
        pending_requests = prebook_requests.filter(status='P')
        accepted_requests = prebook_requests.filter(status='A')
        declined_requests = prebook_requests.filter(status='D')
        borrowed_requests = prebook_requests.filter(status='B')
        cancelled_requests = prebook_requests.filter(status='C')
        completed_requests = prebook_requests.filter(status='X')

        # Walk-in request statuses
        walkin_pending_requests = walkin_requests.filter(status='P')
        walkin_accepted_requests = walkin_requests.filter(status='A')
        walkin_declined_requests = walkin_requests.filter(status='D')
        walkin_borrowed_requests = walkin_requests.filter(status='B')
        walkin_cancelled_requests = walkin_requests.filter(status='C')
        walkin_completed_requests = walkin_requests.filter(status='X')

        return render(request, 'mod_borrowing/borrowing_studentViewPreBookRequests.html', {
            'pending_requests': pending_requests,
            'accepted_requests': accepted_requests,
            'declined_requests': declined_requests,
            'borrowed_requests': borrowed_requests,
            'cancelled_requests': cancelled_requests,
            'completed_requests': completed_requests,
            'walkin_pending_requests': walkin_pending_requests,
            'walkin_accepted_requests': walkin_accepted_requests,
            'walkin_declined_requests': walkin_declined_requests,
            'walkin_borrowed_requests': walkin_borrowed_requests,
            'walkin_cancelled_requests': walkin_cancelled_requests,
            'walkin_completed_requests': walkin_completed_requests,
            'lab_config': lab_config,  # Pass borrowing config to template
            'prebook_requests_all': prebook_requests,
        })
    except Exception as e:
        logger.error(f"Unexpected error in borrowing_student_viewPreBookRequestsview: {e}", exc_info=True)
        return render(request, 'error_page.html', {'message': 'An unexpected error occurred. Please try again.'})

@transaction.atomic
@login_required
@lab_permission_required('borrow_items')
def cancel_borrow_request(request):
    try:
        data = json.loads(request.body)
        borrow_id = data.get('borrow_id')

        borrow_entry = borrow_info.objects.get(borrow_id=borrow_id)
        # Only allow cancel if the status is pending
        if borrow_entry.status == 'P':
            borrow_entry.status = 'C'  # Canceled
            borrow_entry.save()
            logger.info(f"Borrow request {borrow_id} canceled by {request.user}")
            return JsonResponse({'success': True, 'message': 'Request successfully canceled.'})
        return JsonResponse({'success': False, 'message': 'Only pending requests can be canceled.'})
    except borrow_info.DoesNotExist:
        logger.error(f"Attempt to cancel non-existent borrow request {borrow_id} by {request.user}", exc_info=True)
        return JsonResponse({'success': False, 'message': 'Request not found.'})

    except Exception as e:
        logger.error(f"Unexpected error in cancel_borrow_request: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': 'An unexpected error occurred.'})

@login_required
@lab_permission_required('borrow_items')
def borrowing_student_WalkInRequestsview(request):
    try:
        if not request.user.is_authenticated:
            return redirect('userlogin')

        current_user = request.user

        # Annotate the request date as truncated date (without time) and exclude records where dates match
        prebook_requests = borrow_info.objects.annotate(
            truncated_request_date=TruncDate('request_date')
        ).filter(
            user=current_user,
            request_date__isnull=False,
            borrow_date__isnull=False
        ).exclude(
            truncated_request_date=F('borrow_date')
        ).order_by('-request_date')

        return render(request, 'mod_borrowing/borrowing_studentViewWalkInRequests.html', {
            'prebook_requests': prebook_requests,
        })
    except Exception as e:
        logger.error(f"Error fetching walk-in requests for {request.user}: {e}", exc_info=True)
        return render(request, 'error_page.html', {'message': 'An unexpected error occurred.'})

@login_required
@lab_permission_required('borrow_items')
def borrowing_student_detailedPreBookRequestsview(request, borrow_id):
    try:
        if not request.user.is_authenticated:
            return redirect('userlogin')

        # Get the borrow_info instance using the borrow_id
        borrow_request = get_object_or_404(borrow_info, borrow_id=borrow_id)

        # Get all the items that were borrowed under this request
        borrowed_items_list = borrowed_items.objects.filter(borrow=borrow_request)
        qr_data = f"{borrow_id}"
        qr_details = f"{borrow_id}"
        qr_code_data_borrowid = generate_qr_code(qr_data, qr_details)

        return render(request, 'mod_borrowing/borrowing_studentDetailedPreBookRequests.html', {
            'borrow_request': borrow_request,
            'borrowed_items': borrowed_items_list,
            'qrcode': qr_code_data_borrowid,
        })
    except Http404:
        logger.warning(f"User {request.user} tried to access non-existent borrow ID {borrow_id}")
        return render(request, 'error_page.html', {'message': 'Borrow request not found.'})

    except Exception as e:
        logger.error(f"Error in borrowing_student_detailedPreBookRequestsview for {request.user}: {e}", exc_info=True)
        return render(request, 'error_page.html', {'message': 'An unexpected error occurred.'})

    
@login_required
@lab_permission_required('borrow_items')
def borrowing_student_detailedWalkInRequestsview(request):
    try: 
        if not request.user.is_authenticated:
            return redirect('userlogin')

        # Get the borrow_info instance using the borrow_id
        borrow_request = get_object_or_404(borrow_info, borrow_id=borrow_id)

        # Get all the items that were borrowed under this request
        borrowed_items_list = borrowed_items.objects.filter(borrow=borrow_request)

        return render(request, 'mod_borrowing/borrowing_studentDetailedWalkInRequests.html', {
            'borrow_request': borrow_request,
            'borrowed_items': borrowed_items_list,
        })
    except Http404:
        logger.warning(f"User {request.user} tried to access non-existent walk-in request {borrow_id}")
        return render(request, 'error_page.html', {'message': 'Walk-in request not found.'})

    except Exception as e:
        logger.error(f"Error in borrowing_student_detailedWalkInRequestsview for {request.user}: {e}", exc_info=True)
        return render(request, 'error_page.html', {'message': 'An unexpected error occurred.'})

    
@transaction.atomic
@login_required
@lab_permission_required('view_booking_requests')
def borrowing_labcoord_prebookrequests(request):
    try:
        selected_laboratory_id = request.session.get('selected_lab')
        if not request.user.is_authenticated:
            logger.warning(f"Unauthorized attempt to access borrowing_labcoord_prebookrequests by {request.user}")
            return redirect('userlogin')

        # Get selected status from the GET request, default to 'P' (Pending)
        selected_status = request.GET.get('status', 'P')

        # Filter borrowing requests based on the selected status
        if selected_status == 'all':
            borrowing_requests = borrow_info.objects.filter(laboratory_id=selected_laboratory_id).order_by('request_date')
        else:
            borrowing_requests = borrow_info.objects.filter(laboratory_id=selected_laboratory_id, status=selected_status).order_by('request_date')

        # Check if there are new requests that have not yet been notified
        new_requests = borrow_info.objects.filter(
            laboratory_id=selected_laboratory_id, 
            status='P', 
            request_date__gte=timezone.now()-timezone.timedelta(minutes=5), 
            notification_sent=False
        )

        # Send notifications for new borrow requests
        for borrow_request in new_requests:
            message = f"New borrow request for borrow ID: {borrow_request.borrow_id}"
            create_notification(request.user, message)

            # Mark the borrow request as having been notified
            borrow_request.notification_sent = True
            borrow_request.save()

        # Handle POST request for approval or rejection of borrow requests
        if request.method == 'POST':
            # Get borrow_id and action from form submission
            borrow_id = request.POST.get('borrow_id')
            action = request.POST.get('action')

            # Retrieve the borrow_info object
            borrow_request = get_object_or_404(borrow_info, borrow_id=borrow_id)

            # Update status based on the action
            if action == 'approve':
                borrow_request.status = 'A'
                borrow_request.approved_by = request.user  # Set the approving lab coordinator
                borrow_request.save()
                messages.success(request, f"Borrow request {borrow_id} has been approved.")
            elif action == 'reject':
                borrow_request.status = 'D'
                borrow_request.approved_by = request.user
                borrow_request.save()
                messages.success(request, f"Borrow request {borrow_id} has been declined.")

            return redirect('borrowing_labcoord_prebookrequests')

        # Get updated notifications count and list
        notifications = get_notifications(request)

        return render(request, 'mod_borrowing/borrowing_labcoord_prebookrequests.html', {
            'borrowing_requests': borrowing_requests,
            'selected_status': selected_status,  # Pass the selected status to the template
            'notifications': notifications,
        })
    
    except Exception as e:
        logger.error(f"Error in borrowing_labcoord_prebookrequests for {request.user}: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred while fetching booking requests.")
        return redirect('home')

@transaction.atomic
@login_required
@lab_permission_required('configure_borrowing')
def borrowing_labcoord_borrowconfig(request):
    try: 
        selected_laboratory_id = request.session.get('selected_lab')
        items = item_description.objects.filter(laboratory_id=selected_laboratory_id, is_disabled=False)

        # Fetch the current total quantity from item_inventory for each item
        items_with_qty = []
        for item in items:
            # Retrieve and sum all qty values for each item from item_inventory
            total_qty = item_inventory.objects.filter(item_id=item.item_id).aggregate(total_qty=models.Sum('qty'))['total_qty']
            current_quantity = total_qty if total_qty else 0  # Default to 0 if no inventory entry exists

            item.current_quantity = current_quantity  # Attach current quantity to item
            items_with_qty.append(item)

        item_types_list = item_types.objects.filter(laboratory_id=selected_laboratory_id)
        
        lab, created = borrowing_config.objects.get_or_create(laboratory_id=selected_laboratory_id)

        # Annotate each item type to check if all items under it are borrowable
        # for type in item_types_list:
        #     type.all_items_borrowable = item_description.objects.filter(itemType_id=type.itemType_id, allow_borrow=False).count() == 0
        #     type.all_items_consumable = item_description.objects.filter(itemType_id=type.itemType_id, is_consumable=False).count() == 0

        if request.method == 'POST':
            if 'lab_config_form' in request.POST:
                lab.allow_walkin = 'allow_walkin' in request.POST
                lab.allow_prebook = 'allow_prebook' in request.POST
                lab.prebook_lead_time = request.POST.get('prebook_lead_time') or 0
                lab.allow_shortterm = 'allow_shortterm' in request.POST
                lab.allow_longterm = 'allow_longterm' in request.POST
                lab.save()
                messages.success(request, 'Borrowing configurations updated successfully!')
                return redirect('borrowing_labcoord_borrowconfig')

            elif 'borrow_config_form' in request.POST:
                borrowMode = request.POST.get('borrowMode')
                if borrowMode == 'item':
                    allowed_items = request.POST.getlist('borrow_item')  # Items explicitly checked
                    is_consumable_list = request.POST.getlist('is_consumable')
                    
                    # Update logic
                    items.update(allow_borrow=0, is_consumable=0)

                    # Handle individual items: Set allow_borrow=True and is_consumable=True for explicitly checked items
                    if allowed_items:
                        item_description.objects.filter(item_id__in=allowed_items).update(allow_borrow=True)
                    if is_consumable_list:
                        item_description.objects.filter(item_id__in=is_consumable_list).update(is_consumable=True)

                    # Validate and update qty_limit for each item
                    for item in items_with_qty:
                        lead_time_prep = request.POST.get(f'lead_time_prep_{item.item_id}')
                        if lead_time_prep is not None:
                            lead_time_prep = int(lead_time_prep) if lead_time_prep else None
                            item_description.objects.filter(item_id=item.item_id).update(lead_time_prep=lead_time_prep)

                        qty_limit = request.POST.get(f'qty_limit_{item.item_id}')
                        if qty_limit is not None:
                            qty_limit = int(qty_limit) if qty_limit else None
                            
                            # Check if qty_limit exceeds the current quantity
                            if qty_limit and qty_limit > item.current_quantity:
                                messages.error(request, f"Quantity limit for item '{item.item_name}' exceeds the available quantity of {item.current_quantity}.")
                                return redirect('borrowing_labcoord_borrowconfig')  # Redirect with error
                            # Save the qty_limit
                            # item.qty_limit = qty_limit
                            # item.save()

                            item_description.objects.filter(item_id=item.item_id).update(qty_limit=qty_limit)

                
                        

                else:
                    allowed_item_types = request.POST.getlist('borrow_item_type')  # Item types explicitly checked
                    is_consumable_type_list = request.POST.getlist('is_consumable_type')

                    print(allowed_item_types)
                    print(is_consumable_type_list)
                    
                    # Handle item types: Set allow_borrow=True for all items under the checked item types
                    if allowed_item_types:
                        item_description.objects.filter(itemType_id__in=allowed_item_types).update(allow_borrow=True)
                        for type in item_types_list:
                            # If the item type is checked in the form, mark all items under this type as borrowable and update consumable status
                            if str(type.itemType_id) in allowed_item_types:
                                item_description.objects.filter(itemType_id=type.itemType_id).update(allow_borrow=True)
                            # else:
                            #     # If unchecked, ensure items under this type are set to allow_borrow=False and is_consumable=False
                            #     item_description.objects.filter(itemType_id=type.itemType_id).update(allow_borrow=False)
                    
                    if is_consumable_type_list:
                        item_description.objects.filter(itemType_id__in=is_consumable_type_list).update(is_consumable=True)
                        for type in item_types_list:
                            # If the item type is checked in the form, mark all items under this type as borrowable and update consumable status
                            if str(type.itemType_id) in is_consumable_type_list:
                                item_description.objects.filter(itemType_id=type.itemType_id).update(is_consumable=True)
                            # else:
                            #     item_description.objects.filter(itemType_id=type.itemType_id).update(is_consumable=False)
                        
                messages.success(request, "Borrowing configuration updated successfully!")
                return redirect('borrowing_labcoord_borrowconfig')
            
            elif 'add_question_form' in request.POST:
                question_text = request.POST.get('question_text')
                input_type = request.POST.get('input_type')
                borrowing_mode = request.POST.get('borrowing_mode')  # New: borrow mode (walk-in, pre-book, both)
                dropdown_choices = request.POST.get('dropdown_choices', '').split(',') if input_type == 'dropdown' else None

                lab.add_question(question_text, input_type, borrowing_mode, dropdown_choices)
                messages.success(request, 'Question added successfully!')
                return redirect('borrowing_labcoord_borrowconfig')

            elif 'update_question_form' in request.POST:
                index = int(request.POST.get('question_index'))
                question_text = request.POST.get('question_text')
                input_type = request.POST.get('input_type')
                borrowing_mode = request.POST.get('borrowing_mode')  # New: borrow mode (walk-in, pre-book, both)
                dropdown_choices = request.POST.get('dropdown_choices', '').split(',') if input_type == 'dropdown' else None

                lab.update_question(index, question_text, input_type, borrowing_mode, dropdown_choices)
                messages.success(request, 'Question updated successfully!')
                return redirect('borrowing_labcoord_borrowconfig')

            elif 'remove_question_form' in request.POST:
                index = int(request.POST.get('question_index'))

                lab.remove_question(index)
                messages.success(request, 'Question removed successfully!')
                return redirect('borrowing_labcoord_borrowconfig')

        return render(request, 'mod_borrowing/borrowing_labcoord_borrowconfig.html', {
            'items': items_with_qty,
            'item_types_list': item_types_list,
            'lab': lab,
            'questions': lab.get_questions()  # Get the questions to display them
        })
    except Exception as e:
        logger.error(f"Error in borrowing_labcoord_borrowconfig for {request.user}: {e}", exc_info=True)
        messages.error(request, "An error occurred while updating borrowing configurations.")
        return redirect('home')

@transaction.atomic
@login_required
@lab_permission_required('view_booking_requests')
def borrowing_labcoord_detailedPrebookrequests(request, borrow_id):
    # Fetch the borrow request and associated borrowed items
    try:
        selected_laboratory_id = request.session.get('selected_lab')
        borrow_request = get_object_or_404(borrow_info, borrow_id=borrow_id, laboratory_id=selected_laboratory_id)
        borrowed_items_list = borrowed_items.objects.filter(borrow=borrow_request)
        
        # Prepare original quantities in JSON format for JavaScript
        borrowed_items_json = json.dumps(
            [{'id': item.id, 'qty': item.qty} for item in borrowed_items_list],
            cls=DjangoJSONEncoder
        )

        if request.method == 'POST':
            action = request.POST.get('action')
            edited = request.POST.get('edited')  # Check if any edits were made by checking 'edited' input

            # Decline action
            if action == 'decline':
                decline_reason = request.POST.get('decline_reason')
                if decline_reason:
                    borrow_request.status = 'D'  # Declined
                    borrow_request.remarks = decline_reason
                    borrow_request.approved_by = request.user
                    borrow_request.save()
                    messages.success(request, 'The request has been declined.')
                    logger.info(f"Borrow request {borrow_id} declined by {request.user}")
                else:
                    messages.error(request, 'Please provide a reason for declining.')

            # Accept action
            elif action == 'accept':
                # Check if any quantities were edited
                print('pass')
                if edited == '1':
                    # Require an edit reason if quantities were modified
                    edit_reason = request.POST.get('edit_reason')
                    if not edit_reason:
                        messages.error(request, 'Please provide a reason for the quantity update.')
                        return redirect('borrowing_labcoord_detailedPrebookrequests', borrow_id=borrow_id)
                    
                    # Update each item's quantity if it was changed
                    for item in borrowed_items_list:
                        new_qty = int(request.POST.get(f'qty_{item.id}', item.qty))  # Retrieve new quantity
                        if new_qty != item.qty:  # Check if quantity was updated
                            item.qty = new_qty  # Update quantity in the database
                            item.save()
                    
                    borrow_request.remarks = edit_reason  # Save the edit reason
                    messages.success(request, 'The request was accepted with updated quantities.')

                else:
                    borrow_request.remarks = "Accepted"
                    messages.success(request, 'The request has been accepted without quantity updates.')
                
                borrow_request.approved_by = request.user
                borrow_request.status = 'A'  # Set status to Accepted
                borrow_request.save()

            return redirect('borrowing_labcoord_detailedPrebookrequests', borrow_id=borrow_id)

        # Only show action buttons if the request is pending, accepted, or declined (to allow modifications)
        show_action_buttons = borrow_request.status in ['P', 'A', 'D']

        return render(request, 'mod_borrowing/borrowing_labcoord_DetailedPrebookRequests.html', {
            'borrow_request': borrow_request,
            'borrowed_items_list': borrowed_items_list,
            'borrowed_items_json': borrowed_items_json,
            'show_action_buttons': show_action_buttons,
        })
    
    except Exception as e:
        logger.error(f"Error in borrowing_labcoord_detailedPrebookrequests for {request.user}: {e}", exc_info=True)
        messages.error(request, "An error occurred while processing the request.")
        return redirect('borrowing_labcoord_prebookrequests')

@transaction.atomic
@login_required
@lab_permission_required('return_item')
def return_borrowed_items(request):
    b_user_id = request.GET.get('b_user_id', '')  # Fetch b_user_id from GET request
    borrow_entries = None
    selected_laboratory_id = request.session.get('selected_lab')

    if b_user_id:
        try:
            borrow_entries = borrow_info.objects.filter(
                user__personal_id=b_user_id, 
                laboratory_id=selected_laboratory_id, 
                status='B'
            ).select_related('user')

            #debugging
            print(f"Personal ID entered: {b_user_id}")
            print(f"Query Executed: {borrow_entries.query}")
            print(f"Results Found: {borrow_entries.exists()}")

            if not borrow_entries.exists():
                messages.error(request, "No active borrow requests found for this ID.")
                return redirect('return_borrowed_items')

        except borrow_info.DoesNotExist:
            messages.error(request, "Invalid Personal ID.")
            return redirect('return_borrowed_items')

    if request.method == 'POST' and 'return_items' in request.POST:
        try:
            for borrow_entry in borrow_entries:
                borrowed_items_list = borrowed_items.objects.filter(borrow=borrow_entry, item__is_consumable=False)
                consumed_items_list = borrowed_items.objects.filter(borrow=borrow_entry, item__is_consumable=True)

                for item in borrowed_items_list:
                    returned_all = request.POST.get(f'returned_all_{item.item.item_id}', False) == 'on'
                    qty_returned = int(request.POST.get(f'return_qty_{item.item.item_id}', 0))
                    hold_clearance = request.POST.get(f'hold_clearance_{item.item.item_id}', False) == 'on'
                    remarks = request.POST.get(f'remarks_{item.item.item_id}', '').strip()
                    amount_to_pay = request.POST.get(f'amount_to_pay_{item.item.item_id}', 0)

                    if returned_all:
                        item.returned_qty = item.qty
                    else:
                        item.returned_qty = qty_returned
                    item.save()

                    if hold_clearance and remarks:
                        reported_items.objects.create(
                            borrow=borrow_entry,
                            item=item.item,
                            qty_reported=item.qty - item.returned_qty,
                            report_reason=remarks,
                            amount_to_pay=amount_to_pay or 0,
                            laboratory_id=selected_laboratory_id,
                            user=borrow_entry.user
                        )

                for consumed_item in consumed_items_list:
                    consumed_item.returned_qty = consumed_item.qty
                    consumed_item.save()

                # If all borrowed and consumed items are returned, mark as completed
                if all(item.qty == item.returned_qty for item in borrowed_items_list) and \
                        all(item.qty == item.returned_qty for item in consumed_items_list):
                    borrow_entry.status = 'X'
                    borrow_entry.save()

            messages.success(request, 'Successfully Returned Items')
            return redirect('return_borrowed_items')
        except Exception as e:
            logger.error(f"Error while returning items for Personal ID {b_user_id}: {e}", exc_info=True)
            messages.error(request, "An error occurred while returning items.")


    return render(request, 'mod_borrowing/borrowing_return_borrowed_items.html', {
        'borrow_entries': borrow_entries,
        'b_user_id': b_user_id,
    })


@transaction.atomic
@login_required
@lab_permission_required('view_borrowed_items')
def borrowing_labtech_prebookrequests(request):
    try:
        selected_laboratory_id = request.session.get('selected_lab')
        today = date.today()

        today_borrows = borrow_info.objects.filter(borrow_date=today, status='A', laboratory_id=selected_laboratory_id).select_related('user')
        future_borrows = borrow_info.objects.filter(borrow_date__gt=today, status='A', laboratory_id=selected_laboratory_id).select_related('user')
        past_borrows = borrow_info.objects.filter(borrow_date__lt=today, status='A', laboratory_id=selected_laboratory_id).select_related('user')
        
        cancelled_borrows = borrow_info.objects.filter(status='L', laboratory_id=selected_laboratory_id).select_related('user')
        borrowed_borrows = borrow_info.objects.filter(status='B', laboratory_id=selected_laboratory_id).select_related('user')
        accepted_borrows = borrow_info.objects.filter(status__in=['A', 'B', 'L', 'X', 'Y'], laboratory_id=selected_laboratory_id).order_by('-request_date').select_related('user')

        separated_future_borrows = []

        for borrow in future_borrows:
            borrowed_items_qs = borrowed_items.objects.filter(borrow=borrow).select_related('item')

            for borrowed_item in borrowed_items_qs:
                lead_time_days = borrowed_item.item.lead_time_prep
                if lead_time_days and lead_time_days > 0:
                    prep_date = borrow.borrow_date - timedelta(days=lead_time_days)

                    separated_future_borrows.append({
                        "borrow_id": borrow.borrow_id,
                        "prep_date": prep_date,
                        "item_name": borrowed_item.item.item_name,
                        "borrow_date": borrow.borrow_date,
                        "status": borrow.get_status_display(),
                        "qty": borrowed_item.qty,
                        "unit": borrowed_item.unit,
                    })

        separated_future_borrows.sort(key=lambda x: x["prep_date"])

        if request.method == 'POST':
            borrow_id = request.POST.get('borrow_id')
            action = request.POST.get('action')
            remarks = request.POST.get('remarks', '')

            borrow_entry = get_object_or_404(borrow_info, borrow_id=borrow_id)
            
            if action == 'borrowed':
                borrow_entry.status = 'B'
            elif action == 'cancel':
                borrow_entry.status = 'L'
                borrow_entry.remarks = remarks  
            borrow_entry.save()

            return redirect('borrowing_labtech_prebookrequests')

        highlight_borrow_id = request.GET.get('highlight', '')
        
        return render(request, 'mod_borrowing/borrowing_labtech_prebookrequests.html', {
            'today_borrows': today_borrows,
            'future_borrows': future_borrows,
            'past_borrows': past_borrows,
            'cancelled_borrows': cancelled_borrows,
            'accepted_borrows': accepted_borrows,
            'borrowed_borrows': borrowed_borrows,
            'separated_future_borrows': separated_future_borrows,  
            'today': today
        })
    except Exception as e:
        logger.error(f"Error fetching borrowing requests for {request.user}: {e}", exc_info=True)
        messages.error(request, "An error occurred while fetching borrow requests.")
        return redirect('home')

@transaction.atomic
@login_required
@lab_permission_required('view_borrowed_items')
def borrowing_labtech_detailedprebookrequests(request, borrow_id):
    borrow_entry = get_object_or_404(borrow_info, borrow_id=borrow_id)
    borrowed_items1 = borrowed_items.objects.filter(borrow=borrow_entry)

    inventory_items_map = {}
    for item in borrowed_items1:
        inventory_items_map[item.id] = item_inventory.objects.filter(item=item.item)

    print(inventory_items_map)

    today = date.today()

    borrowed_items_json = json.dumps(
        [{'id': item.id, 'qty': item.qty} for item in borrowed_items1],
        cls=DjangoJSONEncoder
    )

    if request.method == 'POST':
        try:
            action = request.POST.get('action')
            edited = request.POST.get('edited')
            print('pass', )

            if action == 'borrowed':
                if edited == '1':  # Check if any edits were made
                    edit_reason = request.POST.get('edit_reason')
                    if not edit_reason:
                        messages.error(request, 'Please provide a reason for the quantity update.')
                        return redirect('borrowing_labtech_detailedprebookrequests', borrow_id=borrow_id)
                    
                    for item in borrowed_items1:
                        new_qty = request.POST.get(f'qty_{item.id}')
                        if new_qty and int(new_qty) != item.qty:
                            item.qty = int(new_qty)
                            item.save()
                    borrow_entry.remarks = edit_reason
                    messages.success(request, 'The borrow request has been marked as borrowed with updated quantities.')
                else:
                    messages.success(request, 'The borrow request has been marked as borrowed.')
                
                borrow_entry.status = 'B'  # Mark as borrowed
                borrow_entry.save()

                # ================== Inventory Item Selection ==================
                for item in borrowed_items1:
                    item_expity_type = item.item.expiry_type
                    if item.unit == "pcs":
                        inventory_item_ids = request.POST.getlist(f'inventory_items_{item.id}')

                        if len(inventory_item_ids) != item.qty:
                            messages.error(request, f"Must select {item.qty} inventory items for '{item.item.item_name}'")
                            return redirect('borrowing_labtech_detailedprebookrequests', borrow_id=borrow_id)

                        selected_inventory_items = []
                        for inv_id in inventory_item_ids:
                            if inv_id:
                                inv_item = get_object_or_404(item_inventory, inventory_item_id=inv_id)
                                inv_item.uses += 1
                                inv_item.save()

                                if (item_expity_type == 'D'): 
                                    exp_record = get_object_or_404(item_expirations, inventory_item=inv_item)
                                    if exp_record.remaining_uses is not None and exp_record.remaining_uses > 0:
                                        exp_record.remaining_uses -= 1
                                        exp_record.save()
                                    else:
                                        messages.error(request, f"Inventory Item {inv_id} has no remaining uses.")
                                        return redirect('borrowing_labtech_detailedprebookrequests', borrow_id=borrow_id)

                                selected_inventory_items.append(inv_id)

                        item.inventory_item = ', '.join(selected_inventory_items)
                        item.save()

                    # messages.success(request, 'Inventory items recorded successfully.')
                
            elif action == 'cancel':
                cancel_reason = request.POST.get('cancel_reason')
                borrow_entry.status = 'L'  # Mark as cancelled
                borrow_entry.remarks = cancel_reason
                borrow_entry.save()
                messages.success(request, 'The borrow request has been cancelled.')

            return redirect('borrowing_labtech_prebookrequests')
        
        except Exception as e:
            logger.error(f"Error processing Borrow ID {borrow_id}: {e}", exc_info=True)
            messages.error(request, 'An error occurred while processing your request.')
            return redirect('borrowing_labtech_detailedprebookrequests', borrow_id=borrow_id)
        
    return render(request, 'mod_borrowing/borrowing_labtech_detailedprebookrequests.html', {
        'borrow_entry': borrow_entry,
        'borrowed_items': borrowed_items1,
        'borrowed_items_json': borrowed_items_json,
        'today': today,
        'inventory_items_map': inventory_items_map,
    })

def validate_borrow_id(request):
    try:
        borrow_id = request.GET.get('borrow_id', '').strip()
        if borrow_id:
            is_valid = borrow_info.objects.filter(borrow_id=borrow_id).exists()
            messages.error(request, 'Invalid Borrow ID')
            return JsonResponse({'valid': is_valid})
        return JsonResponse({'valid': False})
    except Exception as e:
        logger.error(f"Error validating borrow ID: {e}", exc_info=True)
        return JsonResponse({'valid': False, 'error': str(e)}, status=500)


#CLEARANCE
@login_required
@lab_permission_required('view_own_clearance')
def clearance_view(request):
    return render(request, 'mod_clearance/clearance.html')

@login_required
@lab_permission_required('view_own_clearance')
def clearance_student_viewClearance(request):
    try:
        # Get the currently logged-in user
        user = request.user
        # Debugging output to verify the user instance
        if not user.is_authenticated:
            return render(request, 'mod_clearance/student_viewClearance.html', {'error': 'User is not authenticated.'})

        # Use the user instance's ID for querying
        selected_laboratory_id = request.session.get('selected_lab')
        current_user = request.user

        try:

            reports = reported_items.objects.filter(user=current_user, laboratory_id=selected_laboratory_id)

                # Handle the filter by status
            status = request.GET.get('status', 'All')
            if status != 'All':
                if status == 'Cleared':
                    reports = reports.filter(status=0)  # Clear status
                elif status == 'Pending':
                    reports = reports.filter(status=1)  # Pending status
       

        except Exception as e: 
            reports = reported_items.objects.none()  # If there's an error, return no reports
            print(f"Error fetching reports: {e}")  # Debugging output for error tracking

        context = {
            'reports': reports,
        }
        return render(request, 'mod_clearance/student_viewClearance.html', context)
    except Exception as e:
        reports = reported_items.objects.none()
        logger.error(f"Error fetching reports for {user}: {e}", exc_info=True)
        return redirect('home')


@login_required
@lab_permission_required('view_own_clearance')
def clearance_student_viewClearanceDetailed(request, report_id):
    # Get the currently logged-in user
    try:
        user = request.user
        selected_laboratory_id = request.session.get('selected_lab')
        lab = get_object_or_404(laboratory, laboratory_id=selected_laboratory_id)
        report = get_object_or_404(reported_items, report_id=report_id, user=user)
        borrow = report.borrow  # Access related borrow_info directly from report

        # Retrieve all reported items for the specific borrow entry
        report_details = reported_items.objects.filter(borrow=borrow, user=user)

        # Context for rendering details of borrow and reports
        context = {
            'report': report,                  # Main report entry
            'borrow_details': borrow,          # Borrow details
            'report_details': report_details,   # All reported items for this borrow
            'laboratory_name': lab.name
        }
        return render(request, 'mod_clearance/student_viewClearanceDetailed.html', context)
    except Exception as e:
        logger.error(f"Error fetching detailed report {report_id} for {user}: {e}", exc_info=True)
        messages.error(request, "An error occurred while retrieving report details.")
        return redirect('clearance_student_viewClearance')


@login_required
@lab_permission_required('view_student_clearance')
def clearance_labtech_viewclearance(request):
    selected_laboratory_id = request.session.get('selected_lab')
    users = user.objects.filter(is_deactivated=False)
    items = item_description.objects.filter(is_disabled=0, allow_borrow=1)

    if request.method == 'POST':
        try:
            selected_user_id = request.POST.get('user')
            item_id = request.POST.get('item_name')
            reason = request.POST.get('reason')
            amount = request.POST.get('amount')
            quantity = request.POST.get('quantity')

            selected_user = user.objects.get(user_id=selected_user_id)
            item_obj = get_object_or_404(item_description, item_id=item_id)

            # Save the manual entry with the selected user
            reported_items.objects.create(
                laboratory_id=selected_laboratory_id,
                borrow=None,
                user=selected_user,  # Save the user object
                item=item_obj,
                qty_reported=quantity,
                report_reason=reason,
                amount_to_pay=amount,
                status=1,
            )

            messages.success(request, "Manual clearance added successfully!")
        except Exception as e:
            messages.error(request, f"Error adding clearance: {e}")
            logger.error(f"Error adding clearance: {e}", exc_info=True)
    
    reports = reported_items.objects.filter(
       laboratory_id=selected_laboratory_id
    )
    status_filter = request.GET.get('status', 'All')
    if status_filter == 'Cleared':
        reports = reports.filter(status=0)
    elif status_filter == 'Pending':
        reports = reports.filter(status=1)

    report_data = []

    try:
        for report in reports:
                borrow_info_obj = report.borrow
                user_obj = borrow_info_obj.user if borrow_info_obj else report.user

                report_data.append({
                    'report_id': report.report_id,
                    'borrow_id': borrow_info_obj.borrow_id if borrow_info_obj else "Manual Entry",
                    'user_name': f"{user_obj.firstname} {user_obj.lastname}" if user_obj else "Unknown",
                    'id_number': user_obj.personal_id if user_obj else "N/A",
                    'item_name': report.item.item_name,
                    'reason': report.report_reason,
                    'amount_due': report.amount_to_pay,
                    'status': 'Pending' if report.status == 1 else 'Cleared',
                    'quantity': report.qty_reported,
                })


        context = {
            'reports': report_data,
            'users': users,
            'items': items,
        }
        
        return render(request, 'mod_clearance/labtech_viewclearance.html', context)

    except Exception as e:
        logger.error(f"Error processing report data: {e}", exc_info=True)
        messages.error(request, "Error processing report data.")
        return redirect('home')

@transaction.atomic
@login_required
@lab_permission_required('view_student_clearance')
def clearance_labtech_viewclearanceDetailed(request, report_id):
    try:
        # Get the reported item by ID
        report = get_object_or_404(reported_items, report_id=report_id)
        selected_laboratory_id = request.session.get('selected_lab')
        lab = get_object_or_404(laboratory, laboratory_id=selected_laboratory_id)

        # Prepare report data in a similar structure as in 'clearance_labtech_viewclearance'
        borrow_info_obj = report.borrow
        user_obj = borrow_info_obj.user if borrow_info_obj else report.user

        report_data = {
            'report_id': report.report_id,
            'borrow_id': borrow_info_obj.borrow_id if borrow_info_obj else "Manual Entry",
            'user_name': f"{user_obj.firstname} {user_obj.lastname}" if user_obj else "Unknown",
            'id_number': user_obj.personal_id if user_obj else "N/A",
            'item_name': report.item.item_name,
            'reason': report.report_reason,
            'amount_due': report.amount_to_pay,
            'status': 'Pending' if report.status == 1 else 'Cleared',
            'quantity': report.qty_reported,
            'remarks': report.remarks if report.remarks else '',
     
        }

        if request.method == 'POST':
            # Handle remarks submission and marking as cleared
            remarks = request.POST.get('remarks', '').strip()
            if remarks:
                report.remarks = remarks
            
            # Update the status to Cleared
            report.status = 0  # Assuming status 0 means Cleared
            report.save()

            # Redirect to the same page or to the view clearance page
            logger.info(f"Labtech cleared report {report_id} with remarks: {remarks}")
            return HttpResponseRedirect(request.path_info)

        # Pass the report_data to the context for rendering
        context = {
            'report_data': report_data,
            'laboratory_name': lab.name,
        }

        return render(request, 'mod_clearance/labtech_viewclearanceDetailed.html', context)
    except Exception as e:
        logger.error(f"Error fetching detailed clearance report {report_id}: {e}", exc_info=True)
        messages.error(request, "Error fetching report details.")
        return redirect('clearance_labtech_viewclearance')




# lab reserv ================================================================= 
@login_required
@lab_permission_required('reserve_laboratory')
def lab_reservation_view(request):
    return render(request, 'mod_labRes/lab_reservation.html')

@login_required
@lab_permission_required('reserve_laboratory')
def lab_reservation_preapproval(request):
    selected_laboratory_id = request.session.get('selected_lab')
    try:
        reservation_config_obj = get_object_or_404(reservation_config, laboratory_id=selected_laboratory_id)
        if request.method == 'POST':
            contact_name = request.POST.get('contact_name')
            contact_email = request.POST.get('contact_email')
            num_people = request.POST.get('num_people')
            purpose = request.POST.get('purpose')

            uploaded_form = request.FILES.get('approval_form')
            if reservation_config_obj.approval_form and not uploaded_form:
                messages.error(request, "Please upload the required approval form.")
                return redirect('lab_reservation_preapproval')

            # Save reservation as pending
            reservation = laboratory_reservations.objects.create(
                user=request.user,
                room=None,  # Room selection will happen after approval
                start_date=None,
                start_time=None,
                end_time=None,
                contact_name=contact_name,
                contact_email=contact_email,
                num_people=num_people,
                purpose=purpose,
                status='P',  # Set to pending for approval
                filled_approval_form=uploaded_form,
                laboratory_id=selected_laboratory_id

            )
            
            messages.success(request, "Your reservation is now pending approval. You’ll be notified once it’s approved.")
            return redirect('lab_reservation_detail', reservation.reservation_id)  # Redirect to home or any page you want

        return render(request, 'mod_labRes/lab_reservation_preapproval.html', {
            'reserv_config': reservation_config_obj,
        })
    except Exception as e:
        logger.error(f"Error in lab_reservation_preapproval: {e}", exc_info=True)
        messages.error(request, "An error occurred while processing your reservation. Please try again later.")
        return redirect('home')

@login_required
@lab_permission_required('reserve_laboratory')
def lab_reservation_student_reserveLabChooseRoom(request):
    selected_laboratory_id = request.session.get('selected_lab')
    try:
        reservation_config_obj = reservation_config.objects.get(laboratory_id=selected_laboratory_id)

        res_id = None
        # Check if the lab requires approval and if the user has an approved reservation
        if reservation_config_obj.require_approval:
            reservation_data = request.session.get('reservation_id')
            if reservation_data:
                res_id = reservation_data.get('res_id')
            else:
                res_id = None

            if not res_id:
                return redirect('lab_reservation_preapproval')  

        today = timezone.localtime().date()
        min_reservation_date = today + timedelta(days=reservation_config_obj.leadtime)

        reservation_date = request.GET.get('reservationDate')
        error_message = None

        if reservation_date:
            reservation_date = datetime.strptime(reservation_date, '%Y-%m-%d').date()
            if reservation_date < min_reservation_date:
                error_message = f"Selected date must be at least {reservation_config_obj.leadtime} days from today."
        else:
            error_message = "Please select a reservation date."

        if error_message:
            context = {
                'res_id': res_id,
                'error_message': error_message,
                'min_reservation_date': min_reservation_date,
            }
            return render(request, 'mod_labRes/lab_reservation_studentReserveLabChooseRoom.html', context)

        
        capacity_filter = request.GET.get('capacityFilter', '')

        # Fetch rooms based on selected lab and not disabled
        rooms_query = rooms.objects.filter(laboratory_id=selected_laboratory_id, is_disabled=False, is_reservable=True)

        # Filter by room capacity if provided
        if capacity_filter:
            rooms_query = rooms_query.filter(capacity__gte=capacity_filter)

        # Fetch reservation configuration for the laboratory
        start_time = reservation_config_obj.start_time
        end_time = reservation_config_obj.end_time

        # Determine time slots based on reservation type
        if reservation_config_obj.reservation_type == 'class':
            time_slots = [
                '7:30-9:00',
                '9:15-10:45',
                '11:00-12:30',
                '12:45-14:15',
                '14:30-16:00',
                '16:15-17:45',
                '18:00-19:30',
            ]
        elif reservation_config_obj.reservation_type == 'hourly':
            # Generate hourly intervals between start_time and end_time
            time_slots = []
            current_time = start_time
            while current_time < end_time:
                next_time = (datetime.combine(today, current_time) + timedelta(hours=1)).time()
                time_slots.append(f"{current_time.strftime('%H:%M')}-{next_time.strftime('%H:%M')}")
                current_time = next_time

        # Fetch existing reservations for the selected date
        existing_reservations = laboratory_reservations.objects.filter(
        room__laboratory_id=selected_laboratory_id, start_date=reservation_date, room__is_disabled=0)

        room_reservation_times = {room.room_id: [] for room in rooms_query}

        # Populate the room_reservation_times dictionary
        for reservation in existing_reservations:
            start_time_str = reservation.start_time.strftime('%H:%M')
            end_time_str = reservation.end_time.strftime('%H:%M')
            room_reservation_times[reservation.room.room_id].append(f"{start_time_str}-{end_time_str}")

        # Sort the reservation times for each room by start time
        for room_id, times in room_reservation_times.items():
            room_reservation_times[room_id] = sorted(times, key=lambda x: datetime.strptime(x.split('-')[0], '%H:%M'))

        print('room res: ', room_reservation_times)

        # Get the day of the week for the selected date
        day_of_week = reservation_date.strftime('%A')  # E.g., 'Monday', 'Tuesday', etc.

        room_availability = {}
        for room in rooms_query:
            availability = {}
            
            # Load and parse the blocked times (assuming it's stored as a JSON string)
            blocked_times = json.loads(room.blocked_time) if room.blocked_time else {}

            for start in time_slots:
                time_key = f"{start}"

                # Check if the time slot is blocked for the selected day of the week
                is_blocked = time_key in blocked_times.get(day_of_week, [])

                # Split the time slot into start and end times
                start_time_str, end_time_str = start.split('-')
                start_time_obj = datetime.strptime(start_time_str, '%H:%M').time()
                end_time_obj = datetime.strptime(end_time_str, '%H:%M').time()

                # Check if the time slot is reserved
                reserved = existing_reservations.filter(
                    room=room,
                    start_time__lt=end_time_obj,
                    end_time__gt=start_time_obj,
                    status='R'
                ).exists()

                # Get the reservation info for the current time slot
                reservation_info = []
                for time in room_reservation_times[room.room_id]:
                    ss_start_time_str, ss_end_time_str = time.split('-')

                    # Convert the reservation times to time objects for comparison
                    ss_start_time_obj = datetime.strptime(ss_start_time_str, '%H:%M').time()
                    ss_end_time_obj = datetime.strptime(ss_end_time_str, '%H:%M').time()

                    # Check if the reservation overlaps with the requested time slot
                    if (ss_start_time_obj < end_time_obj and ss_end_time_obj > start_time_obj):
                        reservation_info.append(time)
                
                print(reservation_info)

                # Mark as 'red' if reserved or blocked
                if reserved or is_blocked:
                    availability[time_key] = {
                        'color': 'red',
                        'reservation_info': reservation_info  # Store specific reservation times
                    }  # Unavailable
                else:
                    availability[time_key] = {
                        'color': 'green',
                        'reservation_info': None  # No reservation info
                    }

            room_availability[room.room_id] = availability
            # print(room_availability)

    
        room_tables = {}
        for room in rooms_query:
            tables = RoomTable.objects.filter(room=room)
            room_tables[room.room_id] = tables

        
        context = {
            'rooms': rooms_query,
            'time_slots': time_slots,
            'room_availability': room_availability,
            'reservation_date': reservation_date,
            'min_reservation_date': min_reservation_date,
            'capacity_filter': capacity_filter,
            'error_message': error_message,
            'reservation_config_obj': reservation_config_obj,
            'res_id': res_id,
            'room_tables': room_tables 
        }

        return render(request, 'mod_labRes/lab_reservation_studentReserveLabChooseRoom.html', context)
    except Exception as e:
        logger.error(f"Error in lab_reservation_student_reserveLabChooseRoom: {e}", exc_info=True)
        messages.error(request, "An error occurred while processing your reservation. Please try again later.")
        return redirect('home')

def get_room_tables(request):
    try:
        room_id = request.GET.get('room_id')
        selected_date = request.GET.get('selected_date')
        selected_start_time = request.GET.get('selected_start_time')
        selected_end_time = request.GET.get('selected_end_time')

        # Convert selected times to datetime.time if they are strings
        start_time = datetime.strptime(selected_start_time, '%H:%M').time()
        end_time = datetime.strptime(selected_end_time, '%H:%M').time()

        # Filter out tables that are already reserved for the selected date and time
        reserved_tables = laboratory_reservations.objects.filter(
            room_id=room_id,
            start_date=selected_date,
            status__in=['R', 'A', 'P']
        ).filter(
            start_time__lt=end_time,
            end_time__gt=start_time
        ).values_list('table_id', flat=True)

        tables = RoomTable.objects.filter(room_id=room_id).exclude(table_id__in=reserved_tables).values('table_id', 'table_name', 'capacity')
        return JsonResponse({'tables': list(tables)})
    except ValueError as e:
        logger.error(f"Invalid time format in get_room_tables: {e}")
        return JsonResponse({'error': 'Invalid time format'}, status=400)
    
    except Exception as e:
        logger.error(f"Error in get_room_tables: {e}", exc_info=True)
        return JsonResponse({'error': 'An unexpected error occurred'}, status=500)

@login_required
@lab_permission_required('reserve_laboratory')
def lab_reservation_student_reserveLabConfirm(request):
    try:
        selected_laboratory_id = request.session.get('selected_lab')
        reservation_config_obj = reservation_config.objects.get(laboratory_id=selected_laboratory_id)

        if reservation_config_obj.require_approval:
            del request.session['reservation_id']
        res_id = None
        if request.method == 'POST':
            selected_room_id = request.POST.get('selectedRoom')
            selected_table_id = request.POST.get('selectedTable')  
            selected_date = request.POST.get('selectedDate')
            selected_start_time = request.POST.get('selectedStartTime')
            selected_end_time = request.POST.get('selectedEndTime')

            if reservation_config_obj.require_approval:
                res_id = request.POST.get('reservation_id')

            # Fetch room information
            selected_room = get_object_or_404(rooms, room_id=selected_room_id)
            selected_table = get_object_or_404(RoomTable, table_id=selected_table_id)

            print(selected_start_time, '--', selected_end_time)

            # Check if the table is already reserved for the selected date and time
            existing_reservation = laboratory_reservations.objects.filter(
                table=selected_table,
                start_date=selected_date,
                status__in=['R', 'A', 'P']
            ).filter(
                # Check for overlapping time slots
                start_time__lt=selected_end_time,  # Existing reservation starts before new reservation ends
                end_time__gt=selected_start_time     # Existing reservation ends after new reservation starts
            ).exists()


            # Fetch blocked times for the room
            blocked_times = json.loads(selected_room.blocked_time) if selected_room.blocked_time else {}
            day_of_week = datetime.strptime(selected_date, '%Y-%m-%d').strftime('%A')  # Get the day of the week

            # Convert selected times to datetime.time if they are strings
            if isinstance(selected_start_time, str):
                check_selected_start_time = datetime.strptime(selected_start_time, '%H:%M').time()
            if isinstance(selected_end_time, str):
                check_selected_end_time = datetime.strptime(selected_end_time, '%H:%M').time()

            print(selected_start_time, selected_end_time)

            # Check for overlaps with blocked times
            is_blocked = False
            if day_of_week in blocked_times:
                for blocked_time in blocked_times[day_of_week]: 
                    print('pass')             
                    blocked_start, blocked_end = blocked_time.split('-')
                    blocked_start_time = datetime.strptime(blocked_start.strip(), '%H:%M').time()
                    blocked_end_time = datetime.strptime(blocked_end.strip(), '%H:%M').time()
                    if (check_selected_start_time < blocked_end_time) and (check_selected_end_time > blocked_start_time):
                        print('pass')
                        is_blocked = True
                        break
            
            print(existing_reservation, is_blocked)

            if existing_reservation or is_blocked:
                error_message = "The selected time slot for this room is not available (blocked or already reserved)."
                messages.error(request, error_message)
                return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

            # Save the reservation data in session temporarily for confirmation
            request.session['reservation_data'] = {
                'res_id': res_id,
                'room_id': selected_room.room_id,
                'room_name': selected_room.name,
                'selected_date': selected_date,
                'start_time': selected_start_time,
                'end_time': selected_end_time,
                'table_id': selected_table.table_id,  
                'table_name': selected_table.table_name,  
            }

            return redirect('lab_reservation_student_reserveLabConfirmDetails')

        return HttpResponse("Invalid request", status=400)
    except Exception as e:
        logger.error(f"Error in lab_reservation_student_reserveLabConfirm: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('home')
    
@transaction.atomic
@login_required
@lab_permission_required('reserve_laboratory')
def lab_reservation_student_reserveLabConfirmDetails(request):
    try:
        reservation_data = request.session.get('reservation_data')
        current_user = request.user
        selected_laboratory_id = request.session.get('selected_lab')
        reservation_config_obj = get_object_or_404(reservation_config, laboratory_id=selected_laboratory_id)
        preapproval_details = None

        if not reservation_data:
            return redirect('lab_reservation_student_reserveLabChooseRoom')

        room_id = reservation_data.get('room_id')
        room = get_object_or_404(rooms, room_id=room_id)

        table_id = reservation_data.get('table_id')
        table = get_object_or_404(RoomTable, table_id=table_id)

        if reservation_config_obj.require_approval:
            preapproval_details = get_object_or_404(laboratory_reservations, reservation_id=reservation_data.get('res_id'))

        if request.method == 'POST':
            # Get user details from the form
            contact_name = request.POST.get('contact_name')
            contact_email = request.POST.get('contact_email')
            num_people = request.POST.get('num_people')
            purpose = request.POST.get('purpose')
            selected_table_name = request.POST.get('selectedTable')
            

            # Set reservation status
            status = 'P'  # Set to pending initially
            message = 'Room Reserved Successfully'

            if room.capacity < int(num_people):
                messages.error(request, 'Number of people reached the maximum capacity.')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
            
            # Get the selected table and update its capacity
            selected_table = get_object_or_404(RoomTable, table_id=table_id)
            blocked_times = selected_table.blocked_time.get(reservation_data['selected_date'], [])
    
            # Convert selected times to datetime.time if they are strings
            start_time = datetime.strptime(reservation_data['start_time'], '%H:%M').time()
            end_time = datetime.strptime(reservation_data['end_time'], '%H:%M').time()

            # Check for overlaps with blocked times
            is_blocked = False
            for blocked_time in blocked_times:
                blocked_start, blocked_end = blocked_time.split('-')
                blocked_start_time = datetime.strptime(blocked_start.strip(), '%H:%M').time()
                blocked_end_time = datetime.strptime(blocked_end.strip(), '%H:%M').time()
                if (start_time < blocked_end_time) and (end_time > blocked_start_time):
                    is_blocked = True
                    break

            if is_blocked:
                messages.error(request, 'Selected table is already reserved for the selected time slot.')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

            # Mark the table as reserved for the selected time slot
            blocked_times.append(f"{reservation_data['start_time']}-{reservation_data['end_time']}")
            selected_table.blocked_time[reservation_data['selected_date']] = blocked_times
            selected_table.save()

            if selected_table.capacity < int(num_people):
                messages.error(request, 'Selected table does not have enough capacity.')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

            # selected_table.capacity -= int(num_people)
            # selected_table.save()

            # Check if all tables are reserved for the selected time slot
            all_tables_reserved = all(
                any(
                    (start_time < datetime.strptime(bt.split('-')[1], '%H:%M').time()) and
                    (end_time > datetime.strptime(bt.split('-')[0], '%H:%M').time())
                    for bt in table.blocked_time.get(reservation_data['selected_date'], [])
                ) for table in room.tables.all()
            )
            if all_tables_reserved:
                status = 'R'  # Set to reserved if all tables are reserved for the selected time slot


            # Handle the uploaded PDF file if required
            if reservation_config_obj.require_approval:
                reservation = get_object_or_404(laboratory_reservations, reservation_id=reservation_data['res_id'])
                reservation.room = room
                reservation.table = selected_table  
                reservation.start_date = reservation_data['selected_date']
                reservation.start_time = reservation_data['start_time']
                reservation.end_time = reservation_data['end_time']
                reservation.status = status
                reservation.save()
            else:
                reservation = laboratory_reservations.objects.create(
                    user=current_user,
                    room=room,
                    table=selected_table,
                    start_date=reservation_data['selected_date'],
                    start_time=reservation_data['start_time'],
                    end_time=reservation_data['end_time'],
                    contact_name=contact_name,
                    contact_email=contact_email,
                    num_people=num_people,
                    purpose=purpose,
                    status=status,
                    laboratory_id = selected_laboratory_id
                )

            # Clear session data and redirect
            del request.session['reservation_data']
            messages.success(request, message)
            return redirect('lab_reservation_detail', reservation.reservation_id)

        # Check if approval form exists for download link
        approval_form_exists = bool(reservation_config_obj.approval_form)
        print(preapproval_details)
        
        return render(request, 'mod_labRes/lab_reservation_studentReserveLabConfirm.html', {
            'reservation_data': reservation_data,
            'preapproval_details': preapproval_details,
            'user': current_user if request.user.is_authenticated else None,
            'reserv_config': reservation_config_obj,
            'approval_form_exists': approval_form_exists,
            'room': room
        })
    except Exception as e:
        logger.error(f"Error in lab_reservation_student_reserveLabConfirmDetails: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('home')

# not used for now
@login_required
@lab_permission_required('reserve_laboratory')
def lab_reservation_student_reserveLabChooseTime(request):
     return render(request, 'mod_labRes/lab_reservation_studentReserveLabChooseTime.html')


@login_required
@lab_permission_required('reserve_laboratory')
def lab_reservation_student_reserveLabSummary(request):
    try:
        current_user = request.user
        selected_laboratory_id = request.session.get('selected_lab')
        
        # Get today's date
        today = timezone.localtime().date()

        # Filter reservations by categories
        tab = request.GET.get('tab', 'all')  # Default to 'today' tab
        reservations = laboratory_reservations.objects.filter(user=current_user, laboratory_id=selected_laboratory_id)

        if tab == 'today':
            reservations = reservations.filter(start_date=today)
        elif tab == 'previous':
            reservations = reservations.filter(start_date__lt=today)
        elif tab == 'future':
            reservations = reservations.filter(start_date__gt=today)
        elif tab == 'cancelled':
            reservations = reservations.filter(status='C')  # Assuming 'C' is the status for cancelled
        else:
            reservations = reservations.all()

        context = {
            'reservations': reservations,
            'tab': tab,
        }

        return render(request, 'mod_labRes/lab_reservation_studentReserveLabSummary.html', context)
    except Exception as e:
        logger.error(f"Error in lab_reservation_student_reserveLabSummary: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred while fetching reservations.")
        return redirect('home')

@transaction.atomic
@login_required
@lab_permission_required('reserve_laboratory')
def cancel_reservation(request, reservation_id):
    try:
        # Get the reservation object by id
        current_user = request.user
        reservation = get_object_or_404(laboratory_reservations, reservation_id=reservation_id, user=current_user)

        # Change the status to 'Cancelled' (assuming 'C' represents cancelled)
        reservation.status = 'C'
        reservation.save()

        # Redirect to the reservation summary page
        return redirect('lab_reservation_student_reserveLabSummary')
    except laboratory_reservations.DoesNotExist:
        messages.error(request, "Reservation not found.")
        return redirect('lab_reservation_student_reserveLabSummary')
    except Exception as e:
        logger.error(f"Error cancelling reservation {reservation_id}: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred while cancelling your reservation.")
        return redirect('lab_reservation_student_reserveLabSummary')

@login_required
@lab_permission_required('reserve_laboratory')
def lab_reservation_detail(request, reservation_id):
    try:
        # Get the reservation object by its ID
        current_user = request.user
        reservation = get_object_or_404(laboratory_reservations, reservation_id=reservation_id, user=current_user)
        if request.method == "POST":
            if 'chooseroom' in request.POST:
                print('pass')
                res_id = request.POST.get('reservation_id')
                request.session['reservation_id'] = {
                    'res_id': res_id,
                }
                return redirect('lab_reservation_student_reserveLabChooseRoom')
        context = {
            'reservation': reservation,
        }
        return render(request, 'mod_labRes/lab_reservation_detail.html', context)
    except laboratory_reservations.DoesNotExist:
        messages.error(request, "Reservation not found.")
        return redirect('lab_reservation_student_reserveLabSummary')
    except Exception as e:
        logger.error(f"Error in lab_reservation_detail: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred.")
        return redirect('lab_reservation_student_reserveLabSummary')

@transaction.atomic
@login_required
@lab_permission_required('view_reservations')
def export_schedule_to_excel(request):
    try:
        selected_room = request.GET.get('roomSelect')
        selected_month = request.GET.get('selectMonth')
        if not selected_room or not selected_month:
            return JsonResponse({"error": "Invalid parameters. Please provide a valid room and month."}, status=400)

        # Fetch reservations
        room = rooms.objects.get(room_id=selected_room)
        year, month = map(int, selected_month.split('-'))
        start_date = datetime(year, month, 1)
        end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)

        reservations = laboratory_reservations.objects.filter(
            room_id=room.room_id,
            start_date__range=[start_date, end_date]
        )

        # Prepare reservations by day
        reservations_by_day = {}
        for reservation in reservations:
            day = reservation.start_date.day
            if day not in reservations_by_day:
                reservations_by_day[day] = []
            reservations_by_day[day].append(reservation)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{room.name} - {selected_month} ({year})"

        # Write headers for the calendar
        headers = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        ws.append(headers)

        # Fill in the calendar
        first_day_of_month = start_date.weekday()  # 0 = Monday, 6 = Sunday
        total_days = (end_date - start_date).days + 1

        row = []
        # Add empty cells for days before the start of the month
        for _ in range(first_day_of_month):
            row.append("")

        for day in range(1, total_days + 1):
            day_reservations = reservations_by_day.get(day, [])
            cell_content = f"{day}\n"  # Add the day number
            for res in day_reservations:
                cell_content += f"{res.start_time.strftime('%I:%M %p')} - {res.end_time.strftime('%I:%M %p')}\n{res.contact_name}\n"

            row.append(cell_content.strip())  # Add to the current row

            # Start a new row every 7 days or at the end of the month
            if len(row) == 7 or day == total_days:
                ws.append(row)
                row = []

        # Merge cells to add a title above the calendar
        title_cell = f"Laboratory Schedule - {room.name} ({month:02d}/{year})"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        ws.cell(row=1, column=1, value=title_cell)
        ws.row_dimensions[1].height = 20
        title_font = openpyxl.styles.Font(size=14, bold=True)
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal="center")

        # Adjust cell alignment for better readability
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                if cell.value:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

        # Return Excel file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename=Laboratory_Schedule_{month:02d}_{year}.xlsx"
        wb.save(response)
        return response
    except rooms.DoesNotExist:
        return JsonResponse({"error": "Room not found."}, status=404)
    except Exception as e:
        logger.error(f"Error exporting schedule: {e}", exc_info=True)
        return JsonResponse({"error": "An unexpected error occurred while exporting the schedule."}, status=500)

@login_required
@lab_permission_required('view_reservations')
def labres_lab_schedule(request):
    try:
        selected_laboratory_id = request.session.get('selected_lab')
        room_list = []
        reservations = []
        selected_month = None
        selected_room = None
        reservations_by_day = {}
        days_range = None
        # Get the current month
        current_month = timezone.localtime().strftime('%Y-%m')

        if selected_laboratory_id:
            room_list = rooms.objects.filter(laboratory_id=selected_laboratory_id, is_disabled=False)

        if request.method == "GET":
            selected_room = request.GET.get('roomSelect')
            selected_month = request.GET.get('selectMonth', current_month)
            
            if selected_room and selected_month:
                try:
                    room = rooms.objects.get(room_id=selected_room, laboratory_id=selected_laboratory_id)

                    # Get the first and last day of the selected month
                    year, month = map(int, selected_month.split('-'))
                    start_date = datetime(year, month, 1)
                    end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)

                    # Get reservations for the selected month
                    reservations = laboratory_reservations.objects.filter(
                        room_id=room.room_id,  # Ensure you use room.room_id to access the ID
                        start_date__range=[start_date, end_date]
                    )

                    # Format start_time and end_time for each reservation
                    for reservation in reservations:
                        reservation.formatted_start_time = reservation.start_time.strftime("%I:%M%p") if reservation.start_time else "N/A"
                        reservation.formatted_end_time = reservation.end_time.strftime("%I:%M%p") if reservation.end_time else "N/A"

                    # Prepare a list to hold reservations indexed by day
                    days_in_month = monthrange(year, month)[1]
                    reservations_by_day = {day: [] for day in range(1, days_in_month + 1)}

                    # Organize reservations by day
                    for reservation in reservations:
                        reservations_by_day[reservation.start_date.day].append(reservation)

                    # Pass the range of days to the template
                    days_range = range(1, days_in_month + 1)

                    

                except rooms.DoesNotExist:
                    print(f"Room '{selected_room}' does not exist.")
                except Exception as e:
                    print(f"Error occurred: {str(e)}")

        context = {
            'room_list': room_list,
            'reservations_by_day': reservations_by_day,  # Pass the organized reservations
            'selected_month': selected_month,
            'selected_room': selected_room,
            'days_range': days_range,  # Pass the range of days
        }

        return render(request, 'mod_labRes/labres_lab_schedule.html', context)
    except Exception as e:
        logger.error(f"Error in labres_lab_schedule: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred.")
        return redirect('home')

@transaction.atomic
@login_required
@lab_permission_required('approve_deny_reservations')
def labres_lab_reservationreqs(request):
    try:
        selected_laboratory_id = request.session.get('selected_lab')
        require_approval = reservation_config.objects.filter(laboratory_id=selected_laboratory_id).values('require_approval')

        reservations = []
        selected_room = None
        selected_date = None
        room_list = []

        requests = laboratory_reservations.objects.filter(laboratory_id=selected_laboratory_id, status='P').annotate(room_name=F('room__name'))

        if selected_laboratory_id:
            room_list = rooms.objects.filter(laboratory_id=selected_laboratory_id, is_disabled=False)

        if request.method == "GET":
            selected_room = request.GET.get('roomSelect')
            selected_date = request.GET.get('selectDate')

            if selected_room and selected_date:
                try:
                    room = rooms.objects.get(room_id=selected_room, laboratory_id=selected_laboratory_id)
                    reservations = laboratory_reservations.objects.filter(
                        room=room,
                        start_date=selected_date,
                        room__is_disabled=0
                    ).order_by('request_date')

                    # Format start_time, end_time and calculate interval
                    for reservation in reservations:
                        # Format time as 9:00AM
                        reservation.formatted_start_time = reservation.start_time.strftime("%I:%M%p")
                        reservation.formatted_end_time = reservation.end_time.strftime("%I:%M%p")

                        # Combine start_date with start_time and end_time to get datetime objects
                        start_datetime = datetime.combine(reservation.start_date, reservation.start_time)
                        end_datetime = datetime.combine(reservation.start_date, reservation.end_time)

                        # Calculate the time difference in minutes
                        time_difference = end_datetime - start_datetime
                        reservation.time_interval = int(time_difference.total_seconds() // 60)


                except rooms.DoesNotExist:
                    pass
            else:
                reservations = laboratory_reservations.objects.filter(
                    laboratory_id=selected_laboratory_id
                ).exclude(room__isnull=True).annotate(room_name=F('room__name')).order_by('-request_date')
                # Format start_time, end_time and calculate interval
                for reservation in reservations:
                    # Format time as 9:00AM
                    reservation.formatted_start_time = reservation.start_time.strftime("%I:%M%p")
                    reservation.formatted_end_time = reservation.end_time.strftime("%I:%M%p")

                    # Combine start_date with start_time and end_time to get datetime objects
                    start_datetime = datetime.combine(reservation.start_date, reservation.start_time)
                    end_datetime = datetime.combine(reservation.start_date, reservation.end_time)

                    # Calculate the time difference in minutes
                    time_difference = end_datetime - start_datetime
                    reservation.time_interval = int(time_difference.total_seconds() // 60)

        # Handle POST requests for Accept and Delete actions
        if request.method == "POST":
            action = request.POST.get('action')
            reservation_id = request.POST.get('reservation_id')

            if action == "accept":
                reservation = get_object_or_404(laboratory_reservations, reservation_id=reservation_id, status='P')
                reservation.status = 'A'  # Set to approved
                reservation.save()
                messages.success(request, f"Reservation {reservation_id} has been approved.")
            elif action == "decline":
                reservation = get_object_or_404(laboratory_reservations, reservation_id=reservation_id, status='P')
                reservation.status = 'D'  # Set to declined
                reservation.save()
                messages.success(request, f"Reservation {reservation_id} has been declined.")
            
            return redirect('labres_lab_reservationreqs')  # Redirect to avoid resubmission
        
        # if request.method == "POST":
        #     action = request.POST.get('action')

        #     if action:
        #         # Check if the action is accept or delete for a specific reservation
        #         if action.startswith("accept_"):
        #             reservation_id = action.split("_")[1]
        #             reservation = get_object_or_404(laboratory_reservations, reservation_id=reservation_id)

        #             if reservation.status == 'P':  # Only update if pending
        #                 reservation.status = 'R'  # Change status to Approved
        #                 reservation.save()

        #         elif action.startswith("delete_"):
        #             reservation_id = action.split("_")[1]
        #             reservation = get_object_or_404(laboratory_reservations, reservation_id=reservation_id)

        #             if reservation.status == 'P':  # Only update if pending
        #                 reservation.status = 'D'  # Change status to Cancelled by Lab Tech
        #                 reservation.save()

        #     # Redirect to avoid form resubmission on page reload
        #     return HttpResponseRedirect(reverse('labres_lab_reservationreqs'))
        print('test:',require_approval)
        context = {
            'room_list': room_list,
            'reservations': reservations,
            'selected_room': selected_room,
            'selected_date': selected_date,

            'requests': requests,
            'require_approval': require_approval
        }

        return render(request, 'mod_labRes/labres_lab_reservationreqs.html', context)
    except Exception as e:
        logger.error(f"Error in labres_lab_reservationreqs: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred.")
        return redirect('home')

@transaction.atomic
@login_required
@lab_permission_required('approve_deny_reservations')
def labres_lab_reservationreqsDetailed(request, reservation_id):
    selected_laboratory_id = request.session.get('selected_lab')
    try:
        try:
            reservation = laboratory_reservations.objects.get(reservation_id=reservation_id)
            if(reservation.room_id):
                room = rooms.objects.get(room_id=reservation.room_id, laboratory_id=selected_laboratory_id)
            else:
                room=None
        except (laboratory_reservations.DoesNotExist, rooms.DoesNotExist):
            reservation = None
            room = None

        if request.method == "POST":
            action = request.POST.get('action')
            if action == 'accept':
                # Change status to 'Approved' ('A')
                reservation.status = 'R'
            elif action == 'delete':
                # Change status to 'Cancelled' ('L')
                reservation.status = 'D'
            reservation.save()

            return redirect('labres_lab_reservationreqsDetailed', reservation_id=reservation_id)

        context = {
            'reservation': reservation,
            'room': room,
        }

        return render(request, 'mod_labRes/labres_lab_reservationreqsDetailed.html', context)
    except Exception as e:
        logger.error(f"Error fetching reservation details: {e}", exc_info=True)
        messages.error(request, "An error occurred while fetching reservation details.")
        return redirect('labres_lab_reservationreqs')

@transaction.atomic
@login_required
@lab_permission_required('configure_lab_reservation')
def labres_labcoord_configroom(request):
    selected_laboratory_id = request.session.get('selected_lab')
    message = None
    room_configured = None

    try:
        reservation_config_obj, created = reservation_config.objects.get_or_create(laboratory_id=selected_laboratory_id)
    except Exception as e:
        logger.error(f"Error fetching reservation config: {e}", exc_info=True)
        messages.error(request, "Failed to retrieve reservation configuration.")
        return redirect('home')

    if request.method == "POST":
        try:
            if 'save_rooms' in request.POST:
                # Update is_reservable field for rooms
                for room in rooms.objects.filter(laboratory_id=selected_laboratory_id):
                    room.is_reservable = f'room_{room.room_id}_enabled' in request.POST
                    room.save()
                message = 'Successfully saved room configuration'

            elif 'delete_room' in request.POST:
                delete_room_id = request.POST.get('delete_room')
                room = get_object_or_404(rooms, room_id=delete_room_id)
                room.is_disabled = True
                room.save()
                message = 'Successfully deleted a room'

            elif 'add_room' in request.POST:
                room_name = request.POST.get('room_name')
                room_capacity = request.POST.get('room_capacity')
                room_description = request.POST.get('room_description')

                if room_name and room_capacity:
                    new_room = rooms(
                        laboratory_id=selected_laboratory_id,
                        name=room_name,
                        capacity=room_capacity,
                        description=room_description
                    )
                    new_room.save()
                message = f'Successfully added Room {room_name}'

            elif 'save_time' in request.POST:
                # Save the general reservation configuration for the lab (applies to all rooms)
                reservation_type = request.POST.get('reservation_type')
                reservation_config_obj.reservation_type = reservation_type

                if reservation_type == 'hourly':
                    reservation_config_obj.start_time = request.POST.get('hourly_start_time')
                    reservation_config_obj.end_time = request.POST.get('hourly_end_time')
                else:
                    reservation_config_obj.start_time = '07:30'
                    reservation_config_obj.end_time = '19:30'

                reservation_leadtime = request.POST.get('lead_time')
                reservation_config_obj.leadtime = reservation_leadtime

                room = rooms.objects.filter(laboratory_id=selected_laboratory_id)
                room.update(blocked_time='{}')

                reservation_config_obj.save()
                message = f'Time configuration saved'

            elif 'save_timeblocked' in request.POST:
                room_id = request.POST.get('room_id')
                room_configured = get_object_or_404(rooms, pk=room_id)

                blocked_times = {}
                for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']:
                    time_slots = request.POST.getlist(f'{day}_time_slots')
                    if time_slots:
                        blocked_times[day] = time_slots

                room_configured.blocked_time = json.dumps(blocked_times)
                room_configured.save()
                message = f'Time configuration saved for {room_configured.name}'
            
            elif 'save_approval' in request.POST:
                # Save approval settings and optional PDF form
                reservation_config_obj.require_approval = 'require_approval' in request.POST

                if 'approval_form' in request.FILES:
                    reservation_config_obj.approval_form = request.FILES['approval_form']

                reservation_config_obj.save()
                message = 'Approval settings saved'

            elif 'save_tc' in request.POST:
                # Save terms and conditions description
                reservation_config_obj.tc_description = request.POST.get('tc_description')
                reservation_config_obj.save()
                message = 'Terms and conditions saved'

            elif 'add_table' in request.POST:
                room_id = request.POST.get('room_id')
                table_name = request.POST.get('table_name')
                table_capacity = request.POST.get('table_capacity')
                room = get_object_or_404(rooms, room_id=room_id)

                if table_name and table_capacity:
                    new_table = RoomTable(
                        room=room,
                        table_name=table_name,
                        capacity=table_capacity
                    )
                    new_table.save()
                message = f'Successfully added Table {table_name} to Room {room.name}'

            elif 'delete_table' in request.POST:
                table_id = request.POST.get('table_id')
                table = get_object_or_404(RoomTable, table_id=table_id)
                table.delete()
                message = f'Successfully deleted Table {table.table_name} from Room {table.room.name}'
        except rooms.DoesNotExist:
            messages.error(request, "The selected room does not exist.")
        except RoomTable.DoesNotExist:
            messages.error(request, "The selected table does not exist.")
        except Exception as e:
            logger.error(f"Error updating room configuration: {e}", exc_info=True)
            messages.error(request, "An error occurred while updating room configuration.")


        messages.success(request, message)
    # Fetch rooms and reservation config
    rooms_query = rooms.objects.filter(laboratory_id=selected_laboratory_id, is_disabled=False)
    reservation_config_data = {
        'reservation_type': reservation_config_obj.reservation_type,
        'start_time': reservation_config_obj.start_time,
        'end_time': reservation_config_obj.end_time,
        'require_approval': reservation_config_obj.require_approval,
        'tc_description': reservation_config_obj.tc_description,
        'leadtime': reservation_config_obj.leadtime,
        'approval_form': reservation_config_obj.approval_form,
    }

    context = {
        'rooms': rooms_query,
        'room_configured': room_configured,
        'reservation_config': reservation_config_data,
    }

    return render(request, 'mod_labRes/labres_labcoord_configroom.html', context)

@transaction.atomic
@login_required
@lab_permission_required('configure_lab_reservation')
def labres_bulk_upload(request):
    selected_laboratory_id = request.session.get('selected_lab')
    try:
        rooms_query = rooms.objects.filter(laboratory_id=selected_laboratory_id, is_disabled=False)

        if request.method == "POST" and 'download_template' in request.POST:
            selected_room_ids = request.POST.getlist('selected_rooms')
            print(f"Selected Room IDs: {selected_room_ids}")  # Debugging line
            if not selected_room_ids:
                print("No rooms selected!")  # Debugging line

            selected_rooms = rooms.objects.filter(room_id__in=selected_room_ids)
            
            # If no rooms are selected, return an error message
            if not selected_rooms.exists():
                print("No rooms found!")  # Debugging line
                return HttpResponse("No rooms found.", status=400)

            # Create Excel template with selected room names
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Room Time Slots"
            
            # Add headers
            ws.append(["Room ID", "Room Name", "Time Slot"])

            # Add rows for each selected room
            for room in selected_rooms:
                ws.append([room.room_id, room.name, ''])  # Add an empty column for Time Slot

            # Create an in-memory file
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="room_time_slots_template.xlsx"'

            # Save the workbook to the response object
            with BytesIO() as buffer:
                wb.save(buffer)
                buffer.seek(0)
                response.write(buffer.read())

            return response

        context = {
            'rooms': rooms_query,
        }

        return render(request, 'mod_labRes/labres_bulk_upload.html', context)
    except Exception as e:
        logger.error(f"Error in labres_bulk_upload: {e}", exc_info=True)
        messages.error(request, f"An error occurred while generating the template: {str(e)}")
        return redirect('labres_labcoord_configroom')

@transaction.atomic
@login_required
@lab_permission_required('configure_lab_reservation')
def labres_bulk_upload_time(request):   
    selected_laboratory_id = request.session.get('selected_lab')
    
    # Define allowed time slots
    ALLOWED_TIME_SLOTS = [
        '7:30-9:00',
        '9:15-10:45',
        '11:00-12:30',
        '12:45-2:15',
        '2:30-4:00',
        '4:15-5:45',
        '6:00-7:30',
    ]
    
    if request.method == 'POST' and 'upload_file' in request.FILES:
        excel_file = request.FILES['upload_file']
        
        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            messages.error(request, f"Error reading the Excel file: {str(e)}")
            return redirect('labres_labcoord_configroom')  # Redirect to the lab configuration page

        # Validate that required columns exist
        required_columns = ['Room ID', 'Room Name', 'Time Slot']
        for column in required_columns:
            if column not in df.columns:
                messages.error(request, f"Missing required column: {column}")
                return redirect('labres_labcoord_configroom')
        
        # Iterate through the rows and update the `blocked_time` for each room
        for index, row in df.iterrows():
            try:
                room_id = row.get('Room ID')
                time_slot = row.get('Time Slot')

                if not room_id or not time_slot:
                    messages.error(request, f"Missing Room ID or Time Slot in row {index + 1}.")
                    return redirect('labres_labcoord_configroom')
                
                room = rooms.objects.filter(room_id=room_id).first()
                if room:
                    blocked_times = json.loads(room.blocked_time) if room.blocked_time else {}

                    # Split the time slots by commas if multiple slots are provided
                    time_slots = time_slot.split(',')  # This will handle multiple time slots in a single cell
                    
                    for ts in time_slots:
                        ts = ts.strip()  # Remove any extra spaces around the time slot
                        
                        # Extract the time range (e.g., "7:30-9:00" from "Monday 7:30-9:00")
                        try:
                            _, time_range = ts.split(' ', 1)  # Split day and time
                            
                            # Validate if the extracted time is in the allowed list
                            if time_range not in ALLOWED_TIME_SLOTS:
                                messages.error(request, f"Invalid time slot '{time_range}' in row {index + 1}. Allowed time slots are: {', '.join(ALLOWED_TIME_SLOTS)}.")
                                return redirect('labres_labcoord_configroom')

                            # Update blocked times for the room
                            day = ts.split(' ', 1)[0]  # Extract the day
                            if day not in blocked_times:
                                blocked_times[day] = []
                            blocked_times[day].append(time_range)

                        except ValueError:
                            messages.error(request, f"Invalid time slot format in row {index + 1}. Correct format: 'Day time' (e.g., 'Monday 7:30-9:00').")
                            return redirect('labres_labcoord_configroom')

                    room.blocked_time = json.dumps(blocked_times)
                    room.save()

                else:
                    messages.error(request, f"Room with ID {room_id} not found in row {index + 1}.")
                    return redirect('labres_labcoord_configroom')
            except Exception as e:
                logger.error(f"Error processing row {index + 1}: {e}", exc_info=True)
                messages.error(request, f"Error processing row {index + 1}: {e}")
                continue

        messages.success(request, "Bulk upload of time slots was successful.")
        return redirect('labres_labcoord_configroom')  # Redirect back to the lab configuration page
    
    return redirect('labres_labcoord_configroom')  # In case of non-POST request, just redirect

@login_required
@lab_permission_required('configure_lab_reservation')
def get_room_configuration(request, room_id):
    try:
        room = get_object_or_404(rooms, pk=room_id)
        reservation_config_obj = reservation_config.objects.get(laboratory=room.laboratory)

        # Return the room's configuration as JSON
        response_data = {
            'reservation_type': reservation_config_obj.reservation_type,
            'start_time': reservation_config_obj.start_time.strftime('%H:%M') if reservation_config_obj.start_time else None,
            'end_time': reservation_config_obj.end_time.strftime('%H:%M') if reservation_config_obj.end_time else None,
            'blocked_time': room.blocked_time if room.blocked_time else {} 
        }

        return JsonResponse(response_data)
    except Exception as e:
        logger.error(f"Error fetching room configuration: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)



# ======================= WIP ============================

@login_required
@lab_permission_required('view_wip')
def view_wip(request, wip_id):
    try:
        selected_laboratory_id = request.session.get('selected_lab')
        selected_laboratory = get_object_or_404(laboratory, laboratory_id=selected_laboratory_id)

        current_user = request.user
        user_roles = laboratory_users.objects.filter(
            laboratory=selected_laboratory, 
            user=current_user, 
            is_active=True, 
            status='A'
        ).values_list('role', flat=True)
        user_permissions = laboratory_permissions.objects.filter(
            role__in=user_roles, 
            laboratory=selected_laboratory
        )
        permission_ids = user_permissions.values_list('permissions__permission_id', flat=True)

        if 22 in permission_ids:
            wip = get_object_or_404(WorkInProgress, wip_id=wip_id)
        elif 23 in permission_ids:
            wip = get_object_or_404(WorkInProgress, wip_id=wip_id, user=current_user)
        else:
            wip = WorkInProgress.objects.none()

        return render(request, 'mod_labRes/wip_view.html', {'wip': wip})
    except Exception as e:
        logger.error(f"Error viewing wip: {e}", exc_info=True)
        messages.error(request, "An error occured unexpectedly.")
        return redirect('home')

@transaction.atomic
@login_required
@lab_permission_required('view_wip')
def list_wip(request):
    try:
        selected_laboratory_id = request.session.get('selected_lab')
        selected_laboratory = get_object_or_404(laboratory, laboratory_id=selected_laboratory_id)

        current_user = request.user
        user_roles = laboratory_users.objects.filter(
            laboratory=selected_laboratory, 
            user=current_user, 
            is_active=True, 
            status='A'
        ).values_list('role', flat=True)
        user_permissions = laboratory_permissions.objects.filter(
            role__in=user_roles, 
            laboratory=selected_laboratory
        )
        permission_ids = user_permissions.values_list('permissions__permission_id', flat=True)

        if 22 in permission_ids:
            wip_experiments = WorkInProgress.objects.filter(laboratory=selected_laboratory_id).order_by('-start_time')
        elif 23 in permission_ids:
            wip_experiments = WorkInProgress.objects.filter(laboratory=selected_laboratory_id, user=current_user).order_by('-start_time')
        else:
            wip_experiments = WorkInProgress.objects.none()

        rooms_query = rooms.objects.filter(laboratory_id=selected_laboratory_id, is_disabled=False, is_reservable=True)

        if request.method == 'POST':
            selected_user_id = request.POST.get('user')
            room_id = request.POST.get('room')
            description = request.POST.get('description')
            remarks = request.POST.get('remarks')
            selected_user = user.objects.get(user_id=selected_user_id)
            selected_room = rooms.objects.get(room_id=room_id)
            start_time = timezone.now()

            WorkInProgress.objects.create(
                user_id=selected_user.user_id,
                laboratory=selected_laboratory,
                room=selected_room,
                description=description,
                start_time=start_time,
                remarks=remarks,
                status='A'
            )
            messages.success(request, "WIP created successfully!")
            return redirect('list_wip')
        
        context = {
            'wip_experiments': wip_experiments,
            'rooms': rooms_query,
        }
        return render(request, 'mod_labRes/wip_list.html', context )
    except Exception as e:
        logger.error(f"Error viewing list wip: {e}", exc_info=True)
        messages.error(request, "An error occured unexpectedly.")
        return redirect('home')

@transaction.atomic
@login_required
@lab_permission_required('clear_wip')
def clear_wip(request, wip_id):
    try:
        wip = get_object_or_404(WorkInProgress, wip_id=wip_id)
        if request.method == 'POST':
            wip.end_time = timezone.now()
            wip.status = 'C'  # Mark as completed
            wip.save()
            messages.success(request, 'WIP experiment cleared successfully.')
            return redirect('list_wip')
        return render(request, 'mod_labRes/wip_clear.html', {'wip': wip})
    except Exception as e:
        logger.error(f"Error clearing wip: {e}", exc_info=True)
        messages.error(request, "An error occured unexpectedly.")
        return redirect('home')
#  ================================================================= 


#REPORTS module
@login_required
@lab_permission_required('view_reports')
def reports_view(request):
    selected_laboratory_id = request.session.get('selected_lab')
    try:
        lab_info = get_object_or_404(
            laboratory.objects.annotate(user_count=Count('laboratory_users')),
            laboratory_id=selected_laboratory_id
        )

        # Get all roles that are either default or specific to the selected laboratory
        all_roles = laboratory_roles.objects.filter(
            Q(laboratory_id=0) | Q(laboratory_id=selected_laboratory_id)
        )

        # Annotate each role with the count of users
        role_user_counts = all_roles.annotate(
            user_count=Count('users', filter=Q(users__laboratory_id=selected_laboratory_id))
        )
        role_user_data = {
            'labels': [role.name for role in role_user_counts],
            'series': [role.user_count for role in role_user_counts]
        }
        

        lab_users = laboratory_users.objects.filter(laboratory_id=selected_laboratory_id, is_active=1).select_related('user', 'role').annotate(
            username=F('user__username'),
            user_email=F('user__email'),
            full_name=Concat(F('user__firstname'), Value(' '), F('user__lastname'), output_field=CharField()),
            role_name=F('role__name'),
            personal_id=F('user__personal_id'),
        )
        
        context = {
            'lab_info': lab_info,
            'role_user_data': role_user_data,

            'lab_users': lab_users
        }

        return render(request, 'mod_reports/reports.html', context)
    except laboratory.DoesNotExist:
        logger.error(f"Laboratory with ID {selected_laboratory_id} does not exist.")
        messages.error(request, "Selected laboratory does not exist.")
        return redirect('home')  # Redirect to a safe page

    except Exception as e:
        logger.error(f"Unexpected error in reports_view: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred while loading the reports.")
        return redirect('home')  # Redirect to a safe page

def inventory_reports(request):
    try:
        selected_laboratory_id = request.session.get('selected_lab')
        current_date = date.today()
        purchased_filter_type = request.GET.get('purchased_filter', 'this_week')
        used_filter_type = request.GET.get('used_filter', 'this_week')

        print(purchased_filter_type, '---',used_filter_type )

        total_purchased = item_handling.objects.filter(
                inventory_item__item__laboratory_id=selected_laboratory_id,
                qty__gt=0,
                timestamp__date__range = calculate_date_range(request, purchased_filter_type)
            ).aggregate(total_qty=Sum('qty'))['total_qty'] or 0

        total_used = item_handling.objects.filter(
            inventory_item__item__laboratory_id=selected_laboratory_id,
            qty__lt=0,
            timestamp__date__range = calculate_date_range(request, used_filter_type)
        ).aggregate(total_qty=Sum('qty'))['total_qty'] or 0

        print(total_purchased,' --- ', total_used)
        
        # total qty of expired items
        expired_items_qty = item_inventory.objects.filter(
            item_expirations__expired_date__lt=current_date
        ).aggregate(total_qty=Sum('qty'))['total_qty']

        # If there are no expired items, set the quantity to 0
        if expired_items_qty is None:
            expired_items_qty = 0

        # inventory items table
        inventory_items = item_description.objects.filter(
            laboratory_id=selected_laboratory_id,
            is_disabled=False  # Only get items that are enabled
        ).annotate(total_qty=Coalesce(Sum('item_inventory__qty'), 0))  # Calculate total quantity
        inventory_qty_data = {
            'categories': [item.item_name for item in inventory_items],
            'series': [item.total_qty for item in inventory_items]
        }

        
        

        item_types_list = item_types.objects.filter(laboratory_id=selected_laboratory_id)
        item_id = request.GET.get('item_id')

        end_date = timezone.localtime().date()
        start_date = end_date - timedelta(days=365)

        daily_data = {}
        current_day = start_date  # Changed from 'date' to 'current_day'
        while current_day <= end_date:
            daily_data[current_day] = {"added_qty": 0, "removed_qty": 0}
            current_day += timedelta(days=1)

        if item_id:
            added_quantities = item_handling.objects.filter(
                inventory_item__item__laboratory_id=selected_laboratory_id,
                inventory_item__item_id=item_id,
                qty__gt=0,
                timestamp__date__range=(start_date, end_date)
            ).annotate(date=TruncDay('timestamp')).values('date').annotate(total_qty=Sum('qty')).order_by('date')

            removed_quantities = item_handling.objects.filter(
                inventory_item__item__laboratory_id=selected_laboratory_id,
                inventory_item__item_id=item_id,
                qty__lt=0,
                timestamp__date__range=(start_date, end_date)
            ).annotate(date=TruncDay('timestamp')).values('date').annotate(total_qty=Abs(Sum('qty'))).order_by('date')

            for entry in added_quantities:
                entry_date = entry['date'].date()  # Renamed to entry_date to avoid conflict
                daily_data[entry_date]["added_qty"] = entry['total_qty']

            for entry in removed_quantities:
                entry_date = entry['date'].date()
                daily_data[entry_date]["removed_qty"] = entry['total_qty']

        trend_data_added = [
            {"date": day.strftime('%Y-%m-%d'), "total_qty": data["added_qty"]}
            for day, data in daily_data.items()
        ]
        trend_data_removed = [
            {"date": day.strftime('%Y-%m-%d'), "total_qty": data["removed_qty"]}
            for day, data in daily_data.items()
        ]
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'total_purchased': total_purchased,
                'total_used': abs(total_used)  # Convert to positive number
            })
        
        items = item_description.objects.filter(laboratory_id=selected_laboratory_id, is_disabled=False)

        selected_laboratory_id = request.session.get('selected_lab')
        current_date = date.today()
        supplier_filter_type = request.GET.get('supplier_reports_filter', 'this_week')
        supplier_item_id = request.GET.get('item_id')
        start_date, end_date = calculate_date_range(request, supplier_filter_type)

        # Filter supplier data based on the selected item and date range
        supplier_data = []
        if supplier_item_id:
            inventory_items = item_inventory.objects.filter(item_id=supplier_item_id).exclude(supplier_id=None)
            for inventory in inventory_items:
                # Filter based on date range
                first_handling = item_handling.objects.filter(inventory_item=inventory, timestamp__date__range=calculate_date_range(request, supplier_filter_type)).order_by('timestamp').first()
                expiration_date = item_expirations.objects.filter(inventory_item=inventory).first()
                duration = (inventory.date_received - inventory.date_purchased).days if inventory.date_purchased and inventory.date_received else 'N/A'
                
                if first_handling:
                    supplier_data.append({
                        'supplier_name': inventory.supplier.supplier_name if inventory.supplier else 'N/A',
                        'inventory_item_id': inventory.inventory_item_id,
                        'timestamp': first_handling.timestamp,
                        'date_purchased': inventory.date_purchased,
                        'date_received': inventory.date_received,
                        'duration': duration,
                        'qty': first_handling.qty,
                        'purchase_price': inventory.purchase_price,
                        'expiration': expiration_date.expired_date if expiration_date else 'None'
                    })


        loss_filter_type = request.GET.get('loss_reports_filter', 'this_week')
        loss_start_date, loss_end_date= calculate_date_range(request, loss_filter_type)

        # Query the loss data
        loss_data_query = item_handling.objects.filter(
            changes='D',
            inventory_item__item__laboratory_id=selected_laboratory_id,
            inventory_item__item__is_disabled=0
        ).select_related(
            'inventory_item__item', 'inventory_item__item__itemType', 'updated_by'
        )

        # Apply date filtering
        if loss_start_date and loss_end_date:
            loss_data_query = loss_data_query.filter(
                timestamp__date__range=(loss_start_date, loss_end_date)
            )

        # Prepare data for the template
        loss_data = [
            {
                'item_name': record.inventory_item.item.item_name,
                'item_type': record.inventory_item.item.itemType.itemType_name,
                'inventory_item_id': record.inventory_item.inventory_item_id,
                'qty': record.qty,
                'timestamp': record.timestamp,
                'remarks': record.remarks,
                'updated_by': record.updated_by if record.updated_by else 'Unknown'
            }
            for record in loss_data_query
        ]


        context = {
            'inventory_qty_data': inventory_qty_data,
            'item_types_list': item_types_list,
            'laboratory_id': selected_laboratory_id,

            'trend_data_added': json.dumps(trend_data_added),
            'trend_data_removed': json.dumps(trend_data_removed),
            'items': items,
            'selected_item_id': item_id,


            'total_purchased': total_purchased,
            'total_used': abs(total_used),  # Convert to positive number
            'purchased_filter_display': purchased_filter_type.replace("_", " ").title(),
            'used_filter_display': used_filter_type.replace("_", " ").title(),
            'expired_items_qty': expired_items_qty,

            'supplier_data': supplier_data,
            'supplier_item_id': supplier_item_id,
            'start_date': start_date,
            'end_date': end_date,
            'filter_type': supplier_filter_type,

            'loss_data': loss_data,
            'loss_filter_type': loss_filter_type,
            'loss_start_date': loss_start_date,
            'loss_end_date': loss_end_date,
        }

        return render(request, 'mod_reports/inventory_reports.html', context)
    except ValueError as e:
        logger.error(f"Value error in inventory_reports: {e}")
        messages.error(request, str(e))
        return redirect('home')

    except DatabaseError as e:
        logger.error(f"Database error in inventory_reports: {e}", exc_info=True)
        messages.error(request, "Database error occurred. Please try again later.")
        return redirect('home')

    except Exception as e:
        logger.error(f"Unexpected error in inventory_reports: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred while loading inventory reports.")
        return redirect('home')

def borrowing_reports(request):
    try:
        selected_laboratory_id = request.session.get('selected_lab')
        current_date = date.today()
        one_year_ago = current_date - timedelta(days=365)

        # Query to get the total number of borrows per day with status 'B', 'X', 'Y' in the last year
        borrows_per_day = borrow_info.objects.filter(
            status__in=['B', 'X', 'Y'],
            borrow_date__range=[one_year_ago, current_date],
            laboratory_id=selected_laboratory_id
        ).values('borrow_date').annotate(total_borrows=Count('borrow_id')).order_by('borrow_date')

        # Prepare data for the chart
        if borrows_per_day:
            start_date = one_year_ago
            end_date = current_date
            date_range = (end_date - start_date).days + 1
            all_dates = {start_date + timedelta(days=i): 0 for i in range(date_range)}

            # Update the dictionary with actual counts
            for entry in borrows_per_day:
                all_dates[entry['borrow_date']] = entry['total_borrows']

            # Convert data to JSON format for ApexCharts
            # day_borrowing_data = {
            #     'dates': [date.strftime('%Y-%m-%d') for date in all_dates.keys()],
            #     'counts': list(all_dates.values())
            # }
            # day_borrowing_data = {
            #     'dates': [date.isoformat() for date in all_dates.keys()],
            #     'counts': list(all_dates.values())
            # }
            day_borrowing_data = {
                'dates': [date.strftime('%Y-%m-%d') for date in all_dates.keys()],
                'counts': list(all_dates.values())
            }

        else:
            day_borrowing_data = {
                'dates': [],
                'counts': []
            }
        
        # print(day_borrowing_data)

        # borrow table ============================
        # Get filter parameters from request
        filter_type = request.GET.get('filter_type', 'today')
        start_date, end_date = calculate_date_range(request, filter_type)

        # Get the current date
        current_date = date.today()

        # Query to get the total quantity of items borrowed for each item in a specific laboratory
        borrowed_items_data = borrowed_items.objects.filter(
            borrow__status__in=['B', 'X', 'Y'],
            borrow__borrow_date__range= (start_date, end_date),
            borrow__laboratory_id=selected_laboratory_id
        ).values('item__item_id','item__item_name', 'item__itemType__itemType_name').annotate(total_qty=Sum('qty')).order_by('item__item_name')

        
        # New Query for borrow_info (borrowing requests)
        borrowreq_filter_type = request.GET.get('borrowreq_filter_type', 'today')
        borrowreq_start_date, borrowreq_end_date = calculate_date_range(request, borrowreq_filter_type)

        borrow_requests_data = borrow_info.objects.filter(
            borrow_date__range=(borrowreq_start_date, borrowreq_end_date),
            laboratory_id=selected_laboratory_id
        ).order_by('borrow_date')

        context = {
            'day_borrowing_data': json.dumps(day_borrowing_data),

            'borrowed_items_data': borrowed_items_data,
            'filter_type': filter_type,
            'start_date': start_date,
            'end_date': end_date,

            'borrowreq_filter_type': borrowreq_filter_type,
            'borrowreq_start_date': borrowreq_start_date,
            'borrowreq_end_date': borrowreq_end_date,

            'borrow_requests_data': borrow_requests_data
        }

        return render(request, 'mod_reports/borrowing_reports.html', context)
    except ValueError as e:
        logger.error(f"Value error in borrowing_reports: {e}")
        messages.error(request, str(e))
        return redirect('home')

    except DatabaseError as e:
        logger.error(f"Database error in borrowing_reports: {e}", exc_info=True)
        messages.error(request, "Database error occurred. Please try again later.")
        return redirect('home')

    except Exception as e:
        logger.error(f"Unexpected error in borrowing_reports: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred while loading borrowing reports.")
        return redirect('home')

def clearance_reports(request):
    try:
        # Get filter type from GET parameters, defaulting to 'today'
        selected_laboratory_id = request.session.get('selected_lab')
        filter_type = request.GET.get('filter_type', 'today')
        reports_filter = request.GET.get('reports_filter', 'today')
        start_date = end_date = None
        today = timezone.localtime().date()
        
        if filter_type == 'today':
            start_date = timezone.localtime().replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = timezone.localtime().replace(hour=23, minute=59, second=59, microsecond=999999)
        elif filter_type == 'this_week':
            start_date = today - timedelta(days=today.weekday())  # Start of the week
            end_date = timezone.localtime().replace(hour=23, minute=59, second=59, microsecond=999999)
        elif filter_type == 'this_month':
            start_date = today.replace(day=1)
            end_date = timezone.localtime().replace(hour=23, minute=59, second=59, microsecond=999999)
        elif filter_type == 'this_year':
            start_date = today.replace(month=1, day=1)
            end_date = timezone.localtime().replace(hour=23, minute=59, second=59, microsecond=999999)
        elif filter_type == 'custom':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            if start_date:
                start_date = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d')).replace(hour=0, minute=0, second=0, microsecond=0)
            if end_date:
                end_date = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d')).replace(hour=23, minute=59, second=59, microsecond=999999)

        date_range = [start_date, end_date]
        # print(start_date, ' - ', end_date)

        # Calculate the on-hold user count within the laboratory and date range
        on_hold_users_count = reported_items.objects.filter(
            laboratory_id=selected_laboratory_id,
            status=1,  # Only include pending reports
            user_id__isnull=False
        ).values('user_id').distinct().count()

        # Set incidents_start_date based on reports_filter
        if reports_filter == 'today':
            incidents_start_date = today
        elif reports_filter == 'this_week':
            incidents_start_date = today - timedelta(days=today.weekday())
        elif reports_filter == 'this_month':
            incidents_start_date = today.replace(day=1)
        elif reports_filter == 'this_year':
            incidents_start_date = today.replace(month=1, day=1)
        else:
            incidents_start_date = None

        # Calculate total reports count in the lab based on incidents_start_date
        total_reports_count = reported_items.objects.filter(
            laboratory_id=selected_laboratory_id,
            reported_date__date__gte=incidents_start_date if incidents_start_date else None
        ).count()

        # If the request is AJAX, return JSON with total_reports_count and filter display
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'total_reports_count': total_reports_count,
                'reports_filter_display': reports_filter.replace("_", " ").title()
            })



        # Filter reported_items within the date range and specific laboratory
        clearance_data = reported_items.objects.filter(
            laboratory_id=selected_laboratory_id,
            reported_date__range=date_range,
        ).values(
            reported_user_id=F('user__user_id'),
            user_name=Concat(F('user__firstname'), Value(' '), F('user__lastname')),
            personal_id=F('user__personal_id')
        ).annotate(
            total_amount_due=Sum('amount_to_pay'),
            reported_items_count=Count('report_id'),
            status=Sum('status'),
        )

        # Add clearance status for each user
        for record in clearance_data:
            record['clearance_status'] = 'Cleared' if record['status'] == 0 else 'On Hold'

        
        # Fetch reported items data for the specified laboratory and filter criteria
        # Filters for Reported Items Count
        item_filter_type = request.GET.get('item_filter_type', 'today')
        item_start_date = item_end_date = None
        today = timezone.localtime().date()
        
        if item_filter_type == 'today':
            item_start_date = timezone.localtime().replace(hour=0, minute=0, second=0, microsecond=0)
            item_end_date = timezone.localtime().replace(hour=23, minute=59, second=59, microsecond=999999)
        elif item_filter_type == 'this_week':
            item_start_date = today - timedelta(days=today.weekday())
            item_end_date = timezone.localtime().replace(hour=23, minute=59, second=59, microsecond=999999)
        elif item_filter_type == 'this_month':
            item_start_date = today.replace(day=1)
            item_end_date = timezone.localtime().replace(hour=23, minute=59, second=59, microsecond=999999)
        elif item_filter_type == 'this_year':
            item_start_date = today.replace(month=1, day=1)
            item_end_date = timezone.localtime().replace(hour=23, minute=59, second=59, microsecond=999999)
        elif item_filter_type == 'custom':
            item_start_date = request.GET.get('item_start_date')
            item_end_date = request.GET.get('item_end_date')
            if item_start_date:
                item_start_date = timezone.make_aware(datetime.strptime(item_start_date, '%Y-%m-%d')).replace(
                    hour=0, minute=0, second=0, microsecond=0)
            if item_end_date:
                item_end_date = timezone.make_aware(datetime.strptime(item_end_date, '%Y-%m-%d')).replace(
                    hour=23, minute=59, second=59, microsecond=999999)

        item_date_range = [item_start_date, item_end_date]

        # Fetch Reported Items Count data
        reported_items_summary = reported_items.objects.filter(
            laboratory_id=selected_laboratory_id,
            reported_date__range=item_date_range if item_start_date and item_end_date else None
        ).values(
            reported_item_id=F('item__item_id'),
            item_name=F('item__item_name')
        ).annotate(
            report_count=Count('report_id'),
            total_qty_reported=Sum('qty_reported')
        )


        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'total_reports_count': total_reports_count,
                'reports_filter_display': context['reports_filter_display']
            })    

        context = {
            'reported_items_data': clearance_data,

            'filter_type': filter_type,
            'start_date': start_date,
            'end_date': end_date,

            'total_on_hold_users': on_hold_users_count,
            'total_reports_count': total_reports_count,
            'reports_filter_display': reports_filter.replace("_", " ").title(),

            'reported_items_summary': reported_items_summary,
            'item_filter_type': item_filter_type,
            'item_start_date': item_start_date,
            'item_end_date': item_end_date,

        }

        return render(request, 'mod_reports/clearance_reports.html', context)
    except Exception as e:
        logger.error(f"Unexpected error in clearance_reports: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred while loading clearance reports.")
        return redirect('home')

def labres_reports(request):
    try:
        selected_laboratory_id = request.session.get('selected_lab')
        reservations_filter = request.GET.get('reservations_filter', 'today')
        today = timezone.localtime().date()

        # Get total rooms for the selected laboratory
        total_rooms = rooms.objects.filter(laboratory_id=selected_laboratory_id, is_disabled=0).count()
        reservable_rooms = rooms.objects.filter(laboratory_id=selected_laboratory_id, is_reservable=1, is_disabled=0).count()

        # Set start_date based on filter type
        if reservations_filter == 'today':
            start_date = today
        elif reservations_filter == 'this_week':
            start_date = today - timedelta(days=today.weekday())  # Start of the week
        elif reservations_filter == 'this_month':
            start_date = today.replace(day=1)  # Start of the month
        elif reservations_filter == 'this_year':
            start_date = today.replace(month=1, day=1)  # Start of the year
        else:
            start_date = None  # No filter applied

        # Total number of reservations for selected laboratory with optional filtering
        lab_reservations_qs = laboratory_reservations.objects.filter(
            room__laboratory_id=selected_laboratory_id
        )
        if start_date:
            lab_reservations_qs = lab_reservations_qs.filter(start_date__gte=start_date)
        total_reservations = lab_reservations_qs.count()


        reservation_filter_type = request.GET.get('reservation_filter_type', 'today')
        start_date = end_date = None
        start_date, end_date = calculate_date_range(request, reservation_filter_type)
        lab_reservations_qs = laboratory_reservations.objects.filter(
            laboratory_id=selected_laboratory_id,
            status='R'
        )
        if start_date and end_date:
            lab_reservations_qs = lab_reservations_qs.filter(start_date__range=(start_date, end_date))


        # Room Reservation Filter
        room_filter_type = request.GET.get('room_filter_type', 'today')
        room_start_date, room_end_date = calculate_date_range(request, room_filter_type)

        # Query for room reservation summary with filter
        room_data = rooms.objects.filter(laboratory_id=selected_laboratory_id, is_disabled=False).annotate(
            total_reservations=Count(
                'reservations', 
                filter=Q(reservations__start_date__range=(room_start_date, room_end_date))
            ),
            cumulative_time=Sum(
                ExpressionWrapper(
                    (F('reservations__end_time__hour') * 60 + F('reservations__end_time__minute')) -
                    (F('reservations__start_time__hour') * 60 + F('reservations__start_time__minute')),
                    output_field=IntegerField()
                ),
                filter=Q(reservations__start_date__range=(room_start_date, room_end_date))
            ) / 60.0  # Convert minutes to hours
        )

        # Replace None with 0 for cumulative_time
        for room in room_data:
            if room.cumulative_time is None:
                room.cumulative_time = 0

        context = {
            'total_rooms': total_rooms,
            'reservable_rooms': reservable_rooms,
            'total_reservations': total_reservations,
            'reports_filter_display': reservations_filter.replace("_", " ").title(),

            'laboratory_reservations': lab_reservations_qs,
            'reservation_filter_type': reservation_filter_type,
            'start_date': start_date,
            'end_date': end_date,

            'room_data': room_data,
            'room_filter_type': room_filter_type,
            'room_start_date': room_start_date,
            'room_end_date': room_end_date,
        }

        # If AJAX request, return JSON response for dynamic filter update
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'total_reservations': total_reservations,
                'reports_filter_display': context['reports_filter_display']
            })

        return render(request, 'mod_reports/labres_reports.html', context)
    except ValueError as e:
        logger.error(f"Value error in labres_reports: {e}")
        messages.error(request, str(e))
        return redirect('home')

    except DatabaseError as e:
        logger.error(f"Database error in labres_reports: {e}", exc_info=True)
        messages.error(request, "Database error occurred. Please try again later.")
        return redirect('home')

    except Exception as e:
        logger.error(f"Unexpected error in labres_reports: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred while loading laboratory reservation reports.")
        return redirect('home')

def inventory_data(request, item_type_id, laboratory_id):
    try:
        item_type = item_types.objects.get(pk=item_type_id)
        items = item_description.objects.filter(
            itemType_id=item_type_id,
            laboratory_id=laboratory_id,
            is_disabled=False  # Only get items that are enabled
        ).annotate(total_qty=Coalesce(Sum('item_inventory__qty'), 0))  # Calculate total quantity

        add_cols = json.loads(item_type.add_cols)

        items_data = []
        for item in items:
            item_data = {
                'item_id': item.item_id,
                'item_name': item.item_name,
                'alert_qty': item.alert_qty,
                'rec_expiration': item.expiry_type,
                'allow_borrow': item.allow_borrow,
                'is_consumable': item.is_consumable,
                'total_qty': item.total_qty,
                'add_cols': json.loads(item.add_cols) if item.add_cols else {}
            }
            items_data.append(item_data)

        return JsonResponse({
            'items': items_data,
            'add_cols': add_cols
        })

    except DatabaseError as e:
        logger.error(f"Database error in inventory_data: {e}", exc_info=True)
        return JsonResponse({'error': "Database error occurred. Please try again."}, status=500)

    except Exception as e:
        logger.error(f"Unexpected error in inventory_data: {e}", exc_info=True)
        return JsonResponse({'error': "An unexpected error occurred while fetching inventory data."}, status=500)

def calculate_date_range(request, filter_type):
    """ Helper function to determine start and end dates based on the filter type """
    try:
        today = timezone.localtime().date()
        if filter_type == 'today':
            return today, today
        elif filter_type == 'this_week':
            start_date = today - timedelta(days=today.weekday())
            return start_date, today
        elif filter_type == 'this_month':
            start_date = today.replace(day=1)
            return start_date, today
        elif filter_type == 'this_year':
            start_date = today.replace(month=1, day=1)
            return start_date, today
        elif filter_type == 'custom':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            if start_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            if end_date:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            return start_date, end_date
        return None, None
        
    except Exception as e:
        logger.error(f"Unexpected error in calculate_date_range: {e}", exc_info=True)
        return today, today  # Default to today on unexpected error

@user_passes_test(lambda u: u.is_superuser)
def admin_reports_view(request):
    try:
        # Get separate filters for each card from GET parameters
        total_users_filter = request.GET.get('total_users_filter', 'this_year')
        new_users_filter = request.GET.get('new_users_filter', 'today')
        reports_filter = request.GET.get('reports_filter', 'this_week')

        today = timezone.localtime().date()
        # # Total Users Filter Setup
        # if total_users_filter == 'today':
        #     total_start_date = today
        # elif total_users_filter == 'this_month':
        #     total_start_date = today.replace(day=1)
        # elif total_users_filter == 'this_year':
        #     total_start_date = today.replace(month=1, day=1)
        # else:
        #     total_start_date = today - timedelta(weeks=1)

        # Active Users Count
        total_active_users = user.objects.filter(is_deactivated=False).count()
        total_active_labs = laboratory.objects.filter(is_available=True).count()

        # New Users Filter Setup (default to 'today' for new users card)
        if new_users_filter == 'today':
            new_users_filter = 'Today'
            new_start_date = today
        elif new_users_filter == 'this_month':
            new_users_filter = 'This Month'
            new_start_date = today.replace(day=1)
        elif new_users_filter == 'this_year':
            new_users_filter = 'This Year'
            new_start_date = today.replace(month=1, day=1)
        else:
            new_users_filter = 'This Week'
            new_start_date = today - timedelta(weeks=1)

        # New Users Count
        new_users = user.objects.filter(date_joined__date__gte=new_start_date).count()

        # Reports Filter Setup
        if reports_filter == 'this_week':
            reports_filter_display = 'This Week'
            reports_start_date = today - timedelta(days=today.weekday())  # Start of the current week
            date_range = [reports_start_date + timedelta(days=i) for i in range(7)]
            date_format = "%Y-%m-%d"
        elif reports_filter == 'this_month':
            reports_filter_display = 'This Month'
            reports_start_date = today.replace(day=1)
            last_day = (today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            date_range = [reports_start_date + timedelta(days=i) for i in range((last_day - reports_start_date).days + 1)]
            date_format = "%Y-%m-%d"
        else:  # 'this_year'
            reports_filter_display = 'This Year'
            reports_start_date = today.replace(month=1, day=1)
            date_range = [datetime(today.year, month, 1).date() for month in range(1, 13)]
            date_format = "%Y-%m"

        # Count New Users for Each Date in Range (for Reports Card Line Chart)
        reports_data = []
        for date in date_range:
            if reports_filter == 'this_year':
                user_count = user.objects.filter(
                    date_joined__year=date.year,
                    date_joined__month=date.month
                ).count()
            else:
                user_count = user.objects.filter(date_joined__date=date).count()
            reports_data.append({'date': date.strftime(date_format), 'count': user_count})

        print(reports_data)

        # Retrieve data for roles (separating default and others)
        default_roles = laboratory_roles.objects.filter(laboratory_id=0).values('roles_id', 'name').exclude(roles_id=0)
        default_role_ids = [role['roles_id'] for role in default_roles]

        # Retrieve labs and user counts per role for stacked chart
        user_labs_data = []
        labs = laboratory.objects.exclude(Q(laboratory_id=0) | Q(is_available=0)).annotate(user_count=Count('laboratory_users'))
        for lab in labs:
            lab_roles_data = {'lab_name': lab.name}
            
            # Count users for each default role in this laboratory
            for role in default_roles:
                role_count = laboratory_users.objects.filter(
                    laboratory=lab, role_id=role['roles_id'], is_active=True
                ).count()
                lab_roles_data[role['name']] = role_count
            
            # Count users with roles not in the default roles as "Others"
            others_count = laboratory_users.objects.filter(
                laboratory=lab, is_active=True
            ).exclude(role_id__in=default_role_ids).count()
            lab_roles_data['Others'] = others_count
            
            user_labs_data.append(lab_roles_data)
            
        


        context = {
            'total_active_users': total_active_users,
            'total_active_labs': total_active_labs,
            'new_users': new_users,
            'filter_type': {
                'total_users': total_users_filter,
                'new_users': new_users_filter,
                'reports': reports_filter
            },
            'reports_data': reports_data,
            'reports_filter_display': reports_filter_display,
            
            'user_labs_data': user_labs_data,
            'default_roles': default_roles,

            'labs': labs
        }
        return render(request, 'mod_reports/admin_reports.html', context)
    except DatabaseError as e:
        logger.error(f"Database error in admin_reports_view: {e}", exc_info=True)
        messages.error(request, "Database error occurred. Please try again later.")
        return redirect('home')

    except Exception as e:
        logger.error(f"Unexpected error in admin_reports_view: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred while loading admin reports.")
        return redirect('home')


#  ================================================================= 
#lab setup
@user_passes_test(lambda u: u.is_superuser)
def superuser_manage_labs(request):
    try:
        if not request.user.is_superuser:
            return render(request, 'error_page.html', {'message': 'Module is allowed for this laboratory.'})

        labs = laboratory.objects.exclude(Q(laboratory_id=0) | Q(is_available=0)).annotate(user_count=Count('laboratory_users'))   # Retrieve all laboratory records
        context = {
            'labs': labs,
        }
        return render(request, 'superuser/superuser_manageLabs.html', context)
    except Exception as e:
        logger.error(f"Unexpected error in superuser_lab_info: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('home')

# lab info
@transaction.atomic
@superuser_or_lab_permission_required('configure_laboratory')
def superuser_lab_info(request, laboratory_id):
    lab = get_object_or_404(laboratory, laboratory_id=laboratory_id)
    lab_rooms = rooms.objects.filter(laboratory_id=lab.laboratory_id, is_disabled=False)

    try:
        if request.method == "POST":
            if 'save_rooms' in request.POST:
                # Update is_reservable field for rooms
                for room in rooms.objects.filter(laboratory_id=lab):
                    room.is_reservable = f'room_{room.room_id}_enabled' in request.POST
                    room.save()
                messages.success(request, 'Successfully saved room configuration')

            elif 'delete_room' in request.POST:
                delete_room_id = request.POST.get('delete_room')
                room = get_object_or_404(rooms, room_id=delete_room_id)
                room.is_disabled = True
                room.save()
                messages.success(request,  'Successfully deleted a room')

            elif 'add_room' in request.POST:
                room_name = request.POST.get('room_name')
                room_capacity = request.POST.get('room_capacity')
                room_description = request.POST.get('room_description')

                if room_name and room_capacity:
                    new_room = rooms(
                        laboratory=lab,
                        name=room_name,
                        capacity=room_capacity,
                        description=room_description
                    )
                    new_room.save()
                messages.success( request, f'Successfully added Room: {room_name}')
            elif 'edit_room' in request.POST:
                room_id = request.POST.get('edit_room')
                room = get_object_or_404(rooms, room_id=room_id)
                room.name = request.POST.get('room_name')
                room.capacity = request.POST.get('room_capacity')
                room.description = request.POST.get('room_description')
                room.save()
                messages.success(request, 'Successfully updated the room')
                
            elif 'add_role' in request.POST:
                role_name = request.POST.get('role_name')
                if role_name:
                    new_role = laboratory_roles(
                        laboratory=lab,
                        name=role_name,
                    )
                    new_role.save()
                messages.success( request, f'Successfully added Role: {role_name}')
            elif 'edit_role' in request.POST:
                role_name = request.POST.get('role_name')
                role_id = request.POST.get('edit_role')
                role = get_object_or_404(laboratory_roles, roles_id=role_id)
                role.name = role_name
                role.save()
                messages.success( request, f'Successfully Edited Role: {role_name}')
            elif 'delete_role' in request.POST:
                role_id = request.POST.get('delete_role')
                role = get_object_or_404(laboratory_roles, roles_id=role_id)
                role.delete()
                messages.success(request,  'Successfully deleted a role')

            elif 'accept_user' in request.POST:
                user_id = request.POST.get("user_id")
                lab_user = laboratory_users.objects.filter(user_id=user_id, laboratory_id=laboratory_id, status='P').first()
                if lab_user:
                    lab_user.status = 'A'  # Accepted
                    lab_user.save()
                    messages.success(request, f"User {lab_user.user.get_fullname()} accepted successfully.")

            elif 'decline_user' in request.POST:
                user_id = request.POST.get("user_id")
                lab_user = laboratory_users.objects.filter(user_id=user_id, laboratory_id=laboratory_id, status='P').first()
                if lab_user:
                    lab_user.status = 'D'  # Declined
                    lab_user.is_active = False  # Declined
                    lab_user.save()
                    messages.success(request, f"User {lab_user.user.get_fullname()} declined.")
            
            elif 'toggle_user_status' in request.POST:
                user_id = request.POST.get('user_id')
                lab_user = get_object_or_404(laboratory_users, user_id=user_id, laboratory_id=laboratory_id, is_active=1)
                lab_user.is_active = not lab_user.is_active
                lab_user.save()
                
                # Add feedback message based on new status
                status_message = "activated" if lab_user.is_active else "deactivated"
                messages.success(request, f"User {lab_user.user.get_fullname()} successfully {status_message}.")

            # Existing code for room and role operations goes here...

    
        
        # Get modules active in the lab
        active_module_ids = lab.modules
        modules = Module.objects.filter(id__in=active_module_ids)
        
        # Filter permissions by active modules
        permissions_by_module = {}
        for module in modules:
            permissions_by_module[module.id] = permissions.objects.filter(module=module)

        all_modules = Module.objects.all()  # All available modules

        # Retrieve all lab users and roles
        lab_users = laboratory_users.objects.filter(laboratory_id=lab.laboratory_id, is_active=1, status__in=['A', 'I']).select_related('user', 'role').annotate(
            username=F('user__username'),
            user_email=F('user__email'),
            full_name=Concat(F('user__firstname'), Value(' '), F('user__lastname'), output_field=CharField()),
            role_name=F('role__name')
        )
        lab_roles = laboratory_roles.objects.filter(Q(laboratory_id=0) | Q(laboratory_id=lab.laboratory_id)).annotate(
            usercount=Count('users', filter=Q(users__laboratory_id=laboratory_id))
        )

        print(lab_roles)

        # Retrieve pending users for display in the "Share" tab
        pending_users = laboratory_users.objects.filter(
            laboratory_id=laboratory_id, status='P'
        ).select_related('user', 'role').annotate(
            full_name=Concat(F('user__firstname'), Value(' '), F('user__lastname'), output_field=CharField()),
            user_email=F('user__email'),
            personal_id=F('user__personal_id')
        )
        
        # Current permissions for each role in the lab
        role_permissions = {(perm.role.roles_id, perm.permissions.codename): True 
                            for perm in laboratory_permissions.objects.filter(laboratory=lab)}
        # print(role_permissions)
        
        context = {
            'laboratory_id': laboratory_id,
            'lab': lab,
            'modules': modules,
            'all_modules': all_modules,
            'lab_rooms': lab_rooms,
            'lab_users': lab_users,
            'lab_roles': lab_roles,
            'permissions_by_module': permissions_by_module,
            'role_permissions': role_permissions,
            'pending_users': pending_users
        }

        return render(request, 'superuser/superuser_labInfo.html', context)

    except Exception as e:
        logger.error(f"Unexpected error in superuser_lab_info: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('home')


# @login_required()
# @superuser_or_lab_permission_required('configure_laboratory')
# def add_module_to_lab(request, laboratory_id):
#     if request.method == 'POST':
#         module_id = request.POST.get('module_id')
#         lab = get_object_or_404(laboratory, laboratory_id=laboratory_id)
#         module = get_object_or_404(Module, id=module_id)

#         # Check if the module is already associated with the lab
#         if LaboratoryModule.objects.filter(laboratory=lab, module=module).exists():
#             messages.error(request, "The module is already added to this laboratory.")
#         else:
#             # Create a LaboratoryModule entry
#             LaboratoryModule.objects.get_or_create(laboratory=lab, module=module)
#             messages.success(request, "Module added successfully.")

#         return redirect('superuser_lab_info', laboratory_id=laboratory_id)

#     return redirect('superuser_lab_info', laboratory_id=laboratory_id)

@transaction.atomic
@login_required()
@superuser_or_lab_permission_required('configure_laboratory')
def toggle_module_status(request, laboratory_id):
    try:
        if request.method == 'POST':
            lab = get_object_or_404(laboratory, laboratory_id=laboratory_id)
            module_ids = []

            for module in Module.objects.all():
                if request.POST.get(f'module_{module.id}'):
                    module_ids.append(module.id)

            lab.modules = module_ids
            lab.save()

            messages.success(request, "Module statuses updated successfully.")
            return redirect('superuser_lab_info', laboratory_id=laboratory_id)
    except Exception as e:
        logger.error(f"Error updating module statuses for Lab {laboratory_id}: {e}", exc_info=True)
        messages.error(request, "Failed to update module statuses. Please try again.")
        return redirect('superuser_lab_info', laboratory_id=laboratory_id)

@transaction.atomic
@login_required()
@superuser_or_lab_permission_required('configure_laboratory')
def edit_lab_info(request, laboratory_id):
    try:
        if request.method == 'POST':
            lab = get_object_or_404(laboratory, laboratory_id=laboratory_id)
            lab.name = request.POST.get('name')
            lab.description = request.POST.get('description')
            lab.department = request.POST.get('department')

            if not lab.name:
                return JsonResponse({'status': 'fail', 'message': 'Lab name is required'}, status=400)

            lab.save()
            return JsonResponse({'status': 'success'})
        return JsonResponse({'status': 'fail'}, status=400)
    except Exception as e:
        logger.error(f"Error updating lab info for Lab {laboratory_id}: {e}", exc_info=True)
        return JsonResponse({'status': 'fail', 'message': 'An error occurred while updating lab info'}, status=500)

@transaction.atomic
@login_required()
@superuser_or_lab_permission_required('configure_laboratory')
def deactivate_lab(request, laboratory_id):
    try:
        lab = get_object_or_404(laboratory, laboratory_id=laboratory_id)
        lab.is_available = False  # Set is_available to 0 (inactive)
        lab.save()  # Save the changes to the database
        messages.success(request, "Laboratory deactivated successfully.")  # Optional success message
    except Exception as e:
        logger.error(f"Error deactivating Lab {laboratory_id}: {e}", exc_info=True)
        messages.error(request, "Failed to deactivate laboratory. Please try again.")
    
    return redirect('superuser_manage_labs')


@login_required()
@superuser_or_lab_permission_required('configure_laboratory')
def update_permissions(request, laboratory_id):
    try:
        if request.method == 'POST':
            lab = get_object_or_404(laboratory, laboratory_id=laboratory_id)

            # Store permissions that were checked in the form submission
            permissions_data = {}
            for key, value in request.POST.items():
                if key.startswith('permissions'):
                    role_id = int(key.split('[')[1].split(']')[0])
                    perm_codename = key.split('[')[2].split(']')[0]
                    
                    if role_id not in permissions_data:
                        permissions_data[role_id] = {}
                    permissions_data[role_id][perm_codename] = value == 'on'

            # Query all existing permissions for this lab and role
            existing_permissions = laboratory_permissions.objects.filter(laboratory=lab)
            
            # Determine which permissions to keep and which to delete
            for perm_entry in existing_permissions:
                role_id = perm_entry.role.roles_id
                perm_codename = perm_entry.permissions.codename

                # If the permission is missing from permissions_data, delete it
                if not permissions_data.get(role_id, {}).get(perm_codename, False):
                    perm_entry.delete()

            # Add or update permissions based on the form data
            for role_id, perms in permissions_data.items():
                role = laboratory_roles.objects.get(roles_id=role_id)
                for perm_codename, is_selected in perms.items():
                    perm_obj = permissions.objects.get(codename=perm_codename)

                    if is_selected:
                        # Create or update permission if it’s checked
                        laboratory_permissions.objects.update_or_create(
                            role=role, laboratory=lab, permissions=perm_obj
                        )

            messages.success(request, "Permissions updated successfully.")
        return redirect('superuser_lab_info', laboratory_id=laboratory_id)
    except Exception as e:
        logger.error(f"Error updating permissions for Lab {laboratory_id}: {e}", exc_info=True)
        messages.error(request, "Failed to update permissions. Please try again.")
        return redirect('superuser_lab_info', laboratory_id=laboratory_id)


# users
@transaction.atomic
@login_required
@require_POST
@superuser_or_lab_permission_required('configure_laboratory')
def edit_user_role(request, laboratory_id):
    try: 
        user_id = request.POST.get('user_id')
        new_role_id = request.POST.get('role_id')
        new_status = request.POST.get('Status')
        if not all([user_id, new_role_id, new_status]):
                messages.error(request, "Missing required fields.")
                return redirect('superuser_lab_info', laboratory_id)
                
        lab_user = get_object_or_404(laboratory_users, user_id=user_id, laboratory_id=laboratory_id, is_active=1)
        lab_user.role_id = new_role_id
        lab_user.status = new_status
        lab_user.save()
        messages.success(request, "User role updated successfully.")
    except Exception as e:
        logger.error(f"Error updating user role - edit_user_role: {e}", exc_info=True)
        messages.error(request, f"Error updating user role: {e}")
        print(f"Error updating user role: {e}")

    return redirect ('superuser_lab_info', laboratory_id)

@transaction.atomic
@login_required
@require_POST
@superuser_or_lab_permission_required('configure_laboratory')
def toggle_user_status(request):
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')

        if not user_id:
            return JsonResponse({'success': False, 'message': "Invalid user ID"}, status=400)

        lab_user = get_object_or_404(laboratory_users, user_id=user_id)
        lab_user.is_active = not lab_user.is_active
        lab_user.save()
        messages.success(request, "User status updated successfully.")
        return JsonResponse({'success': True, 'is_active': lab_user.is_active})
    except json.JSONDecodeError:
        logger.error(f"Error json decoder - toggle_user_status: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': "Invalid JSON format"}, status=400)
    except Exception as e:
        logger.error(f"Error toggling user status - toggle_user_status: {e}", exc_info=True)
        print(f"Error toggling user status: {e}")
        return JsonResponse({'success': False, 'message': "An error occurred"}, status=500)

@login_required()
@user_passes_test(lambda u: u.is_superuser)
def bulk_upload_users_to_lab(request, laboratory_id):
    if request.method == 'POST' and request.FILES.get('userfile'):
        userfile = request.FILES['userfile']

        try:
            workbook = openpyxl.load_workbook(userfile)
            sheet = workbook.active

            # Expected format: Email, Role
            required_columns = ['email address', 'role']

            header = [
                str(cell.value).strip().lower()
                for cell in sheet[1] if cell.value  # Filter out empty columns
            ]

            if not set([col.lower() for col in required_columns]).issubset(header):
                messages.error(request, f"Invalid format. Expected columns: {', '.join(required_columns)}")
                return redirect('superuser_manage_labs')

            roles_dict = {role.name.strip().lower(): role.roles_id for role in laboratory_roles.objects.all()}

            error_messages = []
            success_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row = [cell for cell in row if cell]
                if len(row) < 2:
                    continue

                email, role_name = row[:2]

                if not email or not role_name:
                    continue

                role_id = roles_dict.get(role_name.strip().lower())
                if not role_id:
                    error_messages.append(f"Invalid Role '{role_name}' for email {email}. Skipping.")
                    continue

                try:
                    user_instance = user.objects.get(email=email)
                except user.DoesNotExist:
                    error_messages.append(f"User with email {email} does not exist. Skipping.")
                    continue

                user_assigned = laboratory_users.objects.filter(
                    user_id=user_instance.user_id,
                    laboratory_id=laboratory_id,
                    is_active=1
                ).exists()

                if not user_assigned:
                    laboratory_users.objects.create(
                        user_id=user_instance.user_id,
                        laboratory_id=laboratory_id,
                        role_id=role_id,
                        is_active=1,
                        status='A',
                    )
                    success_count += 1
                else:
                    error_messages.append(f"User {email} is already assigned to this laboratory with role {role_name}.")

            if success_count > 0:
                messages.success(request, f"{success_count} user(s) added successfully.")
            if error_messages:
                # Use <br> to separate the messages
                formatted_errors = "<br>".join(error_messages)
                messages.warning(request, formatted_errors)

        except openpyxl.utils.exceptions.InvalidFileException:
            messages.error(request, "Invalid file format. Please upload a valid Excel file.")
        except Exception as e:
            logger.error(f"Error processing the file - bulk_upload_users_to_lab: {e}", exc_info=True)
            messages.error(request, f"Error processing the file: {e}")
            print(f"Error processing the file: {e}")

    return redirect('superuser_lab_info', laboratory_id=laboratory_id)


@superuser_or_lab_permission_required('configure_laboratory')
def add_user_laboratory(request, laboratory_id):
    try:
        if request.method == "POST":
            user_id = request.POST['user']
            role_id = request.POST['role']
            status = request.POST['Status']

            # lab = get_object_or_404(laboratory, laboratory_id=laboratory_id)
            # user_instance = get_object_or_404(user, user_id=user_id)
            # role_instance = get_object_or_404(laboratory_roles, roles_id=role_id)
            
            user_assigned = laboratory_users.objects.filter(
                user_id=user_id,
                laboratory_id=laboratory_id,
                is_active=1
            ).exists()

            if not user_assigned:
                # Add user to laboratory if not already assigned
                laboratory_users.objects.create(
                    user_id=user_id,
                    laboratory_id=laboratory_id,
                    role_id=role_id,
                    is_active=1,
                    status=status,
                )
                messages.success(request, 'User  added successfully')
            else:
                messages.error(request, 'User  is already assigned to this laboratory and is active.')

        return redirect('superuser_lab_info', laboratory_id=laboratory_id)
    except Exception as e:
        logger.error(f"Error adding user to laboratory to laboratory - add_user_laboratory: {e}", exc_info=True)
        messages.error(request, f"Error adding user to laboratory: {e}")
        print(f"Error adding user to laboratory: {e}")
        return redirect('superuser_lab_info', laboratory_id=laboratory_id)

@login_required()
@user_passes_test(lambda u: u.is_superuser)
def superuser_manage_users(request):
    try:
        users = user.objects.all()
        context = {
            'users': users,
        }
        return render(request, 'superuser/superuser_manageusers.html', context)
    except Exception as e:
        logger.error(f"Unexpected error in superuser_manage_users: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('home')

@login_required()
@user_passes_test(lambda u: u.is_superuser)
def bulk_upload_users(request):
    if request.method == 'POST' and request.FILES.get('userfile'):
        userfile = request.FILES['userfile']

        try:
            # Load the Excel file
            workbook = openpyxl.load_workbook(userfile)
            sheet = workbook.active

            # Expected format: First Name, Last Name, ID Number, Email, Password
            required_columns = ['First Name', 'Last Name', 'ID Number', 'Email', 'Password']
            
            # Sanitize the header: Strip spaces and convert to lowercase for case-insensitivity
            header = [str(cell.value).strip().lower() for cell in sheet[1]]  # Strip spaces and lower-case
            
            # Validate columns by comparing sanitized header with required columns
            sanitized_required_columns = [col.lower() for col in required_columns]  # Lowercase for case-insensitivity
            if header != sanitized_required_columns:
                messages.error(request, f"Invalid format. Expected columns: {', '.join(required_columns)}")
                return redirect('superuser_manage_users')

            # Process each row
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Clean up each value by stripping white spaces and ensuring proper case handling
                firstname, lastname, idnum, email, password = [str(cell).strip() if cell else '' for cell in row]

                # Skip rows with missing data
                if not (firstname and lastname and email and password):
                    continue
                
                # Ensure the ID number is an integer
                try:
                    idnum = int(float(idnum))  # Convert to float first to handle cases like 1234.5, then to int
                except (ValueError, TypeError):
                    messages.warning(request, f"Invalid ID Number format for email {email}. Skipping.")
                    continue
                
                # Check if the user already exists (email is case-insensitive)
                if User.objects.filter(email=email.lower()).exists():
                    messages.warning(request, f"Email {email} is already registered. Skipping.")
                    continue

                # Create user
                try:
                    User.objects.create_user(
                        firstname=firstname,
                        lastname=lastname,
                        personal_id=idnum,
                        email=email,
                        username=email.lower(),  # Ensure username is unique and case-insensitive
                        password=password,
                    )
                except ValidationError as e:
                    messages.warning(request, f"Error creating user {email}: {e}")
                    continue

            messages.success(request, "Bulk user upload completed successfully.")
        except Exception as e:
            logger.error(f"Error processing the file: {e}")
            messages.error(request, f"Error processing the file: {e}")

    return redirect('superuser_manage_users')

@login_required()
@user_passes_test(lambda u: u.is_superuser)
def add_user(request):
    try:
        if request.method == 'POST':
            firstname = request.POST.get('firstname')
            lastname = request.POST.get('lastname')
            idnum = request.POST.get('idnum')
            email = request.POST.get('email')
            username = request.POST.get('email')
            password = request.POST.get('password')
            is_superuser = request.POST.get('is_superuser') == 'on'
            
            if email and firstname and lastname and password:
                if User.objects.filter(email=email).exists():
                    messages.error(request, "Email is already registered.")
                elif User.objects.filter(username=username).exists():
                    messages.error(request, "Username is already taken.")
                else:
                    try:
                        User.objects.create_user(
                            email=email, 
                            firstname=firstname, 
                            lastname=lastname, 
                            password=password, 
                            personal_id=idnum, 
                            username=username, 
                            is_superuser=is_superuser
                        )
                        messages.success(request, "User added successfully.")
                    except Exception as e:
                        messages.error(request, f"Error creating user: {e}")
            else:
                messages.error(request, "Missing required fields.")
    except Exception as e:
        logger.error(f"Unexpected error in adding user: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred. Please try again.")

    return redirect('superuser_manage_users')

@login_required()
@user_passes_test(lambda u: u.is_superuser)
def superuser_user_info(request, user_id):
    try:
        try:
            user1 = get_object_or_404(user, user_id=user_id)
        except Exception as e:
            logger.error(f"User with ID {user_id} not found: {e}")
            messages.error(request, "User not found.")
            return redirect('superuser_manage_users')
        lab_users = laboratory_users.objects.filter(user=user1, is_active=True, laboratory_id__is_available=1)
        assigned_laboratories = lab_users.values_list('laboratory', flat=True)

        all_laboratories = laboratory.objects.filter(is_available=1).exclude(laboratory_id__in=assigned_laboratories)
        all_roles = laboratory_roles.objects.filter()
        context = {
            'user': user1,
            'lab_users': lab_users,
            'all_laboratories': all_laboratories,
            'all_roles': all_roles,
        }
        return render(request, 'superuser/superuser_userinfo.html', context)
    except Exception as e:
        logger.error(f"Unexpected error in adding user: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('superuser_manage_users')


@transaction.atomic
@login_required()
@user_passes_test(lambda u: u.is_superuser)
def get_roles(request, laboratory_id):
    try:
        roles = laboratory_roles.objects.filter(Q(laboratory_id=laboratory_id) | Q(laboratory_id=0)).values('roles_id', 'name')
        return JsonResponse({'roles': list(roles)})
    except Exception as e:
        logger.error(f"Error fetching roles for lab {laboratory_id}: {e}")
        return JsonResponse({'error': 'Failed to fetch roles'}, status=500)

@transaction.atomic
@login_required()
@user_passes_test(lambda u: u.is_superuser)
def remove_lab_user(request, lab_user_id):
    try:
        lab_user = get_object_or_404(laboratory_users, id=lab_user_id)

        lab_user.is_active = False
        lab_user.save()

        messages.success(request, 'User has been successfully removed from the laboratory.')
    except Exception as e:
        logger.error(f"Error removing user {lab_user_id}: {e}")
        messages.error(request, "Failed to remove user.")
    return redirect('superuser_manage_users')  # Redirect to an appropriate page

@transaction.atomic
@login_required()
@user_passes_test(lambda u: u.is_superuser)
def edit_user(request, user_id):
    try:
        user1 = get_object_or_404(user, user_id=user_id)
        if request.method == 'POST':
            user1.firstname = request.POST['firstname']
            user1.lastname = request.POST['lastname']
            user1.username = request.POST['username']
            user1.email = request.POST['email']
            user1.personal_id = request.POST['personal_id']
            user1.save()
            messages.success(request, 'User details updated successfully.')
    except ValidationError as e:
        logger.error(f"Validation error updating user {user_id}: {e}")
        messages.error(request, f"Error updating user: {e}")

    except Exception as e:
        logger.error(f"Unexpected error updating user {user_id}: {e}")
        messages.error(request, "Failed to update user.")
    return redirect('superuser_user_info', user_id=user_id)
    
@transaction.atomic
@login_required()
@user_passes_test(lambda u: u.is_superuser)
def deactivate_user(request, user_id):
    try:
        user1 = get_object_or_404(user, user_id=user_id)
        user1.is_deactivated = True
        user1.save()
        messages.success(request, 'User deactivated successfully.')
    except Exception as e:
        logger.error(f"Error deactivating user {user_id}: {e}")
        messages.error(request, "Failed to deactivate user.")
    return redirect('superuser_user_info', user_id=user_id)

@login_required()
@user_passes_test(lambda u: u.is_superuser)
def assign_lab(request, user_id):
    if request.method == 'POST':
        try:
            laboratory_id = request.POST['laboratory_id']
            role_id = request.POST['role_id']

            if not laboratory_id or not role_id:
                messages.error(request, "Invalid data provided.")
                return redirect('superuser_user_info', user_id=user_id)

            user_laboratory = get_object_or_404(laboratory, laboratory_id=laboratory_id)
            role = get_object_or_404(laboratory_roles, roles_id=role_id)
            
            # # Check if user is already assigned to this lab
            # if laboratory_users.objects.filter(user_id=user_id, laboratory=user_laboratory, ).exists():
            #     messages.warning(request, "User is already assigned to this laboratory.")
            #     return redirect('superuser_user_info', user_id=user_id)

            laboratory_users.objects.create(user_id=user_id, laboratory=user_laboratory, role=role)
            messages.success(request, 'Laboratory assigned successfully.')
            logger.info(f"User {user_id} assigned to lab {laboratory_id} with role {role_id}.")

        except IntegrityError:
            logger.error(f"Integrity error while assigning user {user_id} to lab {laboratory_id}.")
            messages.error(request, "Assignment failed due to a database error.")

        except Exception as e:
            logger.error(f"Unexpected error assigning user {user_id} to lab {laboratory_id}: {e}")
            messages.error(request, "Failed to assign user to laboratory.")

    return redirect('superuser_user_info', user_id=user_id)


# Function to handle adding users

@user_passes_test(lambda u: u.is_superuser)
def add_room(request, laboratory_id):
    if request.method == "POST":
        try:
            room_name = request.POST.get('room_name')
            room_description = request.POST.get('room_description')
            room_capacity = request.POST.get('room_capacity')

            # Validate input data
            if room_name and room_description and room_capacity.isdigit():  # Ensure capacity is an integer
                room_capacity = int(room_capacity)  # Convert to integer
                
                # Create a new room instance
                room = rooms.objects.create(
                    laboratory_id=laboratory_id,  # Associate room with laboratory
                    name=room_name,
                    capacity=room_capacity,
                    description=room_description
                )
                messages.success(request, "Room added successfully.")
            else:
                messages.error(request, "All fields are required and capacity must be a number.")
        except Exception as e:
            logger.error(f"Error adding room to lab {laboratory_id}: {e}")
            messages.error(request, "Failed to add room.")

        return redirect('superuser_lab_info', laboratory_id=laboratory_id)
 
    # Handle GET request if necessary (optional)
    return render(request, 'superuser/superuser_labInfo.html')  # Adjust to your actual template name

#def setup_editLab(request, laboratory_id):
    # Retrieve the lab to edit
    #lab = get_object_or_404(laboratory, laboratory_id=laboratory_id)
    #return render(request, 'superuser/superuser_editLab.html', {'lab': lab})

@transaction.atomic
@user_passes_test(lambda u: u.is_superuser)
def setup_createlab(request):
    if not request.user.is_superuser:
        return render(request, 'error_page.html', {'message': 'Access restricted.'})
    
    if request.method == 'POST':
        try:
            lab_name = request.POST.get('labname')
            description = request.POST.get('description')
            department = request.POST.get('department')

            # Create the new lab
            new_lab = laboratory.objects.create(
                name=lab_name,
                description=description,
                department=department,
                is_available=True,
                date_created=timezone.localtime()
            )

            # Convert selected module IDs to integers
            selected_modules = [int(module_id) for module_id in request.POST.getlist('modules[]')]
            new_lab.modules = selected_modules
            new_lab.save()
            
            # Check if module ID 2 is selected and create a borrowing config
            if 2 in selected_modules:
                borrowing_config.objects.create(
                    laboratory=new_lab,
                    allow_walkin=False,  # Set default values or customize as needed
                    allow_prebook=False,
                    prebook_lead_time=1,
                    allow_shortterm=False,
                    allow_longterm=False,
                    questions_config=[]
                )

            # Check if module ID 4 is selected and create a reservation config
            if 4 in selected_modules:
                reservation_config.objects.create(
                    laboratory=new_lab,
                    reservation_type='class',  # Set default values or customize as needed
                    start_time=None,
                    end_time=None,
                    require_approval=False,
                    require_payment=False,
                    approval_form=None,
                    tc_description='',
                    leadtime=0
                )


            # Initialize permissions_data dictionary
            permissions_data = {}

            # Regular expression pattern for extracting role_id and perm_codename
            pattern = r"permissions\[(\d+)\]\[(\w+)\]"

            # Extract permissions data from POST
            for key, value in request.POST.items():
                match = re.match(pattern, key)
                if match:
                    role_id, perm_codename = match.groups()
                    role_id = int(role_id)
                    print(f"Parsed role_id: {role_id}, perm_codename: {perm_codename}")  # Debugging print if match found

                    # Organize data into permissions_data dictionary
                    if role_id not in permissions_data:
                        permissions_data[role_id] = {}
                    permissions_data[role_id][perm_codename] = value

            # Save permissions per role per module
            for role_id, perms in permissions_data.items():
                role = laboratory_roles.objects.get(roles_id=role_id)
                for perm_codename, is_selected in perms.items():
                    perm_obj = permissions.objects.get(codename=perm_codename)
                    
                    if is_selected == 'on':  # Save if checkbox was checked
                        laboratory_permissions.objects.update_or_create(
                            role=role, laboratory=new_lab, permissions=perm_obj
                        )
                    else:
                        # Remove the permission if unchecked
                        laboratory_permissions.objects.filter(
                            role=role, laboratory=new_lab, permissions=perm_obj
                        ).delete()

            messages.success(request, "Laboratory and permissions saved successfully.")
            logger.info(f"Laboratory '{lab_name}' created successfully.")
            return redirect('superuser_lab_info', new_lab.laboratory_id)
        except Exception as e:
            logger.error(f"Error creating lab: {e}")
            messages.error(request, "Failed to create laboratory.")

    # Prepare data for rendering
    roles = laboratory_roles.objects.filter(laboratory=0)
    modules = Module.objects.all()

    # Default permissions per role
    default_permissions = {
        1: [7, 8, 9, 10, 11, 1, 2, 3, 4, 5, 6, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21],
        2: [9, 11, 1, 5, 6, 13, 16, 17, 18, 19, 20, 21],
        3: [8, 10, 1, 2, 3, 4, 13, 15],
        4: [1, 18],
        5: [7, 12, 14]
    }

    # Group permissions by module
    permissions_by_module = {}
    for module in modules:
        permissions_by_module[module.id] = permissions.objects.filter(module=module)

    # Flatten role_permissions for easier access in template
    role_permissions = {(rp.role.roles_id, rp.permissions.codename): True for rp in laboratory_permissions.objects.all()}

    # Add default permissions to the context
    context = {
        'modules': modules,
        'roles': roles,
        'permissions_by_module': permissions_by_module,
        'role_permissions': role_permissions,
        'default_permissions': default_permissions
    }
    return render(request, 'superuser/superuser_createlab.html', context)

@user_passes_test(lambda u: u.is_superuser)
def add_role(request):
    if request.method == 'POST':
        try:
            role_name = request.POST.get('roleName')
            lab_id = request.POST.get('laboratory')  # Ensure lab_id is provided or set up

            if not role_name or not lab_id:
                messages.error(request, "Role name and laboratory must be provided.")
                return redirect('add_role')
            
            # Capture permissions for each module from the form
            permissions = {
                "inventory": {
                    "view_inventory": request.POST.get("inventory_view_inventory") == "on",
                    "add_new_item": request.POST.get("inventory_add_new_item") == "on",
                    "update_item_inventory": request.POST.get("inventory_update_item_inventory") == "on",
                    "physical_count": request.POST.get("inventory_physical_count") == "on",
                    "manage_suppliers": request.POST.get("inventory_manage_suppliers") == "on",
                    "configure_inventory": request.POST.get("inventory_configure_inventory") == "on",
                },
                "borrowing": {
                    "borrow_items": request.POST.get("borrowing_borrow_items") == "on",
                    "view_borrowed_items": request.POST.get("borrowing_view_borrowed_items") == "on",
                    "view_booking_requests": request.POST.get("borrowing_view_booking_requests") == "on",
                    "return_item": request.POST.get("borrowing_return_item") == "on",
                    "configure_borrowing": request.POST.get("borrowing_configure_borrowing") == "on",
                },
                "clearance": {
                    "view_own_clearance": request.POST.get("clearance_view_own_clearance") == "on",
                    "view_student_clearance": request.POST.get("clearance_view_student_clearance") == "on",
                },
                "lab_reservation": {
                    "reserve_laboratory": request.POST.get("reservation_reserve_laboratory") == "on",
                    "view_reservations": request.POST.get("reservation_view_reservations") == "on",
                    "approve_deny_reservations": request.POST.get("reservation_approve_deny_reservations") == "on",
                    "configure_lab_reservation": request.POST.get("reservation_configure_lab_reservation") == "on",
                },
                "reports": {
                    "view_reports": request.POST.get("reports_view_reports") == "on",
                },
            }

            # Create the role and save permissions
            new_role = laboratory_roles.objects.create(
                laboratory_id=lab_id,
                name=role_name,
                permissions=permissions
            )

            messages.success(request, "Role created successfully.")
            logger.info(f"New role '{role_name}' created in lab {lab_id}.")
            return redirect('role_list')  # Redirect to role list or another page
        except IntegrityError:
            messages.error(request, "Role already exists.")
            logger.warning(f"Duplicate role creation attempted: {role_name}")

        except Exception as e:
            logger.error(f"Error creating role: {e}")
            messages.error(request, "Failed to create role.")
    return render(request, 'core/add_role.html')

@user_passes_test(lambda u: u.is_superuser)
def superuser_login(request):
    if request.method == 'POST':
        try:
            form = LoginForm(request, data=request.POST)
            if form.is_valid():
                user = form.get_user()
                auth_login(request, user)
                return redirect('superuser_setup')
        except Exception as e:
            logger.error(f"Error during superuser login: {e}")
            messages.error(request, "Login failed.")
    else:
        form = LoginForm()
    return render(request, 'superuser/superuser_login.html', {'form': form})

@user_passes_test(lambda u: u.is_superuser)
def setup_edituser(request):
    return render(request, 'superuser/superuser_edituser.html')

@user_passes_test(lambda u: u.is_superuser)
def setup_manageRooms(request):
       return render(request, 'superuser/superuser_manageRooms.html')


# no need

@login_required()
@user_passes_test(lambda u: u.is_superuser)
def superuser_setup(request):
    if not request.user.is_superuser:
        return redirect('/login')

    labs = laboratory.objects.all()
    modules = Module.objects.all()
    return render(request, 'superuser/superuser_setup.html', {'labs': labs, 'modules': modules})

@user_passes_test(lambda u: u.is_superuser)
def superuser_logout(request):
    logout(request)
    return redirect("/login")

@transaction.atomic
@login_required(login_url='/login/superuser')
@user_passes_test(lambda u: u.is_superuser)
def add_laboratory(request):
    if not request.user.is_superuser:
        return redirect('login')

    if request.method == 'POST':
        form = LaboratoryForm(request.POST)
        if form.is_valid():
            laboratory = form.save(commit=False)
            # Save the laboratory instance first
            laboratory.save()
            # Then save the selected modules
            form.save_m2m()
            return redirect('superuser_setup')
    else:
        form = LaboratoryForm()

    return render(request, 'superuser/add_laboratory.html', {'form': form})

